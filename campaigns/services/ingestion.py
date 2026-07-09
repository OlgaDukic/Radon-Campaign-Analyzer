import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook


@dataclass(frozen=True)
class ColumnMap:
    timestamp: str | None = None
    date: str | None = None
    time: str | None = None
    radon: str | None = None
    temperature: str | None = None
    humidity: str | None = None
    pressure: str | None = None


COLUMN_ALIASES = {
    "timestamp": (
        "timestamp",
        "time",
        "date time",
        "date and time",
        "datetime",
        "time stamp",
        "logged at",
        "recorded at",
        "measurement time",
        "sample time",
        "created at",
    ),
    "date": ("date", "measurement date", "sample date"),
    "time": ("time", "clock", "measurement time", "sample time"),
    "radon": (
        "radon",
        "radon bq/m3",
        "radon bqm3",
        "radon bq m3",
        "radon concentration",
        "radon value",
        "radon gas",
        "radon avg",
    ),
    "temperature": ("temperature", "temp", "temperature c", "temp c", "air temperature", "temperature deg c"),
    "humidity": ("humidity", "relative humidity", "rh", "humidity %", "hum"),
    "pressure": ("pressure", "air pressure", "atmospheric pressure", "pressure hpa", "barometric pressure"),
}


def read_uploaded_file(uploaded_file):
    debug = _file_debug(uploaded_file)
    extension = Path(debug["filename"]).suffix.lower()
    try:
        if extension == ".csv":
            sheets = [{"name": "CSV", "rows": _read_csv(uploaded_file.file)}]
        elif extension == ".xlsx":
            sheets = _read_excel(uploaded_file.file)
        elif extension == ".xls":
            sheets = _read_xls(uploaded_file.file)
        else:
            debug["skipped_reason"] = "Unsupported file type."
            return [], None, debug
    except Exception as exc:
        debug["skipped_reason"] = f"Could not read file: {exc}"
        return [], None, debug

    debug["detected_sheets"] = [sheet["name"] for sheet in sheets]
    debug["raw_rows_read"] = sum(len(sheet["rows"]) for sheet in sheets)
    if not sheets or debug["raw_rows_read"] == 0:
        debug["skipped_reason"] = "No rows found in file."
        return [], None, debug

    best_result = None
    for sheet in sheets:
        sheet_result = _parse_sheet(sheet["name"], sheet["rows"], uploaded_file)
        debug["sheets"].append(sheet_result["debug"])
        if best_result is None or len(sheet_result["measurements"]) > len(best_result["measurements"]):
            best_result = sheet_result

    if not best_result:
        debug["skipped_reason"] = "No sheets could be inspected."
        return [], None, debug

    selected_debug = best_result["debug"]
    debug.update(
        {
            "selected_sheet": selected_debug["sheet_name"],
            "detected_header_row": selected_debug["detected_header_row"],
            "detected_columns": selected_debug["detected_columns"],
            "mapped_columns": selected_debug["mapped_columns"],
            "parsed_measurement_rows": selected_debug["parsed_measurement_rows"],
            "skipped_reason": selected_debug["skipped_reason"],
        }
    )
    return best_result["measurements"], best_result["column_map"], debug


def detect_columns(headers):
    normalized_lookup = {_normalize_header(header): header for header in headers}
    detected = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        normalized_aliases = tuple(_normalize_header(alias) for alias in aliases)
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if normalized_alias in normalized_lookup:
                detected[canonical] = normalized_lookup[normalized_alias]
                break
        if canonical not in detected:
            for normalized, original in normalized_lookup.items():
                if any(alias in normalized for alias in normalized_aliases):
                    detected[canonical] = original
                    break

    if "timestamp" not in detected and not ("date" in detected and "time" in detected):
        raise ValueError("Could not detect timestamp column.")
    if "radon" not in detected:
        raise ValueError("Could not detect radon column.")

    return ColumnMap(
        timestamp=detected.get("timestamp"),
        date=detected.get("date"),
        time=detected.get("time"),
        radon=detected.get("radon"),
        temperature=detected.get("temperature"),
        humidity=detected.get("humidity"),
        pressure=detected.get("pressure"),
    )


def parse_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\xa0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_timestamp(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        parsed = parse_datetime(text)
        if parsed is None:
            parsed_date = parse_date(text)
            parsed = datetime.combine(parsed_date, time.min) if parsed_date else None
        if parsed is None:
            for fmt in (
                "%d.%m.%Y %H:%M:%S",
                "%d.%m.%Y %H:%M",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%m/%d/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M",
            ):
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
    if parsed and timezone.is_naive(parsed):
        return timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _file_debug(uploaded_file):
    return {
        "source_file_id": uploaded_file.id,
        "filename": uploaded_file.original_name or uploaded_file.file.name,
        "detected_sheets": [],
        "raw_rows_read": 0,
        "selected_sheet": None,
        "detected_header_row": None,
        "detected_columns": [],
        "mapped_columns": _empty_mapped_columns(),
        "parsed_measurement_rows": 0,
        "skipped_reason": "",
        "sheets": [],
    }


def _read_csv(file_obj):
    file_obj.seek(0)
    raw = file_obj.read()
    if isinstance(raw, str):
        text = raw
    else:
        text = raw.decode("utf-8-sig")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(io.StringIO(text), dialect))


def _read_excel(file_obj):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    return [
        {"name": worksheet.title, "rows": [list(row) for row in worksheet.iter_rows(values_only=True)]}
        for worksheet in workbook.worksheets
    ]


def _read_xls(file_obj):
    import xlrd

    file_obj.seek(0)
    workbook = xlrd.open_workbook(file_contents=file_obj.read())
    return [
        {"name": worksheet.name, "rows": [worksheet.row_values(index) for index in range(worksheet.nrows)]}
        for worksheet in workbook.sheets()
    ]


def _parse_sheet(sheet_name, table, uploaded_file):
    debug = {
        "sheet_name": sheet_name,
        "raw_rows_read": len(table),
        "detected_header_row": None,
        "detected_columns": [],
        "mapped_columns": _empty_mapped_columns(),
        "parsed_measurement_rows": 0,
        "skipped_reason": "",
    }
    if not table:
        debug["skipped_reason"] = "Sheet has no rows."
        return {"measurements": [], "column_map": None, "debug": debug}

    header = _detect_header(table)
    if not header:
        debug["skipped_reason"] = "Could not find a header row in the first 30 rows."
        return {"measurements": [], "column_map": None, "debug": debug}

    header_index, headers = header
    debug["detected_header_row"] = header_index + 1
    debug["detected_columns"] = headers
    try:
        column_map = detect_columns(headers)
    except ValueError as exc:
        debug["skipped_reason"] = str(exc)
        return {"measurements": [], "column_map": None, "debug": debug}

    debug["mapped_columns"] = _serialize_column_map(column_map)
    data_rows = [
        list(row[: len(headers)])
        for row in table[header_index + 1 :]
        if any(cell not in (None, "") for cell in row)
    ]
    measurements = []
    skipped_without_timestamp = 0
    skipped_without_radon = 0
    for offset, data_row in enumerate(data_rows, start=header_index + 2):
        row = dict(zip(headers, data_row))
        measured_at = _row_timestamp(row, column_map)
        if not measured_at:
            skipped_without_timestamp += 1
            continue
        radon = parse_decimal(row.get(column_map.radon))
        if radon is None:
            skipped_without_radon += 1
            continue
        measurements.append(
            {
                "measured_at": measured_at,
                "radon_bq_m3": radon,
                "temperature_c": parse_decimal(row.get(column_map.temperature)),
                "humidity_percent": parse_decimal(row.get(column_map.humidity)),
                "pressure_hpa": parse_decimal(row.get(column_map.pressure)),
                "source_file": uploaded_file,
                "source_file_id": uploaded_file.id,
                "source_file_name": uploaded_file.original_name or uploaded_file.file.name,
                "original_row_number": offset,
                "original_timestamp_string": _row_timestamp_string(row, column_map),
            }
        )

    debug["parsed_measurement_rows"] = len(measurements)
    if not measurements:
        debug["skipped_reason"] = (
            "No measurement rows parsed "
            f"({skipped_without_timestamp} rows without valid timestamp, "
            f"{skipped_without_radon} rows without valid radon value)."
        )
    return {"measurements": measurements, "column_map": column_map, "debug": debug}


def _detect_header(table):
    candidates = []
    for index, row in enumerate(table[:30]):
        headers = _header_values(row)
        if not any(headers):
            continue
        try:
            column_map = detect_columns(headers)
        except ValueError:
            continue
        score = _header_score(headers, column_map)
        candidates.append((score, index, headers))
    if not candidates:
        return None
    _score, index, headers = max(candidates, key=lambda candidate: (candidate[0], -candidate[1]))
    return index, headers


def _header_score(headers, column_map):
    mapped = _serialize_column_map(column_map)
    score = sum(1 for value in mapped.values() if value)
    if column_map.timestamp:
        score += 2
    if column_map.date and column_map.time:
        score += 2
    if column_map.radon:
        score += 3
    score += min(len([header for header in headers if header]), 8) / 10
    return score


def _header_values(row):
    width = len(row)
    while width > 0 and row[width - 1] in (None, ""):
        width -= 1
    return ["" if value is None else str(value).strip() for value in row[:width]]


def _row_timestamp(row, column_map):
    if column_map.timestamp:
        parsed = parse_timestamp(row.get(column_map.timestamp))
        if parsed:
            return parsed
    if column_map.date and column_map.time:
        return parse_timestamp(f"{row.get(column_map.date, '')} {row.get(column_map.time, '')}".strip())
    return None


def _row_timestamp_string(row, column_map):
    if column_map.timestamp:
        return row.get(column_map.timestamp)
    if column_map.date and column_map.time:
        return f"{row.get(column_map.date, '')} {row.get(column_map.time, '')}".strip()
    return None


def _serialize_column_map(column_map):
    return {
        "timestamp": column_map.timestamp or _combined_timestamp_label(column_map),
        "radon": column_map.radon,
        "temperature": column_map.temperature,
        "humidity": column_map.humidity,
        "pressure": column_map.pressure,
    }


def _combined_timestamp_label(column_map):
    if column_map.date and column_map.time:
        return f"{column_map.date} + {column_map.time}"
    return None


def _empty_mapped_columns():
    return {
        "timestamp": None,
        "radon": None,
        "temperature": None,
        "humidity": None,
        "pressure": None,
    }


def _normalize_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("³", "3").replace("Â³", "3").replace("°", "")
    text = text.replace("bq/m^3", "bq/m3").replace("bq/m3", "bq m3")
    text = re.sub(r"[^a-z0-9%/]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()
