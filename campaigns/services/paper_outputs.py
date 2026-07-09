import csv
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


PAPER_ONE_SHEETS = [
    "Source File Inventory",
    "Canonical Dataset Summary",
    "Canonical Hourly Data",
    "Quality Flags",
    "Quality Flag Dictionary",
    "Sampling Diagnostics",
    "Overlap Conflicts",
    "DST Diagnostics",
    "Resampling Summary",
    "Regime Sensitivity",
    "Prediction Skill by Regime",
    "Prediction Readiness",
    "SIREM Readiness",
    "Reproducibility Config",
    "Row Reconciliation Summary",
]


def build_row_reconciliation_summary(summary):
    canonical = summary.get("canonical_dataset_summary", {})
    quality = summary.get("quality_flag_counts", {})
    raw_rows = _integer(canonical.get("raw_records"))
    exact_removed = _integer(canonical.get("exact_duplicates_removed"))
    conflict_rows = _integer(canonical.get("conflicts"))
    missing_radon = _integer(quality.get("MISSING_RADON"))
    canonical_valid = _integer(canonical.get("canonical_valid_records"))
    hourly_rows = len(summary.get("canonical_hourly_data") or [])
    removed_total = max(raw_rows - exact_removed - canonical_valid, 0)
    removed_due_to_conflict = min(conflict_rows, removed_total)
    removed_due_to_missing = min(missing_radon, max(removed_total - removed_due_to_conflict, 0))
    removed_due_to_other = max(removed_total - removed_due_to_conflict - removed_due_to_missing, 0)
    return {
        "raw_imported_rows": raw_rows,
        "exact_duplicate_rows_removed": exact_removed,
        "duplicate_conflict_rows": conflict_rows,
        "rows_removed_due_to_conflict_resolution": removed_due_to_conflict,
        "rows_removed_due_to_missing_or_invalid_radon": removed_due_to_missing,
        "rows_removed_due_to_other_quality_rules": removed_due_to_other,
        "canonical_valid_rows": canonical_valid,
        "canonical_hourly_rows": hourly_rows,
        "notes": (
            "Reconciliation is computed from report JSON. Canonical valid rows exclude exact duplicate removals "
            "and rows marked invalid by canonical conflict or radon validity rules; hourly rows are aggregate intervals."
        ),
    }


def build_dst_compact_summary(summary):
    diagnostics = summary.get("dst_diagnostics") or []
    ambiguous = _flag_count(diagnostics, "DST_AMBIGUOUS")
    nonexistent = _flag_count(diagnostics, "DST_MISSING")
    parse_warnings = _flag_count(diagnostics, "TIMESTAMP_PARSE_WARNING")
    return {
        "timezone_audit_rows": len(diagnostics),
        "dst_ambiguous_count": ambiguous,
        "dst_nonexistent_count": nonexistent,
        "dst_transition_related_count": ambiguous + nonexistent,
        "dst_problem_count": ambiguous + nonexistent + parse_warnings,
        "dst_notes": (
            "Timezone audit rows document assumptions for reproducible ordering and are not all DST problems. "
            "Problem counts include ambiguous/nonexistent local timestamps and timestamp parse warnings."
        ),
    }


def build_sampling_gaps_compact_summary(summary):
    diagnostics = summary.get("sampling_diagnostics") or {}
    return {
        "total_sampling_irregularities": _integer(diagnostics.get("irregular_interval_count")),
        "short_gaps": _integer(diagnostics.get("short_gap_count")),
        "long_gaps": _integer(diagnostics.get("long_gap_count")),
        "inter_file_gaps": "N/A",
        "dst_related_gaps": "N/A",
        "prediction_breaking_gaps": _integer(diagnostics.get("long_gap_count")),
        "notes": (
            "Sampling gaps are detected with the configured tolerance multiplier. Inter-file and DST-related "
            "gap attribution is not separately classified by the current prototype."
        ),
    }


def enrich_paper_summary(summary):
    enriched = dict(summary or {})
    enriched["row_reconciliation_summary"] = build_row_reconciliation_summary(enriched)
    enriched["dst_diagnostics_compact_summary"] = build_dst_compact_summary(enriched)
    enriched["sampling_gaps_compact_summary"] = build_sampling_gaps_compact_summary(enriched)
    return enriched


def write_paper_output_package(campaign, report, output_dir, excel_path=None, command_used="", checks=None):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    root_dir = output_dir.parent
    summary = enrich_paper_summary(report.summary_json or {})
    created = []

    created.append(_write_rows(output_dir / "source_file_inventory.csv", summary.get("source_file_inventory", [])))
    created.append(_write_kv(output_dir / "canonical_dataset_summary.csv", summary.get("canonical_dataset_summary", {})))
    created.append(_write_kv(output_dir / "quality_flag_counts.csv", summary.get("quality_flag_counts", {}), "quality_flag", "count"))
    sampling_summary = {key: value for key, value in (summary.get("sampling_diagnostics") or {}).items() if key != "gaps"}
    created.append(_write_kv(output_dir / "sampling_diagnostics_summary.csv", sampling_summary))
    created.append(_write_rows(output_dir / "overlap_conflict_summary.csv", summary.get("overlap_conflicts", [])))
    created.append(_write_rows(output_dir / "dst_diagnostics_summary.csv", summary.get("dst_diagnostics", [])))
    created.append(_write_kv(output_dir / "regime_counts.csv", summary.get("regime_counts", {}), "regime", "count"))
    created.append(_write_rows(output_dir / "regime_sensitivity_summary.csv", summary.get("regime_sensitivity", [])))
    created.append(_write_rows(output_dir / "prediction_skill_by_regime.csv", summary.get("prediction_skill_by_regime", [])))
    created.append(_write_rows(output_dir / "prediction_readiness.csv", summary.get("prediction_readiness", [])))
    created.append(_write_rows(output_dir / "sirem_readiness.csv", summary.get("sirem_readiness", [])))
    created.append(_write_kv(output_dir / "reproducibility_config.csv", summary.get("reproducibility_config", {})))
    created.append(_write_kv(output_dir / "row_reconciliation_summary.csv", summary.get("row_reconciliation_summary", {})))
    created.append(_write_kv(output_dir / "dst_diagnostics_compact_summary.csv", summary.get("dst_diagnostics_compact_summary", {})))
    created.append(_write_kv(output_dir / "sampling_gaps_compact_summary.csv", summary.get("sampling_gaps_compact_summary", {})))

    validation_path = output_dir / "paper1_validation_report.md"
    validation_path.write_text(
        build_validation_report(campaign, report, summary, excel_path, command_used, checks),
        encoding="utf-8",
    )
    convenience_path = root_dir / "paper1_validation_report.md"
    if convenience_path != validation_path:
        shutil.copyfile(validation_path, convenience_path)
    return {
        "csv_paths": created,
        "validation_report": validation_path,
        "convenience_validation_report": convenience_path,
    }


def build_validation_report(campaign, report, summary, excel_path=None, command_used="", checks=None):
    excel_validation = _excel_validation(excel_path)
    reconciliation = summary.get("row_reconciliation_summary", {})
    dst_summary = summary.get("dst_diagnostics_compact_summary", {})
    gap_summary = summary.get("sampling_gaps_compact_summary", {})
    config = summary.get("analysis_config", {})
    repro = summary.get("reproducibility_config", {})
    prediction_metrics = summary.get("prediction_metrics") or {}
    models = sorted({model for results in prediction_metrics.values() for model in results})
    horizons = sorted(prediction_metrics)
    small_warnings = sum(1 for row in summary.get("prediction_skill_by_regime", []) if row.get("small_sample_warning"))
    checks = checks or {}
    lines = [
        "# Paper 1 Workflow Validation Report",
        "",
        "This validation report describes an exploratory research-software output package. It is not a certified radon risk assessment, medical, legal, or regulatory decision report.",
        "",
        "## Campaign and Run",
        f"- Campaign ID: {campaign.id}",
        f"- Campaign name: {campaign.name}",
        f"- Analysis report ID: {report.id}",
        f"- Analysis timestamp: {report.created_at.isoformat()}",
        f"- Software version / git commit: {repro.get('app_version_or_git_commit', 'N/A')}",
        f"- Command used: `{command_used or 'N/A'}`",
        f"- Timezone: {config.get('timezone_name', repro.get('timezone', 'N/A'))}",
        f"- Resampling interval: {config.get('resample_interval', repro.get('resampling_interval', 'N/A'))}",
        f"- Gap tolerance: {config.get('gap_tolerance_multiplier', repro.get('gap_tolerance_multiplier', 'N/A'))}",
        "",
        "## Row Reconciliation",
    ]
    lines.extend(f"- {key}: {value}" for key, value in reconciliation.items())
    lines.extend(
        [
            "",
            "## Compact DST Summary",
        ]
    )
    lines.extend(f"- {key}: {value}" for key, value in dst_summary.items())
    lines.extend(
        [
            "",
            "## Compact Sampling Gap Summary",
        ]
    )
    lines.extend(f"- {key}: {value}" for key, value in gap_summary.items())
    lines.extend(
        [
            "",
            "## Regimes and Prediction",
            f"- Regime labels found: {', '.join((summary.get('regime_counts') or {}).keys()) or 'N/A'}",
            f"- Prediction horizons evaluated: {', '.join(horizons) or 'N/A'}",
            f"- Models evaluated: {', '.join(models) or 'N/A'}",
            "- Prediction evaluation policy: chronological train/test split; training observations precede test observations in time.",
            f"- Small-sample warnings: {small_warnings}",
            "",
            "## Excel and Output Validation",
            f"- Excel workbook: {excel_path or 'N/A'}",
            f"- Missing Excel sheets: {', '.join(excel_validation['missing_sheets']) if excel_validation['missing_sheets'] else 'none'}",
            f"- Empty/suspicious sheets: {', '.join(excel_validation['empty_or_suspicious_sheets']) if excel_validation['empty_or_suspicious_sheets'] else 'none'}",
            "",
            "## Tests and Checks",
            f"- `python manage.py test`: {checks.get('test', 'not recorded in this report')}",
            f"- `python manage.py check`: {checks.get('check', 'not recorded in this report')}",
            f"- `python manage.py makemigrations --check --dry-run`: {checks.get('makemigrations', 'not recorded in this report')}",
            "",
            "## Limitations",
            "- Timezone audit rows document reproducibility assumptions and should not be interpreted as data-quality errors.",
            "- Inter-file and DST-related gap attribution is not separately classified by the current prototype.",
            "- Reconciliation is based on report-level counts and is intended as a paper audit trail, not a regulatory data acceptance statement.",
            "- Outputs remain exploratory research diagnostics, not certified radon risk-assessment results.",
            "",
        ]
    )
    return "\n".join(lines)


def _excel_validation(excel_path):
    if not excel_path or not Path(excel_path).exists():
        return {"missing_sheets": PAPER_ONE_SHEETS, "empty_or_suspicious_sheets": ["Excel workbook not found"]}
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        missing = [sheet for sheet in PAPER_ONE_SHEETS if sheet not in workbook.sheetnames]
        suspicious = []
        for sheet_name in PAPER_ONE_SHEETS:
            if sheet_name in workbook.sheetnames and workbook[sheet_name].max_row <= 1:
                suspicious.append(f"{sheet_name}: header only")
        return {"missing_sheets": missing, "empty_or_suspicious_sheets": suspicious}
    finally:
        workbook.close()


def _write_rows(path, rows, headers=None):
    rows = list(rows or [])
    headers = headers or (_headers_for(rows) if rows else ["note"])
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        if rows:
            for row in rows:
                writer.writerow({header: _clean(row.get(header)) for header in headers})
        else:
            writer.writerow({headers[0]: "N/A"})
    return path


def _write_kv(path, mapping, key_name="field", value_name="value"):
    return _write_rows(
        path,
        [{key_name: key, value_name: value} for key, value in (mapping or {}).items()],
        [key_name, value_name],
    )


def _headers_for(rows):
    headers = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return headers


def _clean(value):
    if value is None:
        return "N/A"
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _integer(value):
    if value in (None, ""):
        return 0
    return int(value)


def _flag_count(diagnostics, flag):
    return sum(1 for row in diagnostics if flag in (row.get("flags") or []))
