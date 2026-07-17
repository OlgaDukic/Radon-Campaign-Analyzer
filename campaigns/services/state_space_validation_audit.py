import csv
import json
from collections import defaultdict
from datetime import timedelta
from io import BytesIO, StringIO
from math import sqrt
from statistics import median

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook

from . import reduced_state_space_experiment as ss
from .baseline_prediction_experiment import run_baseline_prediction_experiment
from .documented_events import PHASES, default_event_cycles_for_campaign


VERSION = "state_space_validation_audit_v1"
NOMINAL_COVERAGE = 0.95


def run_state_space_validation_audit(campaign):
    state_payload = ss.run_reduced_state_space_experiment(campaign)
    definitions = default_event_cycles_for_campaign(campaign)
    measurements = ss._measurement_rows(campaign)
    interval = ss._median_interval(measurements) or 1.0
    evaluations = []
    traces = []
    local_q_audit = []
    if len(definitions) >= 2:
        experiment_specs = [
            ("Experiment A", definitions[0], definitions[1]),
            ("Experiment B", definitions[1], definitions[0]),
        ]
        experiment_lookup = {experiment["label"]: experiment for experiment in state_payload["experiments"]}
        for label, train_cycle, test_cycle in experiment_specs:
            base_experiment = experiment_lookup[label]
            selected_q = base_experiment["selected_q"]
            reduced_parameters = base_experiment["regime_parameters"]
            test_rows = ss._rows_in_cycle(measurements, test_cycle)
            train_rows = ss._rows_in_cycle(measurements, train_cycle)
            local_parameters = _local_level_parameters()
            local_audit, local_q = ss._select_q(label, train_cycle, train_rows, local_parameters, interval)
            local_q_audit.extend([{**row, "comparator": "generic_local_level_kalman"} for row in local_audit])
            for mode in ss.FORECAST_MODES:
                rows, trace = _evaluate(
                    label,
                    train_cycle,
                    test_cycle,
                    test_rows,
                    reduced_parameters,
                    selected_q,
                    ss.PRIMARY_R_SCENARIO,
                    ss.R_SCENARIOS[ss.PRIMARY_R_SCENARIO]["observation_variance"],
                    mode,
                    interval,
                    model_label="open_loop_reduced_transition",
                    measurement_update=False,
                )
                evaluations.extend(rows)
                traces.extend(trace)
                rows, trace = _evaluate(
                    label,
                    train_cycle,
                    test_cycle,
                    test_rows,
                    reduced_parameters,
                    selected_q,
                    ss.PRIMARY_R_SCENARIO,
                    ss.R_SCENARIOS[ss.PRIMARY_R_SCENARIO]["observation_variance"],
                    mode,
                    interval,
                    model_label="reduced_sirem_informed_kalman",
                    measurement_update=True,
                )
                evaluations.extend(rows)
                traces.extend(trace)
                rows, trace = _evaluate(
                    label,
                    train_cycle,
                    test_cycle,
                    test_rows,
                    local_parameters,
                    local_q,
                    ss.PRIMARY_R_SCENARIO,
                    ss.R_SCENARIOS[ss.PRIMARY_R_SCENARIO]["observation_variance"],
                    mode,
                    interval,
                    model_label="generic_local_level_kalman",
                    measurement_update=True,
                )
                evaluations.extend(rows)
                traces.extend(trace)
            for q in ss.Q_GRID:
                for mode in ss.FORECAST_MODES:
                    rows, trace = _evaluate(
                        label,
                        train_cycle,
                        test_cycle,
                        test_rows,
                        reduced_parameters,
                        q,
                        ss.PRIMARY_R_SCENARIO,
                        ss.R_SCENARIOS[ss.PRIMARY_R_SCENARIO]["observation_variance"],
                        mode,
                        interval,
                        model_label="q_sensitivity_reduced_kalman",
                        measurement_update=True,
                    )
                    evaluations.extend(rows)
                    traces.extend(trace)
            for r_name, r_config in ss.R_SCENARIOS.items():
                for mode in ss.FORECAST_MODES:
                    rows, trace = _evaluate(
                        label,
                        train_cycle,
                        test_cycle,
                        test_rows,
                        reduced_parameters,
                        selected_q,
                        r_name,
                        r_config["observation_variance"],
                        mode,
                        interval,
                        model_label="r_sensitivity_reduced_kalman",
                        measurement_update=True,
                    )
                    evaluations.extend(rows)
                    traces.extend(trace)
    payload = {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "source_state_space_algorithm_version": state_payload["algorithm_version"],
        "manifest": _manifest(campaign, state_payload),
        "ablation_performance": _ablation_performance(evaluations),
        "q_sensitivity": _q_sensitivity(evaluations, state_payload),
        "r_sensitivity": _r_sensitivity(evaluations, traces),
        "uncertainty_calibration": _uncertainty_calibration(evaluations),
        "event_knowledge_audit": _event_knowledge_audit(evaluations, definitions),
        "rapid_parameter_stability": _rapid_parameter_stability(measurements, definitions, interval),
        "innovation_diagnostics": _innovation_diagnostics(traces),
        "fair_comparison": _fair_comparison(campaign, evaluations),
        "scientific_decision": _scientific_decision(evaluations),
        "local_level_q_selection_audit": local_q_audit,
    }
    return payload


def build_state_space_validation_workbook(payload):
    workbook = Workbook()
    workbook.active.title = "Manifest"
    _write_rows(workbook["Manifest"], [{"field": key, "value": value} for key, value in payload["manifest"].items()])
    for sheet_name, key in [
        ("Ablation Performance", "ablation_performance"),
        ("Q Sensitivity", "q_sensitivity"),
        ("R Sensitivity", "r_sensitivity"),
        ("Uncertainty Calibration", "uncertainty_calibration"),
        ("Event Knowledge Audit", "event_knowledge_audit"),
        ("Rapid Parameter Stability", "rapid_parameter_stability"),
        ("Innovation Diagnostics", "innovation_diagnostics"),
        ("Fair Comparison", "fair_comparison"),
        ("Scientific Decision", "scientific_decision"),
        ("Local Q Audit", "local_level_q_selection_audit"),
    ]:
        _write_rows(workbook.create_sheet(sheet_name), payload[key])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 62)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_state_space_validation_outputs(campaign, output_dir=None):
    payload = run_state_space_validation_audit(campaign)
    base = output_dir or settings.BASE_DIR / "paper_outputs" / f"campaign_{campaign.pk}" / "state_space_validation_audit"
    base.mkdir(parents=True, exist_ok=True)
    (base / "state_space_validation_audit.json").write_text(json.dumps(payload, indent=2, ensure_ascii=True, default=str), encoding="utf-8")
    (base / "state_space_validation_audit.xlsx").write_bytes(build_state_space_validation_workbook(payload).getvalue())
    for filename, key in [
        ("ablation_performance.csv", "ablation_performance"),
        ("q_sensitivity.csv", "q_sensitivity"),
        ("r_sensitivity.csv", "r_sensitivity"),
        ("uncertainty_calibration.csv", "uncertainty_calibration"),
        ("event_knowledge_audit.csv", "event_knowledge_audit"),
        ("rapid_parameter_stability.csv", "rapid_parameter_stability"),
        ("innovation_diagnostics.csv", "innovation_diagnostics"),
        ("fair_comparison.csv", "fair_comparison"),
    ]:
        (base / filename).write_text(_csv(payload[key]), encoding="utf-8")
    (base / "state_space_validation_summary.md").write_text(_summary_markdown(payload), encoding="utf-8")
    return payload, base


def _evaluate(label, train_cycle, test_cycle, rows, parameters, q, r_name, r_value, mode, interval, model_label, measurement_update):
    if not rows:
        return [], []
    state = float(rows[0]["radon_bq_m3"])
    variance = r_value
    last_time = rows[0]["measured_at"]
    by_time = {row["measured_at"]: row for row in rows}
    forecasts = []
    trace = []
    signed_history = []
    for index, row in enumerate(rows):
        current_time = row["measured_at"]
        if index == 0:
            prior, prior_var = state, variance
        else:
            phase_for_transition = ss._phase_for_time(test_cycle, last_time)
            regime = ss._regime_for_phase(phase_for_transition)
            dt = ss._hours(last_time, current_time)
            prior, prior_var = ss._propagate(state, variance, parameters[regime], q, dt, interval)
        observed = float(row["radon_bq_m3"])
        kalman_gain = prior_var / (prior_var + r_value) if prior_var + r_value else 0.0
        if measurement_update:
            posterior, posterior_var = ss.kalman_update(prior, prior_var, observed, r_value)
        else:
            posterior, posterior_var = prior, prior_var
        innovation = observed - prior
        signed_history.append(1 if innovation > 0 else -1 if innovation < 0 else 0)
        current_phase = ss._phase_for_time(test_cycle, current_time)
        trace.append(
            {
                "experiment": label,
                "model": model_label,
                "forecast_mode": mode,
                "R_scenario": r_name,
                "Q": q,
                "timestamp": ss._iso(current_time),
                "phase": current_phase,
                "prior_variance": ss._round(prior_var),
                "innovation": ss._round(innovation),
                "innovation_variance": ss._round(prior_var + r_value),
                "normalized_innovation_squared": ss._round((innovation * innovation) / (prior_var + r_value)) if prior_var + r_value else None,
                "kalman_gain": ss._round(kalman_gain),
                "measurement_update_applied": measurement_update,
            }
        )
        for horizon_label, horizon_hours in ss.HORIZONS.items():
            target_time = current_time + timedelta(hours=horizon_hours)
            target = by_time.get(target_time)
            target_phase = ss._phase_for_time(test_cycle, target_time)
            valid = target is not None and target_phase is not None
            mean, latent_var = ss._forecast_from_origin(posterior, posterior_var, current_time, target_time, test_cycle, parameters, q, mode, interval)
            obs_var = latent_var + r_value
            actual = float(target["radon_bq_m3"]) if target else None
            latent_lower = mean - 1.96 * sqrt(max(latent_var, 0.0))
            latent_upper = mean + 1.96 * sqrt(max(latent_var, 0.0))
            obs_lower = mean - 1.96 * sqrt(max(obs_var, 0.0))
            obs_upper = mean + 1.96 * sqrt(max(obs_var, 0.0))
            forecasts.append(
                {
                    "experiment": label,
                    "training_cycle": train_cycle.cycle_label,
                    "test_cycle": test_cycle.cycle_label,
                    "model": model_label,
                    "measurement_update_applied": measurement_update,
                    "forecast_mode": mode,
                    "R_scenario": r_name,
                    "Q": q,
                    "horizon": horizon_label,
                    "forecast_origin": ss._iso(current_time),
                    "target_timestamp": ss._iso(target_time),
                    "observed_value": ss._round(actual) if actual is not None else None,
                    "forecast_mean": ss._round(mean),
                    "latent_state_variance": ss._round(latent_var),
                    "future_observation_variance": ss._round(obs_var),
                    "latent_lower": ss._round(latent_lower),
                    "latent_upper": ss._round(latent_upper),
                    "observation_lower": ss._round(obs_lower),
                    "observation_upper": ss._round(obs_upper),
                    "latent_covered": latent_lower <= actual <= latent_upper if actual is not None else None,
                    "observation_covered": obs_lower <= actual <= obs_upper if actual is not None else None,
                    "origin_phase": current_phase,
                    "target_phase": target_phase,
                    "event_knowledge_status": ss._event_knowledge_status(mode, current_time, target_time, test_cycle),
                    "valid_status": "valid" if valid else "excluded",
                    "exclusion_reason": None if valid else "missing_target_or_outside_documented_cycle",
                    "error": ss._round(mean - actual) if actual is not None else None,
                    "absolute_error": ss._round(abs(mean - actual)) if actual is not None else None,
                }
            )
        state, variance = posterior, posterior_var
        last_time = current_time
    return forecasts, trace


def _local_level_parameters():
    payload = ss._parameter_payload(1.0, 0.0, 999, 1.0)
    payload["flags"] = "GENERIC_LOCAL_LEVEL_COMPARATOR"
    return {"closed_reference": payload.copy(), "rapid_removal": payload.copy()}


def _ablation_performance(rows):
    selected = [
        row for row in rows
        if row["valid_status"] == "valid"
        and row["model"] in {"open_loop_reduced_transition", "reduced_sirem_informed_kalman", "generic_local_level_kalman"}
        and row["R_scenario"] == ss.PRIMARY_R_SCENARIO
    ]
    return _metric_rows(selected, ["experiment", "model", "forecast_mode", "horizon"])


def _q_sensitivity(rows, state_payload):
    selected_q = {experiment["label"]: experiment["selected_q"] for experiment in state_payload["experiments"]}
    selected = [row for row in rows if row["valid_status"] == "valid" and row["model"] == "q_sensitivity_reduced_kalman"]
    out = _metric_rows(selected, ["experiment", "forecast_mode", "Q", "horizon"])
    for row in out:
        row["selected_q"] = row["Q"] == selected_q.get(row["experiment"])
        innovation_rows = [trace for trace in rows if False]
        row["innovation_note"] = "Innovation statistics are reported in innovation_diagnostics.csv."
    return out


def _r_sensitivity(rows, traces):
    selected = [row for row in rows if row["valid_status"] == "valid" and row["model"] == "r_sensitivity_reduced_kalman"]
    out = _metric_rows(selected, ["experiment", "forecast_mode", "R_scenario", "horizon"])
    gain_groups = defaultdict(list)
    for row in traces:
        if row["model"] == "r_sensitivity_reduced_kalman":
            gain_groups[(row["experiment"], row["forecast_mode"], row["R_scenario"])].append(row["kalman_gain"])
    for row in out:
        gains = gain_groups.get((row["experiment"], row["forecast_mode"], row["R_scenario"]), [])
        row["mean_kalman_gain"] = ss._round(sum(gains) / len(gains)) if gains else None
        row["interpretation"] = "Higher gain gives more weight to the sensor update; lower gain gives more weight to transition propagation."
    return out


def _uncertainty_calibration(rows):
    valid = [row for row in rows if row["valid_status"] == "valid" and row["model"] in {"reduced_sirem_informed_kalman", "r_sensitivity_reduced_kalman"}]
    grouped = defaultdict(list)
    for row in valid:
        grouped[(row["experiment"], row["model"], row["forecast_mode"], row["R_scenario"], row["horizon"])].append(row)
    out = []
    for key, values in sorted(grouped.items()):
        latent = _coverage(values, "latent")
        observation = _coverage(values, "observation")
        out.append(
            {
                **dict(zip(["experiment", "model", "forecast_mode", "R_scenario", "horizon"], key)),
                "nominal_coverage": NOMINAL_COVERAGE,
                "latent_state_empirical_coverage": latent["coverage"],
                "latent_state_coverage_error": ss._round(latent["coverage"] - NOMINAL_COVERAGE) if latent["coverage"] is not None else None,
                "latent_state_mean_interval_width": latent["width"],
                "future_observation_empirical_coverage": observation["coverage"],
                "future_observation_coverage_error": ss._round(observation["coverage"] - NOMINAL_COVERAGE) if observation["coverage"] is not None else None,
                "future_observation_mean_interval_width": observation["width"],
                "future_observation_interval_score": observation["interval_score"],
                "N": len(values),
                "semantic_note": "Future-observation interval includes observation noise R; latent-state interval does not.",
            }
        )
    return out


def _coverage(rows, interval_kind):
    if interval_kind == "latent":
        covered_key, lower_key, upper_key = "latent_covered", "latent_lower", "latent_upper"
    else:
        covered_key, lower_key, upper_key = "observation_covered", "observation_lower", "observation_upper"
    covered = [bool(row[covered_key]) for row in rows]
    widths = [row[upper_key] - row[lower_key] for row in rows]
    scores = [_interval_score(row["observed_value"], row[lower_key], row[upper_key]) for row in rows]
    return {
        "coverage": ss._round(sum(covered) / len(covered)) if covered else None,
        "width": ss._round(sum(widths) / len(widths)) if widths else None,
        "interval_score": ss._round(sum(scores) / len(scores)) if scores else None,
    }


def _interval_score(y, lower, upper):
    alpha = 0.05
    width = upper - lower
    if y < lower:
        return width + (2 / alpha) * (lower - y)
    if y > upper:
        return width + (2 / alpha) * (y - upper)
    return width


def _event_knowledge_audit(rows, definitions):
    out = []
    for row in rows:
        if row["model"] != "reduced_sirem_informed_kalman" or row["forecast_mode"] != "F1_no_future_event_knowledge":
            continue
        if row["event_knowledge_status"] != "no_future_event_knowledge_closed_transition_used":
            continue
        out.append(
            {
                "experiment": row["experiment"],
                "forecast_origin": row["forecast_origin"],
                "target_timestamp": row["target_timestamp"],
                "target_phase": row["target_phase"],
                "future_phase_label_used_for_transition": False,
                "rapid_removal_transition_used_before_event": False,
                "future_boundary_used_for_evaluation_label_only": True,
                "phase_known_at_origin": row["origin_phase"],
                "future_phase_unknown_in_F1": True,
                "note": "F1 uses closed/reference propagation across a future rapid-removal boundary until observations arrive.",
            }
        )
    return out[:200]


def _rapid_parameter_stability(measurements, definitions, interval):
    rows = []
    if len(definitions) < 2:
        return rows
    for label, cycle in [("Experiment A", definitions[0]), ("Experiment B", definitions[1])]:
        for shift in [-2, -1, 0, 1, 2]:
            start = cycle.rapid_removal_start + timedelta(hours=shift)
            end = cycle.rapid_removal_end + timedelta(hours=shift)
            window_rows = [row for row in measurements if start <= row["measured_at"] <= end]
            pairs = [
                (float(prev["radon_bq_m3"]), float(curr["radon_bq_m3"]), ss._hours(prev["measured_at"], curr["measured_at"]))
                for prev, curr in zip(window_rows, window_rows[1:])
            ]
            fit = ss._fit_ab_for_pairs(pairs, interval)
            residuals = [pair[1] - (fit["a"] * pair[0] + fit["b"]) for pair in pairs]
            flags = []
            if fit["sample_count"] < ss.SMALL_PHASE_N:
                flags.append("SMALL_PHASE_SAMPLE")
            if fit["a"] <= 0.01 or fit["a"] >= 0.99:
                flags.append("PARAMETER_NEAR_BOUNDARY")
            rows.append(
                {
                    "experiment": label,
                    "cycle_label": cycle.cycle_label,
                    "boundary_shift_h": shift,
                    "window_start": ss._iso(start),
                    "window_end": ss._iso(end),
                    "N": len(window_rows),
                    "a": fit["a"],
                    "b": fit["b"],
                    "kappa_h_minus_1": fit["kappa_h_minus_1"],
                    "C_eq_bq_m3": fit["C_eq_bq_m3"],
                    "residual_RMSE": ss._round(sqrt(sum(item * item for item in residuals) / len(residuals))) if residuals else None,
                    "flags": "|".join(flags or ["FIT_OK"]),
                    "interpretation": "Boundary sensitivity only; production parameters are unchanged.",
                }
            )
    return rows


def _innovation_diagnostics(traces):
    grouped = defaultdict(list)
    for row in traces:
        if row["model"] in {"reduced_sirem_informed_kalman", "r_sensitivity_reduced_kalman"}:
            grouped[(row["experiment"], row["model"], row["forecast_mode"], row["R_scenario"], row["phase"])].append(row)
    out = []
    for key, values in sorted(grouped.items()):
        innovations = [row["innovation"] for row in values if row["innovation"] is not None]
        nis = [row["normalized_innovation_squared"] for row in values if row["normalized_innovation_squared"] is not None]
        out.append(
            {
                **dict(zip(["experiment", "model", "forecast_mode", "R_scenario", "phase"], key)),
                "N": len(innovations),
                "mean_innovation": ss._round(sum(innovations) / len(innovations)) if innovations else None,
                "innovation_variance": _variance(innovations),
                "mean_normalized_innovation_squared": ss._round(sum(nis) / len(nis)) if nis else None,
                "lag1_autocorrelation": _lag1(innovations),
                "sustained_signed_innovation_runs": _signed_runs(innovations),
                "mean_kalman_gain": ss._round(sum(row["kalman_gain"] for row in values) / len(values)) if values else None,
            }
        )
    return out


def _fair_comparison(campaign, rows):
    baseline = run_baseline_prediction_experiment(campaign)
    baseline_rows = [row for exp in baseline["experiments"] for row in exp["forecast_rows"]]
    audit_rows = [
        row for row in rows
        if row["valid_status"] == "valid"
        and row["R_scenario"] == ss.PRIMARY_R_SCENARIO
        and row["model"] in {"open_loop_reduced_transition", "generic_local_level_kalman", "reduced_sirem_informed_kalman"}
    ]
    output = []
    for experiment in sorted({row["experiment"] for row in audit_rows}):
        for mode in ss.FORECAST_MODES:
            for horizon in ss.HORIZONS:
                state_by_model = {
                    model: {
                        (row["forecast_origin"], row["target_timestamp"]): row
                        for row in audit_rows
                        if row["experiment"] == experiment and row["forecast_mode"] == mode and row["horizon"] == horizon and row["model"] == model
                    }
                    for model in ["open_loop_reduced_transition", "generic_local_level_kalman", "reduced_sirem_informed_kalman"]
                }
                baseline_by_model = {
                    model: {
                        (row["forecast_origin"], row["target_timestamp"]): row
                        for row in baseline_rows
                        if row["experiment"] == experiment and row["horizon"] == horizon and row["model"] == model
                    }
                    for model in ss.BASELINE_MODELS
                }
                common = set.intersection(*(set(mapping) for mapping in [*state_by_model.values(), *baseline_by_model.values()] if mapping))
                for model, mapping in state_by_model.items():
                    output.append({"experiment": experiment, "forecast_mode": mode, "horizon": horizon, "model": model, "common_target_count": len(common), **_simple_metrics([mapping[key]["observed_value"] for key in common], [mapping[key]["forecast_mean"] for key in common])})
                for model, mapping in baseline_by_model.items():
                    output.append({"experiment": experiment, "forecast_mode": mode, "horizon": horizon, "model": model, "common_target_count": len(common), **_simple_metrics([mapping[key]["actual"] for key in common], [mapping[key]["predicted"] for key in common])})
    return output


def _scientific_decision(rows):
    decisions = []
    reduced = _mean_mae(rows, "reduced_sirem_informed_kalman")
    open_loop = _mean_mae(rows, "open_loop_reduced_transition")
    local = _mean_mae(rows, "generic_local_level_kalman")
    decisions.append({"question": "Does sequential updating improve open-loop transition?", "answer": _yes_no(reduced, open_loop), "evidence": f"Mean MAE reduced sequential={reduced}; open-loop={open_loop}"})
    decisions.append({"question": "Does S.I.R.E.M.-informed transition improve generic local-level Kalman?", "answer": _yes_no(reduced, local), "evidence": f"Mean MAE reduced sequential={reduced}; local-level={local}"})
    decisions.append({"question": "Are future-observation intervals separated from latent-state intervals?", "answer": "yes", "evidence": "Audit reports both latent_state and future_observation coverage; future-observation variance adds R."})
    decisions.append({"question": "Is rapid-removal model stable enough for strong paper claim?", "answer": "caution", "evidence": "Rapid-removal training samples are small and boundary stability is reported separately."})
    return decisions


def _mean_mae(rows, model):
    valid = [row["absolute_error"] for row in rows if row["valid_status"] == "valid" and row["model"] == model and row["R_scenario"] == ss.PRIMARY_R_SCENARIO and row["horizon"] == "1h"]
    return ss._round(sum(valid) / len(valid)) if valid else None


def _yes_no(left, right):
    if left is None or right is None:
        return "insufficient_data"
    return "yes" if left < right else "no"


def _metric_rows(rows, group_keys):
    grouped = defaultdict(list)
    for row in rows:
        grouped[tuple(row[key] for key in group_keys)].append(row)
    output = []
    for key, values in sorted(grouped.items()):
        output.append({**dict(zip(group_keys, key)), **_simple_metrics([row["observed_value"] for row in values], [row["forecast_mean"] for row in values]), "N": len(values), "maximum_AE": max(row["absolute_error"] for row in values) if values else None})
    return output


def _simple_metrics(actual, predicted):
    actual = list(actual)
    predicted = list(predicted)
    if not actual:
        return {"MAE": None, "RMSE": None, "bias": None, "median_AE": None, "maximum_AE": None}
    errors = [p - a for a, p in zip(actual, predicted)]
    abs_errors = [abs(error) for error in errors]
    return {
        "MAE": ss._round(sum(abs_errors) / len(abs_errors)),
        "RMSE": ss._round(sqrt(sum(error * error for error in errors) / len(errors))),
        "bias": ss._round(sum(errors) / len(errors)),
        "median_AE": ss._round(median(abs_errors)),
        "maximum_AE": ss._round(max(abs_errors)),
    }


def _variance(values):
    if len(values) < 2:
        return None
    mean = sum(values) / len(values)
    return ss._round(sum((value - mean) ** 2 for value in values) / len(values))


def _lag1(values):
    if len(values) < 3:
        return None
    left = values[:-1]
    right = values[1:]
    ml = sum(left) / len(left)
    mr = sum(right) / len(right)
    denom = sqrt(sum((x - ml) ** 2 for x in left) * sum((y - mr) ** 2 for y in right))
    return ss._round(sum((x - ml) * (y - mr) for x, y in zip(left, right)) / denom) if denom else None


def _signed_runs(values):
    runs = 0
    current = 0
    last_sign = 0
    for value in values:
        sign = 1 if value > 0 else -1 if value < 0 else 0
        if sign and sign == last_sign:
            current += 1
        else:
            if current >= 3:
                runs += 1
            current = 1 if sign else 0
            last_sign = sign
    if current >= 3:
        runs += 1
    return runs


def _manifest(campaign, state_payload):
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "source_state_space_algorithm_version": state_payload["algorithm_version"],
        "audit_scope": "Read-only validation and ablation audit; production state-space results are not modified.",
        "ablation_models": ["open_loop_reduced_transition", "reduced_sirem_informed_kalman", "generic_local_level_kalman"],
        "q_grid": ss.Q_GRID,
        "r_scenarios": ss.R_SCENARIOS,
        "interval_semantics": "Latent-state intervals exclude R; future-observation predictive intervals include R.",
        "guardrail": "No ACH, physical ventilation, beta, lambda_v, C_bm, C_out or exhalation-rate estimation.",
        "generated_at": timezone.now().replace(microsecond=0).isoformat(),
    }


def _summary_markdown(payload):
    lines = [
        "# State-Space Scientific Validation and Ablation Audit",
        "",
        f"Campaign: {payload['campaign_id']} - {payload['campaign_name']}",
        f"Algorithm: {payload['algorithm_version']}",
        "",
        "This audit is read-only exploratory research software and does not change stored campaign data.",
        "",
        "## Scientific decisions",
    ]
    for row in payload["scientific_decision"]:
        lines.append(f"- {row['question']} {row['answer']}. {row['evidence']}")
    lines.extend(["", "## Interval note", "Future-observation predictive intervals include observation noise R; latent-state intervals do not."])
    return "\n".join(lines) + "\n"


def _csv(rows):
    output = StringIO()
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No rows."})
    return output.getvalue()


def _write_rows(sheet, rows):
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
    else:
        sheet.append(["No rows."])


def _excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value
