import csv
from dataclasses import dataclass
from datetime import datetime, timezone as datetime_timezone
from io import BytesIO, StringIO
from statistics import median, pstdev

from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook


PHASES = ["baseline", "accumulation", "rapid_removal", "post_event"]

DEFAULT_RADONEYE_CYCLES = [
    {
        "cycle_label": "Cycle 1",
        "baseline_start": "2024-05-24 00:47",
        "baseline_end": "2024-05-25 08:47",
        "accumulation_start": "2024-05-25 09:47",
        "accumulation_end": "2024-05-29 22:47",
        "rapid_removal_start": "2024-05-29 22:47",
        "rapid_removal_end": "2024-05-30 01:47",
        "post_event_start": "2024-05-30 02:47",
        "post_event_end": "2024-05-31 23:47",
        "evidence_status": "PROVISIONAL",
        "note": "Provisional stored timestamps; documented ventilation period requires timezone/event-log confirmation.",
    },
    {
        "cycle_label": "Cycle 2",
        "baseline_start": "2024-06-09 00:44",
        "baseline_end": "2024-06-10 08:44",
        "accumulation_start": "2024-06-10 09:44",
        "accumulation_end": "2024-06-13 21:44",
        "rapid_removal_start": "2024-06-13 21:44",
        "rapid_removal_end": "2024-06-14 01:44",
        "post_event_start": "2024-06-14 02:44",
        "post_event_end": "2024-06-15 23:44",
        "evidence_status": "PROVISIONAL",
        "note": "Provisional stored timestamps; documented ventilation period requires timezone/event-log confirmation.",
    },
]


@dataclass
class EventCycleDefinition:
    cycle_label: str = ""
    baseline_start: datetime | None = None
    baseline_end: datetime | None = None
    accumulation_start: datetime | None = None
    accumulation_end: datetime | None = None
    rapid_removal_start: datetime | None = None
    rapid_removal_end: datetime | None = None
    post_event_start: datetime | None = None
    post_event_end: datetime | None = None
    evidence_status: str = ""
    note: str = ""

    @classmethod
    def from_mapping(cls, data):
        return cls(
            cycle_label=data.get("cycle_label") or "Documented cycle",
            baseline_start=parse_event_timestamp(data.get("baseline_start")),
            baseline_end=parse_event_timestamp(data.get("baseline_end")),
            accumulation_start=parse_event_timestamp(data.get("accumulation_start")),
            accumulation_end=parse_event_timestamp(data.get("accumulation_end")),
            rapid_removal_start=parse_event_timestamp(data.get("rapid_removal_start")),
            rapid_removal_end=parse_event_timestamp(data.get("rapid_removal_end")),
            post_event_start=parse_event_timestamp(data.get("post_event_start")),
            post_event_end=parse_event_timestamp(data.get("post_event_end")),
            evidence_status=data.get("evidence_status") or "",
            note=data.get("note") or "",
        )


def default_event_cycles_for_campaign(campaign):
    if "RadonEye Salerno 2024" in campaign.name:
        return [EventCycleDefinition.from_mapping(row) for row in DEFAULT_RADONEYE_CYCLES]
    return []


def parse_event_timestamp(value):
    if not value:
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        parsed = parse_datetime(text)
        if parsed is None:
            parsed = datetime.fromisoformat(text.replace(" ", "T"))
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed, datetime_timezone.utc)
    return parsed


def analyse_documented_cycles(campaign, definitions):
    measurements = list(
        campaign.measurements.exclude(radon_bq_m3=None)
        .order_by("measured_at", "id")
        .values("measured_at", "radon_bq_m3", "regime")
    )
    expected_interval_hours = _expected_interval_hours(measurements)
    timezone_status = _timezone_status(measurements)
    cycles = [
        _analyse_cycle(definition, measurements, expected_interval_hours, timezone_status)
        for definition in definitions
    ]
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "timezone_status": timezone_status,
        "caution": (
            "Rapid-removal windows are associated with documented ventilation periods, but exact opening "
            "times may require timezone and event-log confirmation."
        ),
        "cycles": cycles,
    }


def cycle_rows_for_export(payload):
    rows = []
    for cycle in payload["cycles"]:
        baseline = cycle["metrics"]["baseline"]
        accumulation = cycle["metrics"]["accumulation"]
        removal = cycle["metrics"]["rapid_removal"]
        post = cycle["metrics"]["post_event"]
        rows.append(
            {
                "cycle_label": cycle["definition"]["cycle_label"],
                "evidence_status": cycle["definition"]["evidence_status"],
                "baseline_count": baseline["observation_count"],
                "baseline_mean": baseline["mean"],
                "baseline_median": baseline["median"],
                "accumulation_starting_concentration": accumulation["starting_concentration"],
                "accumulation_max_concentration": accumulation["maximum_concentration"],
                "accumulation_max_timestamp": accumulation["maximum_timestamp"],
                "accumulation_endpoint_change": accumulation["endpoint_change"],
                "accumulation_linear_slope": accumulation["ordinary_linear_slope"],
                "accumulation_theil_sen_slope": accumulation["theil_sen_slope"],
                "rapid_removal_begin_concentration": removal["beginning_concentration"],
                "rapid_removal_min_concentration": removal["minimum_concentration"],
                "rapid_removal_min_timestamp": removal["minimum_timestamp"],
                "rapid_removal_absolute_decrease": removal["absolute_decrease"],
                "rapid_removal_percentage_decrease": removal["percentage_decrease"],
                "time_to_50_bq_m3_hours": removal["time_to_first_lte_50_bq_m3_hours"],
                "time_to_30_bq_m3_hours": removal["time_to_first_lte_30_bq_m3_hours"],
                "post_event_count": post["observation_count"],
                "post_event_mean": post["mean"],
                "post_event_median": post["median"],
                "quality_affected": cycle["quality_affected"],
                "missing_observations": cycle["missing_observations_total"],
                "gap_count": cycle["gap_count_total"],
                "note": cycle["definition"]["note"],
            }
        )
    return rows


def build_documented_events_csv(payload):
    output = StringIO()
    rows = cycle_rows_for_export(payload)
    headers = list(rows[0].keys()) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No documented cycles defined."})
    return output.getvalue()


def build_documented_events_workbook(payload):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Documented Event Summary"
    rows = cycle_rows_for_export(payload)
    _write_rows(summary, rows)

    definitions = workbook.create_sheet("Cycle Definitions")
    definition_rows = [
        {"cycle_label": cycle["definition"]["cycle_label"], **cycle["definition"]}
        for cycle in payload["cycles"]
    ]
    _write_rows(definitions, definition_rows)

    phases = workbook.create_sheet("Phase Metrics")
    phase_rows = []
    for cycle in payload["cycles"]:
        for phase, metrics in cycle["metrics"].items():
            phase_rows.append({"cycle_label": cycle["definition"]["cycle_label"], "phase": phase, **metrics})
    _write_rows(phases, phase_rows)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _analyse_cycle(definition, measurements, expected_interval_hours, timezone_status):
    phase_windows = {
        "baseline": (definition.baseline_start, definition.baseline_end),
        "accumulation": (definition.accumulation_start, definition.accumulation_end),
        "rapid_removal": (definition.rapid_removal_start, definition.rapid_removal_end),
        "post_event": (definition.post_event_start, definition.post_event_end),
    }
    phase_rows = {
        phase: _rows_in_window(measurements, start, end)
        for phase, (start, end) in phase_windows.items()
    }
    phase_quality = {
        phase: _phase_quality(rows, phase_windows[phase][0], phase_windows[phase][1], expected_interval_hours)
        for phase, rows in phase_rows.items()
    }
    return {
        "definition": _definition_payload(definition),
        "timezone_status": timezone_status,
        "metrics": {
            "baseline": _descriptive_stats(phase_rows["baseline"]),
            "accumulation": _accumulation_metrics(phase_rows["accumulation"], definition.accumulation_start),
            "rapid_removal": _rapid_removal_metrics(phase_rows["rapid_removal"], definition.rapid_removal_start),
            "post_event": _descriptive_stats(phase_rows["post_event"]),
        },
        "phase_quality": phase_quality,
        "quality_affected": any(value["quality_affected"] for value in phase_quality.values()),
        "missing_observations_total": sum(value["missing_observations"] for value in phase_quality.values()),
        "gap_count_total": sum(len(value["gaps"]) for value in phase_quality.values()),
        "plot": _plot_payload(phase_rows, phase_windows),
    }


def _definition_payload(definition):
    payload = {"cycle_label": definition.cycle_label, "evidence_status": definition.evidence_status, "note": definition.note}
    for phase in PHASES:
        payload[f"{phase}_start"] = _iso(getattr(definition, f"{phase}_start"))
        payload[f"{phase}_end"] = _iso(getattr(definition, f"{phase}_end"))
    return payload


def _rows_in_window(measurements, start, end):
    if start is None or end is None:
        return []
    return [row for row in measurements if row["measured_at"] and start <= row["measured_at"] <= end]


def _descriptive_stats(rows):
    values = _values(rows)
    return {
        "observation_count": len(values),
        "mean": _round(sum(values) / len(values)) if values else None,
        "median": _round(median(values)) if values else None,
        "standard_deviation": _round(pstdev(values)) if len(values) > 1 else 0 if values else None,
        "minimum": _round(min(values)) if values else None,
        "maximum": _round(max(values)) if values else None,
    }


def _accumulation_metrics(rows, window_start):
    values = _values(rows)
    maximum = _extreme_row(rows, max)
    first = rows[0] if rows else None
    last = rows[-1] if rows else None
    return {
        "starting_concentration": _radon(first),
        "maximum_concentration": _radon(maximum),
        "maximum_timestamp": _iso(maximum["measured_at"]) if maximum else None,
        "time_from_start_to_maximum_hours": _hours_between(window_start, maximum["measured_at"]) if maximum else None,
        "endpoint_change": _round(_radon(last) - _radon(first)) if first and last else None,
        "ordinary_linear_slope": _ordinary_slope(first, last),
        "theil_sen_slope": _theil_sen_slope(rows),
        "observation_count": len(values),
    }


def _rapid_removal_metrics(rows, window_start):
    first = rows[0] if rows else None
    minimum = _extreme_row(rows, min)
    begin = _radon(first)
    min_value = _radon(minimum)
    decrease = _round(begin - min_value) if begin is not None and min_value is not None else None
    return {
        "beginning_concentration": begin,
        "minimum_concentration": min_value,
        "minimum_timestamp": _iso(minimum["measured_at"]) if minimum else None,
        "absolute_decrease": decrease,
        "percentage_decrease": _round((decrease / begin) * 100) if decrease is not None and begin else None,
        "time_to_first_lte_50_bq_m3_hours": _time_to_threshold(rows, window_start, 50),
        "time_to_first_lte_30_bq_m3_hours": _time_to_threshold(rows, window_start, 30),
        "observation_count": len(rows),
    }


def _phase_quality(rows, start, end, expected_interval_hours):
    gaps = _gaps(rows, expected_interval_hours)
    expected_count = _expected_count(start, end, expected_interval_hours)
    missing = max(expected_count - len(rows), 0) if expected_count is not None else 0
    return {
        "missing_observations": missing,
        "gaps": gaps,
        "quality_affected": bool(missing or gaps or any((row.get("regime") or "").lower() == "quality_affected" for row in rows)),
    }


def _plot_payload(phase_rows, phase_windows):
    rows = []
    for phase in PHASES:
        rows.extend({"phase": phase, **row} for row in phase_rows[phase])
    rows = sorted(rows, key=lambda row: row["measured_at"])
    if not rows:
        return {"points": "", "markers": {}, "phase_windows": _phase_windows_payload(phase_windows), "phase_bands": []}
    width, height, pad = 700, 220, 28
    values = [_radon(row) for row in rows]
    times = [row["measured_at"] for row in rows]
    min_value, max_value = min(values), max(values)
    span = max(max_value - min_value, 1)
    total_seconds = max((times[-1] - times[0]).total_seconds(), 1)
    points = []
    for row, value in zip(rows, values):
        x = pad + ((row["measured_at"] - times[0]).total_seconds() / total_seconds) * (width - 2 * pad)
        y = height - pad - ((value - min_value) / span) * (height - 2 * pad)
        points.append(f"{x:.1f},{y:.1f}")
    maximum = _extreme_row(rows, max)
    rapid_rows = phase_rows["rapid_removal"]
    minimum = _extreme_row(rapid_rows, min)
    return {
        "points": " ".join(points),
        "markers": {
            "maximum": _marker(maximum, times[0], total_seconds, min_value, span, width, height, pad),
            "rapid_removal_begin": _marker(rapid_rows[0], times[0], total_seconds, min_value, span, width, height, pad) if rapid_rows else None,
            "minimum_after_removal": _marker(minimum, times[0], total_seconds, min_value, span, width, height, pad),
        },
        "phase_windows": _phase_windows_payload(phase_windows),
        "phase_bands": _phase_bands(phase_windows, times[0], total_seconds, width, pad),
    }


def _phase_windows_payload(phase_windows):
    return {phase: {"start": _iso(start), "end": _iso(end)} for phase, (start, end) in phase_windows.items()}


def _phase_bands(phase_windows, first_time, total_seconds, width, pad):
    bands = []
    colors = {
        "baseline": "#e8eef7",
        "accumulation": "#fff0c9",
        "rapid_removal": "#d9f3ee",
        "post_event": "#eef8e8",
    }
    for phase, (start, end) in phase_windows.items():
        if not start or not end:
            continue
        x = pad + ((start - first_time).total_seconds() / total_seconds) * (width - 2 * pad)
        x_end = pad + ((end - first_time).total_seconds() / total_seconds) * (width - 2 * pad)
        bands.append({"phase": phase, "x": round(x, 1), "width": round(max(x_end - x, 1), 1), "color": colors[phase]})
    return bands


def _marker(row, first_time, total_seconds, min_value, span, width, height, pad):
    if not row:
        return None
    value = _radon(row)
    x = pad + ((row["measured_at"] - first_time).total_seconds() / total_seconds) * (width - 2 * pad)
    y = height - pad - ((value - min_value) / span) * (height - 2 * pad)
    return {"x": round(x, 1), "y": round(y, 1), "timestamp": _iso(row["measured_at"]), "value": value}


def _expected_interval_hours(measurements):
    intervals = [
        (current["measured_at"] - previous["measured_at"]).total_seconds() / 3600
        for previous, current in zip(measurements, measurements[1:])
        if previous["measured_at"] and current["measured_at"] and current["measured_at"] > previous["measured_at"]
    ]
    return median(intervals) if intervals else None


def _expected_count(start, end, expected_interval_hours):
    if not start or not end or not expected_interval_hours:
        return None
    elapsed = (end - start).total_seconds() / 3600
    if elapsed < 0:
        return 0
    return int(round(elapsed / expected_interval_hours)) + 1


def _gaps(rows, expected_interval_hours):
    if not expected_interval_hours:
        return []
    gaps = []
    for previous, current in zip(rows, rows[1:]):
        interval = (current["measured_at"] - previous["measured_at"]).total_seconds() / 3600
        if interval > expected_interval_hours * 1.5:
            gaps.append(
                {
                    "from": _iso(previous["measured_at"]),
                    "to": _iso(current["measured_at"]),
                    "observed_interval_hours": _round(interval),
                    "expected_interval_hours": _round(expected_interval_hours),
                }
            )
    return gaps


def _theil_sen_slope(rows):
    slopes = []
    for left_index, left in enumerate(rows):
        for right in rows[left_index + 1:]:
            hours = (right["measured_at"] - left["measured_at"]).total_seconds() / 3600
            if hours > 0:
                slopes.append((_radon(right) - _radon(left)) / hours)
    return _round(median(slopes)) if slopes else None


def _ordinary_slope(first, last):
    if not first or not last:
        return None
    hours = (last["measured_at"] - first["measured_at"]).total_seconds() / 3600
    if hours <= 0:
        return None
    return _round((_radon(last) - _radon(first)) / hours)


def _time_to_threshold(rows, window_start, threshold):
    for row in rows:
        if _radon(row) <= threshold:
            return _hours_between(window_start, row["measured_at"])
    return None


def _hours_between(start, end):
    if not start or not end:
        return None
    return _round((end - start).total_seconds() / 3600)


def _extreme_row(rows, func):
    if not rows:
        return None
    return func(rows, key=lambda row: _radon(row))


def _values(rows):
    return [_radon(row) for row in rows if row.get("radon_bq_m3") is not None]


def _radon(row):
    if not row or row.get("radon_bq_m3") is None:
        return None
    return float(row["radon_bq_m3"])


def _timezone_status(measurements):
    if not measurements:
        return "No stored measurements available."
    aware = all(timezone.is_aware(row["measured_at"]) for row in measurements if row["measured_at"])
    return "Stored measurement timestamps are timezone-aware." if aware else "Some stored measurement timestamps are timezone-naive."


def _write_rows(sheet, rows):
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([row.get(header) for header in headers])
    else:
        sheet.append(["No documented cycles defined."])


def _iso(value):
    return value.isoformat() if value else None


def _round(value):
    return round(float(value), 3) if value is not None else None
