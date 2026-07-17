import csv
import json
from collections import defaultdict
from datetime import timedelta
from io import BytesIO, StringIO
from math import exp, isfinite, log, sqrt
from statistics import median, pstdev

from django.utils import timezone
from openpyxl import Workbook

from .documented_events import PHASES, default_event_cycles_for_campaign


VERSION = "apparent_dynamics_audit_v1"
LAMBDA_RN_H_MINUS_1 = 0.0075528
LAMBDA_RN_S_MINUS_1 = 2.098e-6
YELLOW_TUFF_STATUS = "supporting_material_measurement_not_directly_substituted"
NOT_IDENTIFIABLE = "NOT_IDENTIFIABLE"
BOUNDARY_SHIFTS_HOURS = [-2, -1, 0, 1, 2]
NOMINAL_VOLUME = {"floor_area_m2": 45.0, "ceiling_height_m": 2.7, "volume_m3": 121.5}
HEIGHT_SCENARIOS = [
    {"floor_area_m2": 45.0, "ceiling_height_m": 2.5, "volume_m3": 112.5},
    NOMINAL_VOLUME,
    {"floor_area_m2": 45.0, "ceiling_height_m": 3.0, "volume_m3": 135.0},
]
GEOMETRY_GRID = [
    {"floor_area_m2": area, "ceiling_height_m": height, "volume_m3": round(area * height, 3)}
    for area in [40.0, 45.0, 50.0]
    for height in [2.5, 2.7, 3.0]
]


def run_apparent_dynamics_audit(campaign):
    definitions = default_event_cycles_for_campaign(campaign)
    measurements = _measurement_rows(campaign)
    interval = _median_interval(measurements)
    cycles = [_audit_cycle(definition, measurements, interval) for definition in definitions]
    payload = {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "manifest": _manifest(campaign, definitions, measurements, interval),
        "cycles": cycles,
        "phase_fit_results": [row for cycle in cycles for row in cycle["phase_fit_results"]],
        "rapid_removal_floor_sensitivity": [row for cycle in cycles for row in cycle["rapid_removal_floor_sensitivity"]],
        "rapid_removal_boundary_sensitivity": [row for cycle in cycles for row in cycle["rapid_removal_boundary_sensitivity"]],
    }
    payload["cycle_comparison"] = _cycle_comparison(cycles)
    payload["volume_sensitivity"] = _volume_sensitivity(payload["phase_fit_results"])
    payload["identifiability_flags"] = _flag_rows(payload)
    payload["apparent_intervention_effect"] = _intervention_effect(cycles)
    return payload


def build_apparent_dynamics_csv(payload):
    return _csv(payload["phase_fit_results"])


def build_apparent_dynamics_workbook(payload):
    workbook = Workbook()
    workbook.active.title = "Manifest"
    _write_rows(workbook["Manifest"], [{"field": key, "value": value} for key, value in payload["manifest"].items()])
    _write_rows(workbook.create_sheet("Phase Fits"), payload["phase_fit_results"])
    _write_rows(workbook.create_sheet("Floor Sensitivity"), payload["rapid_removal_floor_sensitivity"])
    _write_rows(workbook.create_sheet("Boundary Sensitivity"), payload["rapid_removal_boundary_sensitivity"])
    _write_rows(workbook.create_sheet("Cycle Comparison"), payload["cycle_comparison"])
    _write_rows(workbook.create_sheet("Volume Sensitivity"), payload["volume_sensitivity"])
    _write_rows(workbook.create_sheet("Identifiability Flags"), payload["identifiability_flags"])
    _write_rows(workbook.create_sheet("Intervention Effect"), payload["apparent_intervention_effect"])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 58)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _audit_cycle(definition, measurements, interval):
    windows = _phase_windows(definition)
    phase_rows = {phase: _rows_in_window(measurements, *windows[phase]) for phase in PHASES}
    baseline_mean = _mean(_values(phase_rows["baseline"]))
    post_mean = _mean(_values(phase_rows["post_event"]))
    phase_fit_results = []
    for phase in PHASES:
        rows = phase_rows[phase]
        d0 = _descriptive_fit(definition.cycle_label, phase, rows, interval)
        phase_fit_results.append(d0)
        d1 = _free_equilibrium_fit(definition.cycle_label, phase, rows, interval, f"{phase}_D1_free_equilibrium")
        phase_fit_results.append(d1)
        if phase == "rapid_removal":
            phase_fit_results.append(_fixed_floor_fit(definition.cycle_label, phase, rows, baseline_mean, "D2_fixed_pre_event_baseline_floor", "pre_event_baseline_mean"))
            phase_fit_results.append(_fixed_floor_fit(definition.cycle_label, phase, rows, post_mean, "D3_fixed_post_event_floor", "post_event_mean_retrospective"))
            phase_fit_results.append(_endpoint_rate_row(definition.cycle_label, phase, rows, baseline_mean, "D4_endpoint_pre_event_baseline_floor", "pre_event_baseline_mean"))
            phase_fit_results.append(_endpoint_rate_row(definition.cycle_label, phase, rows, post_mean, "D4_endpoint_post_event_floor", "post_event_mean_retrospective"))
    return {
        "cycle_label": definition.cycle_label,
        "definition": _definition_payload(definition),
        "phase_fit_results": phase_fit_results,
        "rapid_removal_floor_sensitivity": _floor_sensitivity(definition.cycle_label, phase_rows["rapid_removal"], baseline_mean, post_mean),
        "rapid_removal_boundary_sensitivity": _boundary_sensitivity(definition, measurements, baseline_mean, post_mean),
        "plot": _cycle_plot_payload(phase_rows, phase_fit_results),
        "rapid_removal_plot": _rapid_plot_payload(phase_rows["rapid_removal"], phase_fit_results),
    }


def _descriptive_fit(cycle_label, phase, rows, interval):
    values = _values(rows)
    flags = _base_flags(rows, interval)
    if len(rows) < 2:
        flags.append("INSUFFICIENT_POINTS")
    if len(rows) and _dynamic_range(values) < _minimum_dynamic_range(values):
        flags.append("INSUFFICIENT_DYNAMIC_RANGE")
    if _non_monotonic(values):
        flags.append("NON_MONOTONIC_WITHIN_PHASE")
    return _result_row(
        cycle_label,
        phase,
        "D0_descriptive_linear_trend",
        rows,
        {
            "phase_mean_bq_m3": _round(_mean(values)) if values else None,
            "endpoint_change": _round(values[-1] - values[0]) if len(values) >= 2 else None,
            "linear_slope_bq_m3_per_h": _ordinary_slope(rows),
            "theil_sen_slope_bq_m3_per_h": _theil_sen_slope(rows),
            "convergence_status": "DESCRIPTIVE",
            "identifiability_status": _status(flags, descriptive=True),
            "flags": _join_flags(flags or ["FIT_OK"]),
        },
    )


def _free_equilibrium_fit(cycle_label, phase, rows, interval, fit_name):
    values = _values(rows)
    flags = _base_flags(rows, interval)
    if len(rows) < 3:
        flags.append("INSUFFICIENT_POINTS")
        return _result_row(cycle_label, phase, fit_name, rows, _empty_fit(flags))
    if _dynamic_range(values) < _minimum_dynamic_range(values):
        flags.append("INSUFFICIENT_DYNAMIC_RANGE")
    if _non_monotonic(values):
        flags.append("NON_MONOTONIC_WITHIN_PHASE")
    times = _relative_hours(rows)
    fit = fit_free_equilibrium(times, values)
    flags.extend(fit["flags"])
    if phase == "rapid_removal" and len(rows) <= 5:
        flags.append("DESCRIPTIVE_ONLY_SMALL_N")
    return _result_row(cycle_label, phase, fit_name, rows, _fit_payload(fit, flags))


def fit_free_equilibrium(times, values):
    if len(times) < 3:
        return {"flags": ["INSUFFICIENT_POINTS"], "convergence_status": "FAILED"}

    def fit_for_kappa(kappa):
        c0 = values[0]
        coeffs = [1 - exp(-kappa * t) for t in times]
        bases = [c0 * exp(-kappa * t) for t in times]
        denom = sum(a * a for a in coeffs)
        ceq = sum(a * (y - b) for a, y, b in zip(coeffs, values, bases)) / denom if denom else 0.0
        ceq = max(0.0, ceq)
        preds = [ceq + (c0 - ceq) * exp(-kappa * t) for t in times]
        return _fit_stats(values, preds, kappa, ceq)

    best = _minimize_positive(fit_for_kappa)
    profile = _profile_kappa(fit_for_kappa, best["sse"])
    best["profile"] = profile
    flags = []
    if best["kappa"] <= 1e-5 or best["ceq"] <= 1e-9:
        flags.append("PARAMETER_AT_BOUNDARY")
    if profile.get("near_optimal_kappa_ratio") and profile["near_optimal_kappa_ratio"] > 20:
        flags.append("KAPPA_NOT_PRACTICALLY_IDENTIFIABLE")
    if profile.get("near_optimal_ceq_range") and profile["near_optimal_ceq_range"] > max(abs(best["ceq"]), 1.0) * 5:
        flags.append("EQUILIBRIUM_NOT_PRACTICALLY_IDENTIFIABLE")
    if not isfinite(best["rmse"]):
        flags.append("NUMERICAL_INSTABILITY")
    best["flags"] = flags
    best["convergence_status"] = "CONVERGED" if isfinite(best["rmse"]) else "FAILED"
    return best


def _fixed_floor_fit(cycle_label, phase, rows, floor, fit_name, floor_source):
    flags = _base_flags(rows, None)
    if floor is None or len(rows) < 2:
        flags.append("INSUFFICIENT_POINTS")
        return _result_row(cycle_label, phase, fit_name, rows, {**_empty_fit(flags), "floor_concentration": floor, "floor_source": floor_source})
    times = _relative_hours(rows)
    values = _values(rows)
    fit = fit_fixed_floor(times, values, floor)
    flags.extend(fit["flags"])
    if len(rows) <= 5:
        flags.append("DESCRIPTIVE_ONLY_SMALL_N")
    payload = _fit_payload(fit, flags)
    payload.update({"floor_concentration": _round(floor), "floor_source": floor_source})
    if floor_source == "post_event_mean_retrospective":
        payload["use_status"] = "retrospective_future_informed_descriptive_fit"
    return _result_row(cycle_label, phase, fit_name, rows, payload)


def fit_fixed_floor(times, values, floor):
    if len(times) < 2:
        return {"flags": ["INSUFFICIENT_POINTS"], "convergence_status": "FAILED"}

    def fit_for_kappa(kappa):
        preds = [floor + (values[0] - floor) * exp(-kappa * t) for t in times]
        return _fit_stats(values, preds, kappa, floor, floor_fixed=True)

    best = _minimize_positive(fit_for_kappa)
    flags = []
    if best["kappa"] <= 1e-5:
        flags.append("PARAMETER_AT_BOUNDARY")
    best["flags"] = flags
    best["convergence_status"] = "CONVERGED" if isfinite(best["rmse"]) else "FAILED"
    return best


def endpoint_kappa(c_start, c_end, floor, elapsed_hours):
    if elapsed_hours is None or elapsed_hours <= 0:
        return NOT_IDENTIFIABLE
    numerator = c_end - floor
    denominator = c_start - floor
    if numerator <= 0 or denominator <= 0 or numerator >= denominator:
        return NOT_IDENTIFIABLE
    return -log(numerator / denominator) / elapsed_hours


def _endpoint_rate_row(cycle_label, phase, rows, floor, fit_name, floor_source):
    values = _values(rows)
    elapsed = _elapsed_hours(rows)
    kappa = endpoint_kappa(values[0], values[-1], floor, elapsed) if len(values) >= 2 and floor is not None else NOT_IDENTIFIABLE
    flags = _base_flags(rows, None)
    if kappa == NOT_IDENTIFIABLE:
        flags.append("NOT_IDENTIFIABLE")
    return _result_row(
        cycle_label,
        phase,
        fit_name,
        rows,
        {
            "kappa_h_minus_1": _round(kappa) if kappa != NOT_IDENTIFIABLE else NOT_IDENTIFIABLE,
            "half_time_h": _round(log(2) / kappa) if kappa != NOT_IDENTIFIABLE and kappa > 0 else None,
            "floor_concentration": _round(floor) if floor is not None else None,
            "floor_source": floor_source,
            "convergence_status": "CALCULATED" if kappa != NOT_IDENTIFIABLE else "NOT_IDENTIFIABLE",
            "identifiability_status": "FIT_OK" if kappa != NOT_IDENTIFIABLE else "NOT_IDENTIFIABLE",
            "flags": _join_flags(flags or (["FIT_OK"] if kappa != NOT_IDENTIFIABLE else ["NOT_IDENTIFIABLE"])),
        },
    )


def _fit_payload(fit, flags):
    kappa = fit.get("kappa")
    ceq = fit.get("ceq")
    return {
        "kappa_h_minus_1": _round(kappa),
        "C_eq_bq_m3": _round(ceq),
        "s_bq_m3_per_h": _round(kappa * ceq) if kappa is not None and ceq is not None else None,
        "half_time_h": _round(log(2) / kappa) if kappa and kappa > 0 else None,
        "apparent_non_decay_relaxation_rate_h_minus_1": _round(kappa - LAMBDA_RN_H_MINUS_1) if kappa is not None else None,
        "rmse": _round(fit.get("rmse")),
        "mae": _round(fit.get("mae")),
        "maximum_absolute_residual": _round(fit.get("maximum_absolute_residual")),
        "convergence_status": fit.get("convergence_status", "CONVERGED"),
        "parameter_at_boundary": "PARAMETER_AT_BOUNDARY" in flags,
        "identifiability_status": _status(flags),
        "profile_near_optimal_kappa_min": _round((fit.get("profile") or {}).get("near_optimal_kappa_min")),
        "profile_near_optimal_kappa_max": _round((fit.get("profile") or {}).get("near_optimal_kappa_max")),
        "profile_near_optimal_kappa_ratio": _round((fit.get("profile") or {}).get("near_optimal_kappa_ratio")),
        "profile_best_C_eq": _round(ceq),
        "profile_C_eq_range": _round((fit.get("profile") or {}).get("near_optimal_ceq_range")),
        "kappa_C_eq_tradeoff_note": (fit.get("profile") or {}).get("tradeoff_note"),
        "flags": _join_flags(flags or ["FIT_OK"]),
    }


def _empty_fit(flags):
    return {
        "kappa_h_minus_1": None,
        "C_eq_bq_m3": None,
        "s_bq_m3_per_h": None,
        "half_time_h": None,
        "rmse": None,
        "mae": None,
        "maximum_absolute_residual": None,
        "convergence_status": "NOT_FITTED",
        "identifiability_status": _status(flags),
        "flags": _join_flags(flags),
    }


def _result_row(cycle_label, phase, fit_name, rows, extra):
    values = _values(rows)
    base = {
        "cycle_label": cycle_label,
        "phase": phase,
        "fit_name": fit_name,
        "observation_count": len(rows),
        "elapsed_span_h": _round(_elapsed_hours(rows)),
        "first_timestamp": _iso(rows[0]["measured_at"]) if rows else None,
        "last_timestamp": _iso(rows[-1]["measured_at"]) if rows else None,
        "initial_concentration_bq_m3": _round(values[0]) if values else None,
        "final_concentration_bq_m3": _round(values[-1]) if values else None,
        "dynamic_range_bq_m3": _round(_dynamic_range(values)) if values else None,
    }
    base.update(extra)
    return base


def _minimize_positive(fit_for_kappa):
    candidates = [10 ** (-4 + i * (4 / 80)) for i in range(81)]
    coarse = min((fit_for_kappa(k) for k in candidates), key=lambda row: row["sse"])
    center = coarse["kappa"]
    low = max(center / 4, 1e-5)
    high = min(center * 4, 10.0)
    for _ in range(80):
        left = low + (high - low) / 3
        right = high - (high - low) / 3
        if fit_for_kappa(left)["sse"] <= fit_for_kappa(right)["sse"]:
            high = right
        else:
            low = left
    return fit_for_kappa((low + high) / 2)


def _fit_stats(values, preds, kappa, ceq, floor_fixed=False):
    residuals = [p - y for p, y in zip(preds, values)]
    abs_residuals = [abs(value) for value in residuals]
    sse = sum(value * value for value in residuals)
    return {
        "kappa": kappa,
        "ceq": ceq,
        "predictions": preds,
        "sse": sse,
        "rmse": sqrt(sse / len(values)) if values else None,
        "mae": sum(abs_residuals) / len(abs_residuals) if abs_residuals else None,
        "maximum_absolute_residual": max(abs_residuals) if abs_residuals else None,
        "floor_fixed": floor_fixed,
    }


def _profile_kappa(fit_for_kappa, best_sse):
    kappas = sorted(set([10 ** (-4 + i * (4 / 120)) for i in range(121)]))
    rows = [fit_for_kappa(kappa) for kappa in kappas]
    tolerance = max(best_sse * 1.05, best_sse + 1.0)
    near = [row for row in rows if row["sse"] <= tolerance]
    if not near:
        return {}
    k_min = min(row["kappa"] for row in near)
    k_max = max(row["kappa"] for row in near)
    ceq_values = [row["ceq"] for row in near]
    return {
        "near_optimal_kappa_min": k_min,
        "near_optimal_kappa_max": k_max,
        "near_optimal_kappa_ratio": k_max / k_min if k_min else None,
        "near_optimal_ceq_range": max(ceq_values) - min(ceq_values) if ceq_values else None,
        "tradeoff_note": "profile-error audit: multiple kappa/C_eq combinations are near-optimal" if len(near) > 3 else "profile-error audit: narrow near-optimal set",
    }


def _boundary_sensitivity(definition, measurements, baseline_mean, post_mean):
    rows = []
    for shift in BOUNDARY_SHIFTS_HOURS:
        start = definition.rapid_removal_start + timedelta(hours=shift)
        end = definition.rapid_removal_end
        phase_rows = _rows_in_window(measurements, start, end)
        if len(phase_rows) < 2:
            rows.append(_boundary_row(definition.cycle_label, shift, start, end, phase_rows, None, None, None, NOT_IDENTIFIABLE))
            continue
        d1 = fit_free_equilibrium(_relative_hours(phase_rows), _values(phase_rows))
        d2 = fit_fixed_floor(_relative_hours(phase_rows), _values(phase_rows), baseline_mean) if baseline_mean is not None else None
        d3 = fit_fixed_floor(_relative_hours(phase_rows), _values(phase_rows), post_mean) if post_mean is not None else None
        rows.append(_boundary_row(definition.cycle_label, shift, start, end, phase_rows, d1, d2, d3, None, baseline_mean, post_mean))
    return rows


def _boundary_row(cycle_label, shift, start, end, rows, d1, d2, d3, status, baseline_mean=None, post_mean=None):
    values = _values(rows)
    elapsed = _elapsed_hours(rows)
    return {
        "cycle_label": cycle_label,
        "boundary_shift_h": shift,
        "start_timestamp": _iso(start),
        "end_timestamp": _iso(end),
        "observation_count": len(rows),
        "initial_removal_concentration_bq_m3": _round(values[0]) if values else None,
        "final_removal_concentration_bq_m3": _round(values[-1]) if values else None,
        "elapsed_hours": _round(elapsed),
        "D1_kappa_h_minus_1": _round(d1.get("kappa")) if d1 else None,
        "D2_kappa_h_minus_1": _round(d2.get("kappa")) if d2 else None,
        "D3_kappa_h_minus_1": _round(d3.get("kappa")) if d3 else None,
        "D2_endpoint_kappa_h_minus_1": _endpoint_value(values, baseline_mean, elapsed),
        "D3_endpoint_kappa_h_minus_1": _endpoint_value(values, post_mean, elapsed),
        "half_time_h": _round(log(2) / d2["kappa"]) if d2 and d2.get("kappa") else None,
        "rmse": _round(d2.get("rmse")) if d2 else None,
        "status": status or "FIT_OK",
    }


def _endpoint_value(values, floor, elapsed):
    if len(values) < 2 or floor is None:
        return NOT_IDENTIFIABLE
    value = endpoint_kappa(values[0], values[-1], floor, elapsed)
    return _round(value) if value != NOT_IDENTIFIABLE else NOT_IDENTIFIABLE


def _floor_sensitivity(cycle_label, rows, baseline_mean, post_mean):
    out = []
    for floor, source, label in [
        (baseline_mean, "pre_event_baseline_mean", "D2"),
        (post_mean, "post_event_mean_retrospective", "D3"),
    ]:
        fit = fit_fixed_floor(_relative_hours(rows), _values(rows), floor) if floor is not None and len(rows) >= 2 else None
        values = _values(rows)
        elapsed = _elapsed_hours(rows)
        out.append(
            {
                "cycle_label": cycle_label,
                "scenario": label,
                "floor_source": source,
                "floor_concentration_bq_m3": _round(floor),
                "kappa_h_minus_1": _round(fit.get("kappa")) if fit else None,
                "half_time_h": _round(log(2) / fit["kappa"]) if fit and fit.get("kappa") else None,
                "endpoint_kappa_h_minus_1": _endpoint_value(values, floor, elapsed),
                "rmse": _round(fit.get("rmse")) if fit else None,
                "use_status": "retrospective_future_informed_descriptive_fit" if source.startswith("post") else "pre_event_available_floor_scenario",
            }
        )
    return out


def _cycle_comparison(cycles):
    rows = []
    for cycle in cycles:
        fits = cycle["phase_fit_results"]
        accumulation_d0 = _find_fit(fits, "accumulation", "D0_descriptive_linear_trend")
        accumulation_d1 = _find_fit(fits, "accumulation", "accumulation_D1_free_equilibrium")
        rapid_d1 = _find_fit(fits, "rapid_removal", "rapid_removal_D1_free_equilibrium")
        rapid_d2 = _find_fit(fits, "rapid_removal", "D2_fixed_pre_event_baseline_floor")
        rapid_d3 = _find_fit(fits, "rapid_removal", "D3_fixed_post_event_floor")
        post_d0 = _find_fit(fits, "post_event", "D0_descriptive_linear_trend")
        boundary = cycle["rapid_removal_boundary_sensitivity"]
        floor = cycle["rapid_removal_floor_sensitivity"]
        rows.append(
            {
                "cycle_label": cycle["cycle_label"],
                "accumulation_linear_slope": accumulation_d0.get("linear_slope_bq_m3_per_h"),
                "accumulation_apparent_kappa": accumulation_d1.get("kappa_h_minus_1"),
                "accumulation_apparent_C_eq": accumulation_d1.get("C_eq_bq_m3"),
                "rapid_removal_D1_kappa": rapid_d1.get("kappa_h_minus_1"),
                "rapid_removal_D2_kappa": rapid_d2.get("kappa_h_minus_1"),
                "rapid_removal_D3_kappa": rapid_d3.get("kappa_h_minus_1"),
                "rapid_removal_half_time_D2_h": rapid_d2.get("half_time_h"),
                "apparent_source_s_accumulation": accumulation_d1.get("s_bq_m3_per_h"),
                "post_event_mean": _mean_from_fit(post_d0),
                "boundary_D2_kappa_range": _range([row.get("D2_kappa_h_minus_1") for row in boundary]),
                "floor_kappa_range": _range([row.get("kappa_h_minus_1") for row in floor]),
            }
        )
    return rows


def _intervention_effect(cycles):
    rows = []
    for cycle in cycles:
        fits = cycle["phase_fit_results"]
        accumulation = _find_fit(fits, "accumulation", "accumulation_D1_free_equilibrium")
        for name in ["rapid_removal_D1_free_equilibrium", "D2_fixed_pre_event_baseline_floor", "D3_fixed_post_event_floor"]:
            rapid = _find_fit(fits, "rapid_removal", name)
            if _number(accumulation.get("kappa_h_minus_1")) is not None and _number(rapid.get("kappa_h_minus_1")) is not None:
                delta = _number(rapid.get("kappa_h_minus_1")) - _number(accumulation.get("kappa_h_minus_1"))
            else:
                delta = None
            rows.append(
                {
                    "cycle_label": cycle["cycle_label"],
                    "rapid_fit": name,
                    "kappa_accumulation": accumulation.get("kappa_h_minus_1"),
                    "kappa_rapid_removal": rapid.get("kappa_h_minus_1"),
                    "intervention_associated_increase_in_apparent_relaxation": _round(delta),
                    "interpretation_guardrail": "Not ventilation rate, not ACH, not measured air exchange.",
                }
            )
    return rows


def _volume_sensitivity(phase_fit_results):
    rows = []
    for fit in phase_fit_results:
        s_value = _number(fit.get("s_bq_m3_per_h"))
        if s_value is None:
            continue
        for scenario in GEOMETRY_GRID:
            rows.append(
                {
                    "cycle_label": fit["cycle_label"],
                    "phase": fit["phase"],
                    "fit_name": fit["fit_name"],
                    **scenario,
                    "kappa_h_minus_1": fit.get("kappa_h_minus_1"),
                    "C_eq_bq_m3": fit.get("C_eq_bq_m3"),
                    "half_time_h": fit.get("half_time_h"),
                    "effective_net_source_loading_Bq_per_h": _round(s_value * scenario["volume_m3"]),
                    "volume_note": "Q_eff is not material exhalation rate, yellow-tuff source rate, or physical wall flux.",
                }
            )
    return rows


def _flag_rows(payload):
    rows = []
    for row in payload["phase_fit_results"]:
        for flag in str(row.get("flags") or "").split("|"):
            if flag:
                rows.append({"cycle_label": row["cycle_label"], "phase": row["phase"], "fit_name": row["fit_name"], "flag": flag})
    return rows


def _cycle_plot_payload(phase_rows, fits):
    rows = sorted([row for phase in PHASES for row in phase_rows[phase]], key=lambda row: row["measured_at"])
    if not rows:
        return {}
    return _plot_payload(rows, fits)


def _rapid_plot_payload(rows, fits):
    return _plot_payload(rows, [fit for fit in fits if fit["phase"] == "rapid_removal"]) if rows else {}


def _plot_payload(rows, fits):
    values = _values(rows)
    times = [row["measured_at"] for row in rows]
    first = times[0]
    total = max((times[-1] - first).total_seconds() / 3600, 1e-9)
    v_min, v_max = min(values), max(values)
    span = max(v_max - v_min, 1)
    observed = _points(times, values, first, total, v_min, span)
    series = {}
    for fit in fits:
        if fit.get("kappa_h_minus_1") in (None, NOT_IDENTIFIABLE) or fit.get("C_eq_bq_m3") is None:
            continue
        kappa = float(fit["kappa_h_minus_1"])
        ceq = float(fit["C_eq_bq_m3"])
        c0 = values[0]
        preds = [ceq + (c0 - ceq) * exp(-kappa * ((time - first).total_seconds() / 3600)) for time in times]
        series[fit["fit_name"]] = _points(times, preds, first, total, v_min, span)
    return {"observed_points": observed, "fit_series": series}


def _points(times, values, first, total_hours, v_min, span):
    points = []
    for time, value in zip(times, values):
        x = 28 + (((time - first).total_seconds() / 3600) / total_hours) * 644
        y = 192 - ((value - v_min) / span) * 164
        points.append(f"{x:.1f},{y:.1f}")
    return " ".join(points)


def _measurement_rows(campaign):
    return list(campaign.measurements.exclude(radon_bq_m3=None).order_by("measured_at", "id").values("measured_at", "radon_bq_m3", "regime"))


def _phase_windows(definition):
    return {phase: (getattr(definition, f"{phase}_start"), getattr(definition, f"{phase}_end")) for phase in PHASES}


def _rows_in_window(measurements, start, end):
    if not start or not end:
        return []
    return [row for row in measurements if row["measured_at"] and start <= row["measured_at"] <= end]


def _definition_payload(definition):
    payload = {"cycle_label": definition.cycle_label, "evidence_status": definition.evidence_status, "note": definition.note}
    for phase in PHASES:
        payload[f"{phase}_start"] = _iso(getattr(definition, f"{phase}_start"))
        payload[f"{phase}_end"] = _iso(getattr(definition, f"{phase}_end"))
    return payload


def _relative_hours(rows):
    first = rows[0]["measured_at"]
    return [(row["measured_at"] - first).total_seconds() / 3600 for row in rows]


def _elapsed_hours(rows):
    if len(rows) < 2:
        return None
    return (rows[-1]["measured_at"] - rows[0]["measured_at"]).total_seconds() / 3600


def _values(rows):
    return [float(row["radon_bq_m3"]) for row in rows if row.get("radon_bq_m3") is not None]


def _mean(values):
    return sum(values) / len(values) if values else None


def _mean_from_fit(row):
    return row.get("phase_mean_bq_m3")


def _ordinary_slope(rows):
    if len(rows) < 2:
        return None
    elapsed = _elapsed_hours(rows)
    return _round((_values(rows)[-1] - _values(rows)[0]) / elapsed) if elapsed else None


def _theil_sen_slope(rows):
    slopes = []
    for i, left in enumerate(rows):
        for right in rows[i + 1 :]:
            hours = (right["measured_at"] - left["measured_at"]).total_seconds() / 3600
            if hours > 0:
                slopes.append((float(right["radon_bq_m3"]) - float(left["radon_bq_m3"])) / hours)
    return _round(median(slopes)) if slopes else None


def _base_flags(rows, interval):
    flags = []
    if interval and len(rows) >= 2:
        for prev, cur in zip(rows, rows[1:]):
            if (cur["measured_at"] - prev["measured_at"]).total_seconds() / 3600 > interval * 1.5:
                flags.append("SAMPLING_GAP_WITHIN_PHASE")
                break
    return flags


def _dynamic_range(values):
    return max(values) - min(values) if values else 0


def _minimum_dynamic_range(values):
    return max(5.0, abs(_mean(values) or 0) * 0.03)


def _non_monotonic(values):
    if len(values) < 3:
        return False
    diffs = [b - a for a, b in zip(values, values[1:]) if b != a]
    return bool(diffs) and not (all(diff >= 0 for diff in diffs) or all(diff <= 0 for diff in diffs))


def _status(flags, descriptive=False):
    if "INSUFFICIENT_POINTS" in flags:
        return "INSUFFICIENT_POINTS"
    if descriptive and not flags:
        return "DESCRIPTIVE_ONLY"
    severe = [flag for flag in flags if flag not in {"NON_MONOTONIC_WITHIN_PHASE", "DESCRIPTIVE_ONLY_SMALL_N"}]
    return "|".join(severe) if severe else "FIT_OK"


def _join_flags(flags):
    return "|".join(dict.fromkeys(flags))


def _find_fit(fits, phase, name):
    for fit in fits:
        if fit.get("phase") == phase and fit.get("fit_name") == name:
            return fit
    return {}


def _range(values):
    nums = [_number(value) for value in values]
    nums = [value for value in nums if value is not None]
    return _round(max(nums) - min(nums)) if nums else None


def _number(value):
    try:
        if value in (None, NOT_IDENTIFIABLE):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _median_interval(rows):
    intervals = [(b["measured_at"] - a["measured_at"]).total_seconds() / 3600 for a, b in zip(rows, rows[1:]) if b["measured_at"] > a["measured_at"]]
    return median(intervals) if intervals else None


def _manifest(campaign, definitions, measurements, interval):
    starts = [row["measured_at"] for row in measurements if row.get("measured_at")]
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "measurement_start": _iso(min(starts)) if starts else None,
        "measurement_end": _iso(max(starts)) if starts else None,
        "measurement_interval_hours": _round(interval),
        "documented_cycle_windows": [_definition_payload(definition) for definition in definitions],
        "provisional_timestamp_warning": "Stored timestamps are provisional and not confirmed as local window-opening times.",
        "equations": "dC/dt = s_r - kappa_r*C = kappa_r*(C_eq,r - C)",
        "radon_decay_constant_s_minus_1": LAMBDA_RN_S_MINUS_1,
        "radon_decay_constant_h_minus_1": LAMBDA_RN_H_MINUS_1,
        "fitting_method": "deterministic bounded one-dimensional search; C_eq solved analytically for each kappa in D1",
        "optimization_bounds": "kappa in [1e-5, 10] h^-1, C_eq >= 0",
        "convergence_tolerance": "80 ternary-search refinements after logarithmic coarse scan",
        "random_seed": None,
        "minimum_point_rules": "D1 requires at least 3 observations; fixed-floor fits require at least 2 observations.",
        "floor_scenarios": ["pre_event_baseline_mean", "post_event_mean_retrospective"],
        "boundary_shifts_hours": BOUNDARY_SHIFTS_HOURS,
        "volume_scenarios": GEOMETRY_GRID,
        "yellow_tuff_measurement_status": YELLOW_TUFF_STATUS,
        "algorithm_version": VERSION,
        "generated_at": timezone.now().replace(microsecond=0).isoformat(),
        "guardrail": "apparent_non_decay_relaxation_rate_h_minus_1 is not ACH or measured air exchange.",
    }


def _csv(rows):
    output = StringIO()
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No apparent dynamics rows."})
    return output.getvalue()


def _write_rows(sheet, rows):
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
    else:
        sheet.append(["No rows."])


def _excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value


def _iso(value):
    return value.isoformat() if value else None


def _round(value):
    return round(float(value), 6) if value is not None and value != NOT_IDENTIFIABLE else value
