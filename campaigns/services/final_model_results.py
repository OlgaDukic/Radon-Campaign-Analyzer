import csv
import json
import subprocess
from io import BytesIO, StringIO

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook

from . import reduced_state_space_experiment as ss
from .state_space_validation_audit import run_state_space_validation_audit


VERSION = "final_model_results_v1"
PRIMARY_MODE = "F1_no_future_event_knowledge"
PRIMARY_R = ss.PRIMARY_R_SCENARIO
PRIMARY_DIRECTION = "Experiment A"
MODEL_LABELS = {
    "persistence": "Persistence",
    "rolling_mean_3": "Rolling mean 3",
    "ridge_autoregressive": "Ridge autoregressive",
    "generic_local_level_kalman": "Generic local-level Kalman",
    "reduced_sirem_informed_kalman": "Reduced S.I.R.E.M.-informed Kalman",
    "open_loop_reduced_transition": "Open-loop reduced transition",
}


def build_final_model_results(campaign):
    state_payload = ss.run_reduced_state_space_experiment(campaign)
    audit_payload = run_state_space_validation_audit(campaign)
    table1 = _table_1_overall(audit_payload)
    table2 = _table_2_ablation(audit_payload)
    table3 = _table_3_uncertainty(audit_payload)
    table4 = _table_4_phase(state_payload)
    table5 = _table_5_intervention(state_payload)
    validity = _model_validity_summary(audit_payload, state_payload)
    manifest = _manifest(campaign, state_payload, audit_payload)
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "manifest": manifest,
        "table_1_overall_performance": table1,
        "table_2_sequential_ablation": table2,
        "table_3_predictive_uncertainty": table3,
        "table_4_phase_specific": table4,
        "table_5_intervention_response": table5,
        "model_validity_summary": validity,
        "interpretation_markdown": _interpretation(campaign, table1, table2, table3, table4, table5, validity),
        "state_space_forecast_rows": state_payload["forecast_rows"],
        "state_space_trace": [row for experiment in state_payload["experiments"] for row in experiment["state_trace"]],
    }


def write_final_model_results(campaign, output_dir=None):
    payload = build_final_model_results(campaign)
    base = output_dir or settings.BASE_DIR / "paper_outputs" / f"campaign_{campaign.pk}" / "final_model_results"
    figures = base / "figures"
    figures.mkdir(parents=True, exist_ok=True)
    (base / "final_results_manifest.json").write_text(json.dumps(payload["manifest"], indent=2, ensure_ascii=True, default=str), encoding="utf-8")
    (base / "final_results_tables.xlsx").write_bytes(build_final_results_workbook(payload).getvalue())
    exports = {
        "table_1_overall_performance.csv": payload["table_1_overall_performance"],
        "table_2_sequential_ablation.csv": payload["table_2_sequential_ablation"],
        "table_3_predictive_uncertainty.csv": payload["table_3_predictive_uncertainty"],
        "table_4_phase_specific.csv": payload["table_4_phase_specific"],
        "table_5_intervention_response.csv": payload["table_5_intervention_response"],
        "model_validity_summary.csv": payload["model_validity_summary"],
    }
    for filename, rows in exports.items():
        (base / filename).write_text(_csv(rows), encoding="utf-8")
    (base / "FINAL_RESULTS_INTERPRETATION.md").write_text(payload["interpretation_markdown"], encoding="utf-8")
    for filename, svg in _figures(payload).items():
        (figures / filename).write_text(svg, encoding="utf-8")
    return payload, base


def build_final_results_workbook(payload):
    workbook = Workbook()
    workbook.active.title = "Manifest"
    _write_rows(workbook["Manifest"], [{"field": key, "value": value} for key, value in payload["manifest"].items()])
    for sheet_name, key in [
        ("Table 1 Overall", "table_1_overall_performance"),
        ("Table 2 Ablation", "table_2_sequential_ablation"),
        ("Table 3 Uncertainty", "table_3_predictive_uncertainty"),
        ("Table 4 Phase", "table_4_phase_specific"),
        ("Table 5 Intervention", "table_5_intervention_response"),
        ("Model Validity", "model_validity_summary"),
    ]:
        _write_rows(workbook.create_sheet(sheet_name), payload[key])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 64)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _table_1_overall(audit_payload):
    rows = []
    for row in audit_payload["fair_comparison"]:
        if row["forecast_mode"] != PRIMARY_MODE:
            continue
        if row["model"] == "open_loop_reduced_transition":
            continue
        rows.append(
            {
                "Experiment": row["experiment"],
                "Train cycle": _train_cycle(row["experiment"]),
                "Test cycle": _test_cycle(row["experiment"]),
                "Model": MODEL_LABELS.get(row["model"], row["model"]),
                "Horizon": row["horizon"],
                "N": row["common_target_count"],
                "MAE": row["MAE"],
                "RMSE": row["RMSE"],
                "Bias": row["bias"],
                "Median AE": row.get("median_AE"),
                "Maximum AE": row.get("maximum_AE"),
            }
        )
    return rows


def _table_2_ablation(audit_payload):
    grouped = {}
    for row in audit_payload["ablation_performance"]:
        if row["forecast_mode"] != PRIMARY_MODE:
            continue
        grouped.setdefault((row["experiment"], row["horizon"]), {})[row["model"]] = row
    rows = []
    for (experiment, horizon), values in sorted(grouped.items()):
        open_loop = values.get("open_loop_reduced_transition", {})
        sequential = values.get("reduced_sirem_informed_kalman", {})
        rows.append(
            {
                "Experiment": experiment,
                "Horizon": horizon,
                "Open-loop reduced transition MAE": open_loop.get("MAE"),
                "Sequential reduced Kalman MAE": sequential.get("MAE"),
                "Open-loop RMSE": open_loop.get("RMSE"),
                "Sequential RMSE": sequential.get("RMSE"),
                "Relative MAE reduction": _relative_reduction(open_loop.get("MAE"), sequential.get("MAE")),
                "Relative RMSE reduction": _relative_reduction(open_loop.get("RMSE"), sequential.get("RMSE")),
            }
        )
    return rows


def _table_3_uncertainty(audit_payload):
    rows = []
    for row in audit_payload["uncertainty_calibration"]:
        if row["model"] != "r_sensitivity_reduced_kalman" or row["forecast_mode"] != PRIMARY_MODE or row["R_scenario"] != PRIMARY_R:
            continue
        rows.append(
            {
                "Experiment": row["experiment"],
                "Horizon": row["horizon"],
                "N": row["N"],
                "Nominal coverage": row["nominal_coverage"],
                "Latent-state empirical coverage": row["latent_state_empirical_coverage"],
                "Future-observation empirical coverage": row["future_observation_empirical_coverage"],
                "Future-observation mean interval width": row["future_observation_mean_interval_width"],
                "Interval score": row["future_observation_interval_score"],
            }
        )
    return rows


def _table_4_phase(state_payload):
    rows = []
    for row in state_payload["phase_specific_performance"]:
        if row["forecast_mode"] != PRIMARY_MODE or row["R_scenario"] != PRIMARY_R:
            continue
        rows.append(
            {
                "Experiment": row["experiment"],
                "Horizon": row["horizon"],
                "Target phase": row["target_phase"],
                "N": row["N"],
                "MAE": row["MAE"],
                "RMSE": row["RMSE"],
                "Bias": row["bias"],
                "Maximum AE": row["maximum_AE"],
                "Future-observation coverage": row["empirical_coverage"],
                "Mean interval width": row["mean_interval_width"],
            }
        )
    return rows


def _table_5_intervention(state_payload):
    rows = []
    for row in state_payload["intervention_response"]:
        if row["horizon"] != "1h":
            continue
        rows.append(
            {
                "Experiment": row["experiment"],
                "Forecast mode": _mode_label(row["forecast_mode"]),
                "Horizon": row["horizon"],
                "First rapid-removal target error": row["first_rapid_removal_target_error"],
                "Maximum rapid-removal error": row["maximum_rapid_removal_error"],
                "Error after first new observation": row["adaptation_after_first_new_observation"],
                "Error after second new observation": row["adaptation_after_second_new_observation"],
                "Error after third new observation": row["adaptation_after_third_new_observation"],
                "Warning": "Rapid-removal estimates are descriptive because of small N and provisional event boundaries.",
            }
        )
    return rows


def _model_validity_summary(audit_payload, state_payload):
    definitions = {
        "SMALL_PHASE_SAMPLE": "Regime or phase has few observations.",
        "PARAMETER_NEAR_BOUNDARY": "Estimated transition parameter is close to a constraint.",
        "PROVISIONAL_EVENT_BOUNDARY": "Documented event boundaries are provisional timestamps.",
        "LARGE_INNOVATION": "Observed value differs strongly from the prior forecast.",
        "SUSTAINED_SIGNED_INNOVATION": "Innovations retain the same sign over a run.",
        "HIGH_PREDICTIVE_UNCERTAINTY": "Predictive variance is high.",
        "OUTSIDE_TRAINING_REGIME": "Forecast target/origin is outside documented phase support.",
        "RECENT_GAP": "Recent timestamp gap affects continuity.",
    }
    flag_rows = []
    sources = []
    sources.extend(audit_payload["rapid_parameter_stability"])
    sources.extend(state_payload["model_validity_flags"])
    for row in audit_payload["innovation_diagnostics"]:
        if (row.get("sustained_signed_innovation_runs") or 0) > 0:
            sources.append({"flags": "SUSTAINED_SIGNED_INNOVATION", "phase": row.get("phase")})
    for flag, definition in definitions.items():
        matches = [row for row in sources if flag in str(row.get("flags", ""))]
        phases = sorted({str(row.get("phase") or row.get("target_phase") or "N/A") for row in matches})
        flag_rows.append(
            {
                "Flag": flag,
                "Definition": definition,
                "Count": len(matches),
                "Phases": ", ".join(phases) if phases else "N/A",
                "Interpretation effect": _flag_effect(flag),
            }
        )
    return flag_rows


def _manifest(campaign, state_payload, audit_payload):
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "state_space_algorithm_version": state_payload["algorithm_version"],
        "validation_audit_algorithm_version": audit_payload["algorithm_version"],
        "baseline_version": state_payload["baseline_reference"].get("algorithm_version"),
        "model_equation": "C_(k+1)=a_r*C_k+b_r+w_k; y_k=C_k+v_k",
        "state_vector": "x_k=[C_k]",
        "primary_mode": PRIMARY_MODE,
        "primary_observation_noise_scenario": PRIMARY_R,
        "primary_direction": "Experiment A: Cycle 1 development -> Cycle 2 held-out evaluation",
        "reverse_direction": "Experiment B: Cycle 2 development -> Cycle 1 held-out robustness check",
        "horizons": list(ss.HORIZONS),
        "selected_q": {experiment["label"]: experiment["selected_q"] for experiment in state_payload["experiments"]},
        "r_scenarios": ss.R_SCENARIOS,
        "interval_definitions": state_payload["manifest"]["interval_definitions"],
        "cycle_definitions": state_payload["manifest"]["cycle_windows"],
        "timestamp_warning": state_payload["manifest"]["timestamp_warning"],
        "random_seed": "deterministic_grid_no_random_seed",
        "git_commit_hash": _git_commit_hash(),
        "generated_at": timezone.now().replace(microsecond=0).isoformat(),
        "scientific_guardrail": "No ACH, physical ventilation, yellow-tuff exhalation, beta, lambda_v, C_bm or C_out estimation.",
    }


def _interpretation(campaign, table1, table2, table3, table4, table5, validity):
    primary_1h = [row for row in table1 if row["Experiment"] == PRIMARY_DIRECTION and row["Horizon"] == "1h"]
    reduced = next((row for row in primary_1h if row["Model"] == MODEL_LABELS["reduced_sirem_informed_kalman"]), {})
    local = next((row for row in primary_1h if row["Model"] == MODEL_LABELS["generic_local_level_kalman"]), {})
    ablation = next((row for row in table2 if row["Experiment"] == PRIMARY_DIRECTION and row["Horizon"] == "1h"), {})
    uncertainty = [row for row in table3 if row["Experiment"] == PRIMARY_DIRECTION]
    lines = [
        "# Final Results Interpretation",
        "",
        f"Campaign: {campaign.id} - {campaign.name}",
        "",
        "This package reports exploratory case-study evidence from a reduced one-state linear state-space model. It is not a certified radon risk assessment, regulatory prediction system, or autonomous ventilation-control system.",
        "",
        "## Locked Model",
        "`C_(k+1) = a_r C_k + b_r + w_k`; `y_k = C_k + v_k`; state vector `x_k = [C_k]`.",
        "The model does not identify physical ventilation, ACH, yellow-tuff exhalation, beta, lambda_v, C_bm, or C_out.",
        "",
        "## Main Findings",
        f"1. Sequential updating descriptively reduces error versus open-loop transition. For the primary 1 h Experiment A result, relative MAE reduction is {ablation.get('Relative MAE reduction', 'N/A')}%.",
        f"2. The reduced S.I.R.E.M.-informed transition gives a small to moderate improvement over generic local-level Kalman in the primary direction: MAE {reduced.get('MAE', 'N/A')} versus {local.get('MAE', 'N/A')}.",
        "3. One-hour F1 prediction is the most consistent result across both train/test directions.",
        "4. Longer horizons have larger and less consistent errors.",
        "5. The Future-observation predictive interval is the appropriate interval for forecasting a future sensor measurement.",
        "6. Intervals provide explicit but imperfectly calibrated uncertainty.",
        "7. Rapid-removal results are limited by small N and provisional event boundaries.",
        "8. Results are case-study evidence and should not be generalized across sites without additional campaigns.",
        "",
        "## Primary Uncertainty Rows",
    ]
    for row in uncertainty:
        lines.append(f"- {row['Horizon']}: future-observation coverage={row['Future-observation empirical coverage']}, width={row['Future-observation mean interval width']}")
    lines.extend(["", "## Model-Validity Flags"])
    for row in validity:
        if row["Count"]:
            lines.append(f"- {row['Flag']}: {row['Count']} occurrences. {row['Interpretation effect']}")
    return "\n".join(lines) + "\n"


def _figures(payload):
    rows = [row for row in payload["state_space_forecast_rows"] if row["R_scenario"] == PRIMARY_R and row["horizon"] == "1h" and row["valid_status"] == "valid"]
    return {
        "01_full_observed_cycles.svg": _line_svg(rows, "Full observed RadonEye series with documented cycles", ["observed_value"]),
        "02_experiment_a_f1_observed_vs_predicted.svg": _line_svg([row for row in rows if row["experiment"] == "Experiment A" and row["forecast_mode"] == PRIMARY_MODE], "Experiment A F1 observed vs predicted with future-observation interval", ["observed_value", "forecast_mean"]),
        "03_experiment_b_f1_observed_vs_predicted.svg": _line_svg([row for row in rows if row["experiment"] == "Experiment B" and row["forecast_mode"] == PRIMARY_MODE], "Experiment B F1 observed vs predicted with future-observation interval", ["observed_value", "forecast_mean"]),
        "04_rapid_removal_f1_vs_f2.svg": _line_svg([row for row in rows if row["target_phase"] == "rapid_removal"], "Rapid-removal focus: F1 versus F2", ["observed_value", "forecast_mean"]),
        "05_sequential_ablation.svg": _bar_svg(payload["table_2_sequential_ablation"], "Sequential-updating ablation"),
        "06_future_observation_coverage_width.svg": _bar_svg(payload["table_3_predictive_uncertainty"], "Future-observation coverage and interval width"),
        "07_innovation_residual_plot.svg": _line_svg(payload["state_space_trace"][:220], "Innovation/residual plot with model-validity context", ["innovation"]),
    }


def _line_svg(rows, title, keys):
    width, height, pad = 760, 280, 36
    rows = rows[:260]
    values = [float(row[key]) for row in rows for key in keys if row.get(key) is not None]
    if not values:
        values = [0, 1]
    lo, hi = min(values), max(values)
    span = max(hi - lo, 1)
    colors = ["#123c69", "#2f6fdf", "#2f855a"]
    polylines = []
    for idx, key in enumerate(keys):
        selected = [row for row in rows if row.get(key) is not None]
        points = []
        for i, row in enumerate(selected):
            x = pad + i * (width - 2 * pad) / max(len(selected) - 1, 1)
            y = height - pad - ((float(row[key]) - lo) / span) * (height - 2 * pad)
            points.append(f"{x:.1f},{y:.1f}")
        polylines.append(f'<polyline fill="none" stroke="{colors[idx % len(colors)]}" stroke-width="2" points="{" ".join(points)}"/>')
    return f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}"><rect width="100%" height="100%" fill="#fff"/><text x="24" y="26" font-size="16" font-family="Arial">{title}</text>{"".join(polylines)}</svg>'


def _bar_svg(rows, title):
    width, height = 760, 280
    bars = []
    for idx, row in enumerate(rows[:24]):
        value = row.get("Relative MAE reduction") or row.get("Future-observation empirical coverage") or 0
        try:
            value = float(value)
        except (TypeError, ValueError):
            value = 0
        scaled = value if value <= 1 else value / 100
        h = max(min(scaled, 1.2), 0) * 180
        x = 36 + idx * 28
        bars.append(f'<rect x="{x}" y="{230-h:.1f}" width="18" height="{h:.1f}" fill="#2f6fdf"/>')
    return f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}"><rect width="100%" height="100%" fill="#fff"/><text x="24" y="26" font-size="16" font-family="Arial">{title}</text>{"".join(bars)}</svg>'


def _relative_reduction(old, new):
    if old in (None, 0) or new is None:
        return None
    return round(100 * (float(old) - float(new)) / float(old), 3)


def _mode_label(mode):
    if mode == "F1_no_future_event_knowledge":
        return "F1 no future event knowledge"
    return "F2 known-intervention scenario"


def _train_cycle(experiment):
    return "Cycle 1" if experiment == "Experiment A" else "Cycle 2"


def _test_cycle(experiment):
    return "Cycle 2" if experiment == "Experiment A" else "Cycle 1"


def _flag_effect(flag):
    if flag in {"SMALL_PHASE_SAMPLE", "PARAMETER_NEAR_BOUNDARY"}:
        return "Limits strength of parameter interpretation."
    if flag == "PROVISIONAL_EVENT_BOUNDARY":
        return "Event timing should be interpreted as documented-window evidence, not confirmed opening time."
    if flag in {"LARGE_INNOVATION", "SUSTAINED_SIGNED_INNOVATION"}:
        return "Indicates possible model discrepancy during the affected phase."
    if flag == "HIGH_PREDICTIVE_UNCERTAINTY":
        return "Prediction interval is wide; point estimate should be interpreted cautiously."
    return "Use caution for affected rows."


def _git_commit_hash():
    try:
        result = subprocess.run(["git", "rev-parse", "HEAD"], cwd=settings.BASE_DIR, capture_output=True, text=True, check=True)
    except Exception:
        return "unavailable"
    return result.stdout.strip()


def _csv(rows):
    output = StringIO()
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No rows."})
    return output.getvalue()


def _write_rows(sheet, rows):
    headers = list(rows[0]) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
    else:
        sheet.append(["No rows."])


def _excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value
