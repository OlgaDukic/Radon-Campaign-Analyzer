from io import BytesIO
from datetime import datetime

from django.db.models import Avg, Max, Min
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .paper_outputs import build_row_reconciliation_summary
from .prediction_insights import build_prediction_insights
from .research_context import research_context_rows


SHEETS = [
    "Summary",
    "Segments",
    "Regime Counts",
    "Prediction Metrics",
    "Prediction Insights",
    "Prediction by Regime",
    "Prediction Errors",
    "Gaps",
    "Ingestion Diagnostics",
    "Measurements",
    "Source File Inventory",
    "Canonical Dataset Summary",
    "Canonical Hourly Data",
    "Quality Flags",
    "Quality Flag Dictionary",
    "Sampling Diagnostics",
    "Overlap Conflicts",
    "DST Diagnostics",
    "Resampling Summary",
    "Regime Sensitivity",
    "Prediction Skill by Regime",
    "Prediction Readiness",
    "SIREM Readiness",
    "Reproducibility Config",
    "Row Reconciliation Summary",
    "Campaign Summary",
    "Data Quality Summary",
    "Intervals and Gaps",
    "Measurement Regimes",
    "Episodes",
    "Regime Parameters",
    "Regime Confidence",
    "Important Episodes",
    "Feature Diagnostics",
    "Sudden Event Audit",
    "Episode Reasons",
    "Elevated Period Phases",
    "Profile Applicability",
    "Adaptive Recommendations",
    "Standardized Summary",
    "Transition Merge Audit",
    "Level Sensitivity",
    "Dynamic Sensitivity",
    "Prediction Summary",
    "Prediction Intervals",
    "Largest Errors",
    "Methodology Metadata",
    "Research Context",
]


def build_campaign_report_workbook(campaign, report=None):
    summary = report.summary_json if report and report.summary_json else {}
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary(workbook.create_sheet("Summary"), campaign, report, summary)
    _write_segments(workbook.create_sheet("Segments"), summary)
    _write_regime_counts(workbook.create_sheet("Regime Counts"), summary)
    _write_prediction_metrics(workbook.create_sheet("Prediction Metrics"), summary)
    _write_prediction_insights(workbook.create_sheet("Prediction Insights"), summary)
    _write_prediction_by_regime(workbook.create_sheet("Prediction by Regime"), summary)
    _write_prediction_errors(workbook.create_sheet("Prediction Errors"), summary)
    _write_gaps(workbook.create_sheet("Gaps"), summary)
    _write_ingestion_diagnostics(workbook.create_sheet("Ingestion Diagnostics"), summary)
    _write_measurements(workbook.create_sheet("Measurements"), campaign)
    _write_source_file_inventory(workbook.create_sheet("Source File Inventory"), summary)
    _write_canonical_dataset_summary(workbook.create_sheet("Canonical Dataset Summary"), summary)
    _write_canonical_hourly_data(workbook.create_sheet("Canonical Hourly Data"), summary)
    _write_quality_flags(workbook.create_sheet("Quality Flags"), summary)
    _write_quality_flag_dictionary(workbook.create_sheet("Quality Flag Dictionary"), summary)
    _write_sampling_diagnostics(workbook.create_sheet("Sampling Diagnostics"), summary)
    _write_overlap_conflicts(workbook.create_sheet("Overlap Conflicts"), summary)
    _write_dst_diagnostics(workbook.create_sheet("DST Diagnostics"), summary)
    _write_resampling_summary(workbook.create_sheet("Resampling Summary"), summary)
    _write_regime_sensitivity(workbook.create_sheet("Regime Sensitivity"), summary)
    _write_prediction_skill_by_regime(workbook.create_sheet("Prediction Skill by Regime"), summary)
    _write_prediction_readiness(workbook.create_sheet("Prediction Readiness"), summary)
    _write_sirem_readiness(workbook.create_sheet("SIREM Readiness"), summary)
    _write_reproducibility_config(workbook.create_sheet("Reproducibility Config"), summary)
    _write_row_reconciliation_summary(workbook.create_sheet("Row Reconciliation Summary"), summary)
    _write_campaign_summary(workbook.create_sheet("Campaign Summary"), campaign, report, summary)
    _write_data_quality_summary(workbook.create_sheet("Data Quality Summary"), summary)
    _write_intervals_and_gaps(workbook.create_sheet("Intervals and Gaps"), summary)
    _write_measurement_regimes(workbook.create_sheet("Measurement Regimes"), summary)
    _write_episodes(workbook.create_sheet("Episodes"), summary)
    _write_regime_parameters(workbook.create_sheet("Regime Parameters"), summary)
    _write_regime_confidence(workbook.create_sheet("Regime Confidence"), summary)
    _write_important_episodes(workbook.create_sheet("Important Episodes"), summary)
    _write_feature_diagnostics(workbook.create_sheet("Feature Diagnostics"), summary)
    _write_sudden_event_audit(workbook.create_sheet("Sudden Event Audit"), summary)
    _write_episode_reasons(workbook.create_sheet("Episode Reasons"), summary)
    _write_elevated_period_phases(workbook.create_sheet("Elevated Period Phases"), summary)
    _write_profile_applicability(workbook.create_sheet("Profile Applicability"), summary)
    _write_adaptive_recommendations(workbook.create_sheet("Adaptive Recommendations"), summary)
    _write_standardized_summary(workbook.create_sheet("Standardized Summary"), summary)
    _write_transition_merge_audit(workbook.create_sheet("Transition Merge Audit"), summary)
    _write_level_sensitivity(workbook.create_sheet("Level Sensitivity"), summary)
    _write_dynamic_sensitivity(workbook.create_sheet("Dynamic Sensitivity"), summary)
    _write_prediction_summary_v2(workbook.create_sheet("Prediction Summary"), summary)
    _write_prediction_intervals(workbook.create_sheet("Prediction Intervals"), summary)
    _write_largest_errors(workbook.create_sheet("Largest Errors"), summary)
    _write_methodology_metadata(workbook.create_sheet("Methodology Metadata"), summary)
    _write_research_context(workbook.create_sheet("Research Context"), campaign)

    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_compact_campaign_report_workbook(campaign, report=None):
    summary = report.summary_json if report and report.summary_json else {}
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary(workbook.create_sheet("Summary"), campaign, report, summary)
    _write_segments(workbook.create_sheet("Segments"), summary)
    _write_regime_counts(workbook.create_sheet("Regime Counts"), summary)
    _write_prediction_metrics(workbook.create_sheet("Prediction Metrics"), summary)
    _write_prediction_by_regime(workbook.create_sheet("Prediction by Regime"), summary)
    _write_gaps(workbook.create_sheet("Gaps"), summary)
    _write_ingestion_diagnostics(workbook.create_sheet("Ingestion Diagnostics"), summary)
    _write_source_file_inventory(workbook.create_sheet("Source File Inventory"), summary)
    _write_canonical_dataset_summary(workbook.create_sheet("Canonical Dataset Summary"), summary)
    _write_quality_flags(workbook.create_sheet("Quality Flags"), summary)
    _write_sampling_diagnostics(workbook.create_sheet("Sampling Diagnostics"), summary)
    _write_prediction_readiness(workbook.create_sheet("Prediction Readiness"), summary)
    _write_sirem_readiness(workbook.create_sheet("SIREM Readiness"), summary)
    _write_reproducibility_config(workbook.create_sheet("Reproducibility Config"), summary)
    _write_row_reconciliation_summary(workbook.create_sheet("Row Reconciliation Summary"), summary)
    _write_research_context(workbook.create_sheet("Research Context"), campaign)

    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _write_summary(worksheet, campaign, report, summary):
    stats = campaign.measurements.aggregate(
        first_at=Min("measured_at"),
        last_at=Max("measured_at"),
        mean_radon=Avg("radon_bq_m3"),
        max_radon=Max("radon_bq_m3"),
    )
    rows = [
        ("Field", "Value"),
        ("Campaign name", campaign.name),
        ("Location", _value(campaign.location)),
        ("Measurement date range", _date_range(stats["first_at"], stats["last_at"])),
        ("Uploaded file count", campaign.uploaded_files.count()),
        ("Imported measurement count", _value(summary.get("measurement_count"), campaign.measurements.count())),
        ("Segment count", _value(summary.get("segment_count"), len(summary.get("segments", [])) or None)),
        ("Gap count", _value(summary.get("gap_count"), len(summary.get("gaps", [])) or None)),
        ("Mean radon", _number(stats["mean_radon"])),
        ("Max radon", _number(stats["max_radon"])),
        ("Report created at", _datetime_cell(report.created_at if report else None)),
        ("Report updated at", "N/A"),
        ("Campaign created at", _datetime_cell(campaign.created_at)),
        ("Campaign updated at", _datetime_cell(campaign.updated_at)),
    ]
    for row in rows:
        worksheet.append(row)


def _write_segments(worksheet, summary):
    worksheet.append(
        [
            "Segment ID",
            "Start time",
            "End time",
            "Duration",
            "Mean radon",
            "Max radon",
            "Label",
            "Dominant regime",
            "Interpretation",
        ]
    )
    for segment in summary.get("segments", []):
        stats = segment.get("statistics", {}).get("radon_bq_m3", {})
        worksheet.append(
            [
                _value(segment.get("segment_id")),
                _datetime_cell(segment.get("start")),
                _datetime_cell(segment.get("end")),
                _duration(segment.get("start"), segment.get("end")),
                _number(stats.get("mean")),
                _number(stats.get("max")),
                _value(segment.get("segment_label")),
                _value(segment.get("dominant_regime")),
                _value(segment.get("interpretation_text")),
            ]
        )


def _write_regime_counts(worksheet, summary):
    worksheet.append(["Regime/Label", "Count"])
    for regime, count in summary.get("regime_counts", {}).items():
        worksheet.append([regime, count])


def _write_prediction_metrics(worksheet, summary):
    worksheet.append(
        [
            "Forecast horizon",
            "Model",
            "Samples",
            "Baseline MAE",
            "Model MAE",
            "MAE improvement %",
            "Baseline RMSE",
            "Model RMSE",
            "RMSE improvement %",
            "R2",
        ]
    )
    for horizon, model_results in summary.get("prediction_metrics", {}).items():
        baseline = model_results.get("naive_baseline", {})
        for model_name, metrics in model_results.items():
            worksheet.append(
                [
                    horizon,
                    model_name,
                    _value(metrics.get("samples")),
                    _number(baseline.get("mae")),
                    _number(metrics.get("mae")),
                    _improvement(baseline.get("mae"), metrics.get("mae")),
                    _number(baseline.get("rmse")),
                    _number(metrics.get("rmse")),
                    _improvement(baseline.get("rmse"), metrics.get("rmse")),
                    _number(metrics.get("r2")),
                ]
            )


def _write_prediction_by_regime(worksheet, summary):
    worksheet.append(
        [
            "Forecast horizon",
            "Model",
            "Regime/Label",
            "Samples",
            "MAE",
            "RMSE",
            "MAE improvement %",
            "RMSE improvement %",
        ]
    )
    for row in summary.get("prediction_metrics_by_regime", []):
        worksheet.append(
            [
                _value(row.get("horizon")),
                _value(row.get("model")),
                _value(row.get("regime")),
                _value(row.get("samples")),
                _number(row.get("mae")),
                _number(row.get("rmse")),
                _number(row.get("mae_improvement_percent")),
                _number(row.get("rmse_improvement_percent")),
            ]
        )


def _write_prediction_insights(worksheet, summary):
    worksheet.append(["Prediction Insights"])
    for insight in build_prediction_insights(summary):
        worksheet.append([insight])


def _write_prediction_errors(worksheet, summary):
    worksheet.append(
        [
            "Timestamp",
            "Forecast horizon",
            "Model",
            "Actual radon",
            "Predicted radon",
            "Absolute error",
            "Regime/Label",
            "Segment ID",
        ]
    )
    for row in summary.get("prediction_errors", []):
        worksheet.append(
            [
                _datetime_cell(row.get("timestamp")),
                _value(row.get("horizon")),
                _value(row.get("model")),
                _number(row.get("actual_radon")),
                _number(row.get("predicted_radon")),
                _number(row.get("absolute_error")),
                _value(row.get("regime")),
                _value(row.get("segment_id")),
            ]
        )


def _write_gaps(worksheet, summary):
    worksheet.append(["Gap start time", "Gap end time", "Duration minutes", "Reason/source"])
    for gap in summary.get("gaps", []):
        worksheet.append(
            [
                _datetime_cell(gap.get("from")),
                _datetime_cell(gap.get("to")),
                _number(gap.get("minutes")),
                _value(gap.get("reason") or gap.get("source")),
            ]
        )


def _write_ingestion_diagnostics(worksheet, summary):
    worksheet.append(
        [
            "Uploaded file name",
            "Imported rows/measurements",
            "Skipped rows",
            "Warnings/errors",
            "Detected overlap information",
            "Detected sheets",
            "Header row",
            "Mapped columns",
        ]
    )
    for file_debug in summary.get("ingestion_debug", []):
        mapped_columns = ", ".join(
            f"{key}: {value or 'N/A'}"
            for key, value in file_debug.get("mapped_columns", {}).items()
        )
        skipped_rows = file_debug.get("skipped_rows")
        worksheet.append(
            [
                _value(file_debug.get("filename")),
                _value(file_debug.get("parsed_measurement_rows")),
                _value(skipped_rows),
                _value(file_debug.get("skipped_reason") or file_debug.get("warning") or file_debug.get("error")),
                _value(file_debug.get("overlap_info") or file_debug.get("overlap") or file_debug.get("detected_overlap_information")),
                ", ".join(file_debug.get("detected_sheets", [])) or "N/A",
                _value(file_debug.get("detected_header_row")),
                mapped_columns or "N/A",
            ]
        )


def _write_measurements(worksheet, campaign):
    worksheet.append(["Timestamp", "Radon", "Temperature", "Humidity", "Pressure", "Regime/Label", "Segment ID"])
    for measurement in campaign.measurements.order_by("measured_at", "id"):
        worksheet.append(
            [
                _datetime_cell(measurement.measured_at),
                _number(measurement.radon_bq_m3),
                _number(measurement.temperature_c),
                _number(measurement.humidity_percent),
                _number(measurement.pressure_hpa),
                _value(measurement.regime),
                _value(measurement.segment_id),
            ]
        )


def _write_source_file_inventory(worksheet, summary):
    headers = [
        "Source file ID", "Filename", "Device ID", "Parsed start", "Parsed end",
        "Raw rows", "Imported measurement rows", "Detected columns", "Radon unit",
        "Environmental columns", "Missing values", "Duplicate timestamps",
        "Nominal interval minutes", "Interval distribution", "Irregular intervals",
        "Overlap duration minutes", "Overlap timestamp count", "Warnings/errors",
    ]
    worksheet.append(headers)
    for row in summary.get("source_file_inventory", []):
        worksheet.append([
            _value(row.get("source_file_id")), _value(row.get("filename")), _value(row.get("device_id")),
            _datetime_cell(row.get("parsed_start")), _datetime_cell(row.get("parsed_end")),
            _value(row.get("raw_rows")), _value(row.get("imported_measurement_rows")),
            _join(row.get("detected_columns")), _value(row.get("radon_unit")),
            _join(row.get("environmental_columns_available")), _stringify(row.get("missing_values")),
            _value(row.get("duplicate_timestamps_within_file")), _number(row.get("nominal_sampling_interval_minutes")),
            _stringify(row.get("sampling_interval_distribution")), _value(row.get("irregular_intervals")),
            _number(row.get("overlap_duration_minutes")), _value(row.get("overlap_timestamp_count")),
            _value(row.get("warnings_errors")),
        ])


def _write_canonical_dataset_summary(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in summary.get("canonical_dataset_summary", {}).items():
        worksheet.append([key, _value(value)])


def _write_canonical_hourly_data(worksheet, summary):
    headers = [
        "Interval start", "Radon mean", "Radon median", "Radon min", "Radon max",
        "Radon std", "Radon count", "Temperature mean", "Humidity mean",
        "Pressure mean", "Completeness ratio", "Quality flags",
    ]
    worksheet.append(headers)
    for row in summary.get("canonical_hourly_data", []):
        worksheet.append([
            _datetime_cell(row.get("interval_start")), _number(row.get("radon_mean")),
            _number(row.get("radon_median")), _number(row.get("radon_min")),
            _number(row.get("radon_max")), _number(row.get("radon_std")),
            _value(row.get("radon_count")), _number(row.get("temperature_mean")),
            _number(row.get("humidity_mean")), _number(row.get("pressure_mean")),
            _number(row.get("completeness_ratio")), _join(row.get("quality_flags")),
        ])


def _write_quality_flags(worksheet, summary):
    worksheet.append(["Quality flag", "Count"])
    for flag, count in summary.get("quality_flag_counts", {}).items():
        worksheet.append([flag, count])


def _write_quality_flag_dictionary(worksheet, summary):
    worksheet.append(["Quality flag", "Description"])
    for flag, description in summary.get("quality_flag_dictionary", {}).items():
        worksheet.append([flag, description])


def _write_sampling_diagnostics(worksheet, summary):
    diagnostics = summary.get("sampling_diagnostics", {})
    worksheet.append(["Field", "Value"])
    for key, value in diagnostics.items():
        if key != "gaps":
            worksheet.append([key, _stringify(value)])


def _write_overlap_conflicts(worksheet, summary):
    worksheet.append(["UTC timestamp", "Source file IDs", "Values", "Quality flags", "Note"])
    for row in summary.get("overlap_conflicts", []):
        worksheet.append([
            _datetime_cell(row.get("utc_timestamp")), _join(row.get("source_file_ids")),
            _stringify(row.get("values")), _join(row.get("quality_flags")), _value(row.get("note")),
        ])


def _write_dst_diagnostics(worksheet, summary):
    worksheet.append(["Timestamp", "Local timestamp", "UTC timestamp", "Timezone", "Flags", "Note"])
    for row in summary.get("dst_diagnostics", []):
        worksheet.append([
            _datetime_cell(row.get("timestamp")), _datetime_cell(row.get("local_timestamp")),
            _datetime_cell(row.get("utc_timestamp")), _value(row.get("timezone")),
            _join(row.get("flags")), _value(row.get("note")),
        ])


def _write_resampling_summary(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in summary.get("resampling_summary", {}).items():
        worksheet.append([key, _stringify(value)])


def _write_regime_sensitivity(worksheet, summary):
    worksheet.append(["Threshold multiplier", "Regime counts", "Regime durations", "Agreement %", "Transitions", "Most sensitive regimes", "Adjusted Rand Index"])
    for row in summary.get("regime_sensitivity", []):
        worksheet.append([
            _number(row.get("threshold_multiplier")), _stringify(row.get("regime_counts")),
            _stringify(row.get("regime_durations")), _number(row.get("percentage_agreement_with_baseline")),
            _value(row.get("transitions_count")), _join(row.get("most_sensitive_regimes")),
            _value(row.get("adjusted_rand_index")),
        ])


def _write_prediction_skill_by_regime(worksheet, summary):
    worksheet.append(["Forecast horizon", "Model", "Regime/Label", "Samples", "MAE", "RMSE", "Skill score vs persistence", "Small sample warning"])
    for row in summary.get("prediction_skill_by_regime", []):
        worksheet.append([
            _value(row.get("horizon")), _value(row.get("model")), _value(row.get("regime")),
            _value(row.get("samples")), _number(row.get("mae")), _number(row.get("rmse")),
            _number(row.get("skill_score_vs_persistence")), _value(row.get("small_sample_warning")),
        ])


def _write_prediction_readiness(worksheet, summary):
    worksheet.append(["Segment ID", "Regime", "Score", "Category", "Explanation", "Score reduction flags"])
    for row in summary.get("prediction_readiness", []):
        worksheet.append([
            _value(row.get("segment_id")), _value(row.get("regime")),
            _number(row.get("prediction_readiness_score")), _value(row.get("category")),
            _value(row.get("explanation")), _join(row.get("score_reduction_flags")),
        ])


def _write_sirem_readiness(worksheet, summary):
    worksheet.append(["Checklist item", "Available", "Source", "Notes", "Importance for SIREM"])
    for row in summary.get("sirem_readiness", []):
        worksheet.append([
            _value(row.get("item")), _value(row.get("available")), _value(row.get("source")),
            _value(row.get("notes")), _value(row.get("importance_for_sirem")),
        ])


def _write_reproducibility_config(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in summary.get("reproducibility_config", {}).items():
        worksheet.append([key, _stringify(value)])


def _write_row_reconciliation_summary(worksheet, summary):
    worksheet.append(["Field", "Value"])
    reconciliation = summary.get("row_reconciliation_summary") or build_row_reconciliation_summary(summary)
    for key, value in reconciliation.items():
        worksheet.append([key, _stringify(value)])


def _write_campaign_summary(worksheet, campaign, report, summary):
    rows = {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "report_id": report.id if report else None,
        "report_schema_version": summary.get("report_schema_version", "legacy"),
        "measurement_count": summary.get("measurement_count"),
        "segment_count": summary.get("segment_count"),
        "episode_count": len(summary.get("episodes", [])),
        "exploratory_notice": "Research prototype output; not certified radon risk assessment.",
    }
    worksheet.append(["Field", "Value"])
    for key, value in rows.items():
        worksheet.append([key, _stringify(value)])


def _write_data_quality_summary(worksheet, summary):
    rows = summary.get("quality_flag_details", [])
    _write_dict_rows(
        worksheet,
        rows,
        [
            "quality_flag",
            "flag_occurrences",
            "unique_raw_measurements_affected",
            "unique_canonical_measurements_affected",
            "hourly_or_analysis_rows_affected",
            "percentage_of_measurements_affected",
            "meaning",
        ],
    )


def _write_intervals_and_gaps(worksheet, summary):
    rows = summary.get("time_continuity", {}).get("intervals", [])
    _write_dict_rows(worksheet, rows, ["from", "to", "observed_interval_minutes", "expected_interval_minutes", "interval_class", "creates_segment_boundary", "segment_boundary_reason"])


def _write_measurement_regimes(worksheet, summary):
    rows = summary.get("measurement_regimes_v2", [])
    _write_dict_rows(
        worksheet,
        rows,
        [
            "timestamp",
            "radon_bq_m3",
            "segment_id",
            "observed_interval_hours",
            "concentration_level",
            "candidate_dynamic_state",
            "confirmed_dynamic_state",
            "legacy_regime",
            "adjacent_slope_bq_m3_per_hour",
            "short_slope_bq_m3_per_hour",
            "medium_slope_bq_m3_per_hour",
            "slope_acceleration_bq_m3_per_hour2",
            "rolling_median_radon",
            "local_variability_mad",
            "local_variability_normalized",
            "short_valid_observation_count",
            "medium_valid_observation_count",
            "distance_to_previous_gap_observations",
            "distance_to_previous_gap_hours",
            "distance_to_next_gap_observations",
            "distance_to_next_gap_hours",
            "raw_smoothed_disagreement",
            "confidence_score",
            "confidence_label",
            "confidence_reasons",
            "dynamic_reason_codes",
            "quality_flags",
        ],
    )


def _write_episodes(worksheet, summary):
    rows = summary.get("episodes", [])
    _write_dict_rows(
        worksheet,
        rows,
        [
            "campaign_id",
            "analysis_report_id",
            "segment_id",
            "episode_sequence_number",
            "episode_type",
            "legacy_episode_label",
            "start",
            "end",
            "duration_hours",
            "measurement_count",
            "starting_radon",
            "ending_radon",
            "min_radon",
            "max_radon",
            "mean_radon",
            "median_radon",
            "absolute_concentration_change",
            "relative_concentration_change_percent",
            "mean_adjacent_slope_bq_m3_per_hour",
            "robust_episode_slope_bq_m3_per_hour",
            "mean_slope_bq_m3_per_hour",
            "maximum_positive_slope",
            "maximum_negative_slope",
            "local_variability",
            "concentration_level_distribution",
            "dynamic_state_distribution",
            "distance_from_previous_gap_hours",
            "distance_to_next_gap_hours",
            "quality_status",
            "confidence_score",
            "confidence_category",
            "confidence_reason_codes",
            "dominant_reason_codes",
            "reason_code_summary",
            "positive_confidence_components",
            "negative_confidence_components",
            "start_transition_reasons",
            "end_transition_reasons",
            "regime_algorithm_version",
            "episode_algorithm_version",
            "parameter_set_identifier",
        ],
    )


def _write_regime_parameters(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in summary.get("regime_parameters", {}).items():
        worksheet.append([key, _stringify(value)])


def _write_regime_confidence(worksheet, summary):
    confidence = summary.get("regime_confidence_summary") or {}
    worksheet.append(["Section", "Key", "Value"])
    for key, value in (confidence.get("confidence_category_counts") or {}).items():
        worksheet.append(["confidence_category_counts", key, _stringify(value)])
    for state, distribution in (confidence.get("confidence_distribution_by_dynamic_state") or {}).items():
        worksheet.append(["confidence_distribution_by_dynamic_state", state, _stringify(distribution)])
    for episode_type, distribution in (confidence.get("confidence_distribution_by_episode_type") or {}).items():
        worksheet.append(["confidence_distribution_by_episode_type", episode_type, _stringify(distribution)])
    for reason, count in (confidence.get("reason_code_counts") or {}).items():
        worksheet.append(["reason_code_counts", reason, _stringify(count)])
    worksheet.append(["low_confidence_row_count", "rows", _stringify(confidence.get("low_confidence_row_count"))])
    worksheet.append(["low_confidence_episode_count", "episodes", _stringify(confidence.get("low_confidence_episode_count"))])
    if worksheet.max_row == 1:
        worksheet.append(["note", "N/A", "No regime confidence summary available"])


def _write_important_episodes(worksheet, summary):
    rows = summary.get("important_episodes", [])
    _write_dict_rows(
        worksheet,
        rows,
        [
            "start",
            "end",
            "episode_type",
            "starting_radon",
            "ending_radon",
            "max_radon",
            "duration_hours",
            "robust_episode_slope_bq_m3_per_hour",
            "confidence_score",
            "confidence_category",
            "confidence_reason_codes",
            "distance_from_previous_gap_hours",
            "distance_to_next_gap_hours",
        ],
    )


def _write_feature_diagnostics(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("feature_distribution_diagnostics", []),
        ["feature", "sample_count", "q01", "q05", "q10", "q25", "q50", "q75", "q90", "q95", "q99"],
    )


def _write_sudden_event_audit(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("sudden_event_audit", []),
        [
            "timestamp",
            "segment_id",
            "event_state",
            "previous_radon",
            "current_radon",
            "observed_interval_hours",
            "absolute_change",
            "relative_change_percent",
            "adjacent_slope_bq_m3_per_hour",
            "short_slope_bq_m3_per_hour",
            "threshold_used_bq_m3_per_hour",
            "trigger_rule",
            "threshold_satisfied",
        ],
    )


def _write_episode_reasons(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("episode_reason_summary", []),
        ["segment_id", "episode_sequence_number", "episode_type", "reason_code", "row_count", "row_percent"],
    )


def _write_elevated_period_phases(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("elevated_period_phase_table", []),
        [
            "inspection_period",
            "phase_start",
            "phase_end",
            "concentration_level_distribution",
            "dynamic_state_distribution",
            "episode_type",
            "start_radon",
            "end_radon",
            "min_radon",
            "max_radon",
            "robust_slope",
            "local_variability",
            "confidence",
            "reason_codes",
        ],
    )


def _write_profile_applicability(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in (summary.get("profile_applicability") or {}).items():
        worksheet.append([key, _stringify(value)])


def _write_adaptive_recommendations(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("adaptive_recommendations", []),
        ["parameter", "active_threshold", "recommended_threshold", "source", "override_accepted", "note"],
    )


def _write_standardized_summary(worksheet, summary):
    worksheet.append(["Field", "Value"])
    for key, value in (summary.get("standardized_campaign_summary") or {}).items():
        worksheet.append([key, _stringify(value)])


def _write_transition_merge_audit(worksheet, summary):
    _write_dict_rows(
        worksheet,
        summary.get("transition_merge_audit", []),
        ["merge_id", "segment_id", "episode_type", "original_episode_ids", "merge_decision", "merge_reason", "resulting_episode_id"],
    )


def _write_level_sensitivity(worksheet, summary):
    rows = summary.get("level_sensitivity", [])
    _write_dict_rows(worksheet, rows, ["threshold_multiplier", "level_counts", "level_percentages", "agreement_with_baseline_percent", "elevated_or_high_observations"])


def _write_dynamic_sensitivity(worksheet, summary):
    rows = summary.get("dynamic_sensitivity", [])
    _write_dict_rows(
        worksheet,
        rows,
        [
            "parameter_set_identifier",
            "slope_threshold_multiplier",
            "short_window_observations",
            "medium_window_observations",
            "minimum_state_persistence_observations",
            "variability_threshold_bq_m3",
            "state_counts",
            "state_duration_hours",
            "state_percentages",
            "episode_count_by_type",
            "episode_transition_count",
            "agreement_with_baseline_percent",
            "cohen_kappa",
            "baseline_episode_preservation_percent",
            "metric_note",
        ],
    )


def _write_prediction_summary_v2(worksheet, summary):
    rows = summary.get("prediction_summary_v2", [])
    _write_dict_rows(worksheet, rows, ["horizon", "model", "samples", "mae", "rmse", "bias", "median_absolute_error"])


def _write_prediction_intervals(worksheet, summary):
    rows = summary.get("prediction_intervals", [])
    _write_dict_rows(worksheet, rows, ["model", "nominal_coverage", "empirical_coverage", "average_interval_width", "residual_count", "method", "note"])


def _write_largest_errors(worksheet, summary):
    rows = summary.get("largest_errors_v2", summary.get("prediction_errors", []))
    _write_dict_rows(worksheet, rows, ["timestamp", "horizon", "model", "actual_radon", "predicted_radon", "absolute_error", "regime", "segment_id"])


def _write_methodology_metadata(worksheet, summary):
    config = summary.get("analysis_config", {})
    params = summary.get("regime_parameters", {})
    metadata = {
        "software_version": summary.get("reproducibility_config", {}).get("app_version_or_git_commit"),
        "algorithm_version": params.get("algorithm_version"),
        "run_timestamp": summary.get("reproducibility_config", {}).get("run_timestamp"),
        "sampling_interval": summary.get("time_continuity", {}).get("summary", {}).get("expected_sampling_interval_minutes"),
        "gap_thresholds": {
            "minor_interval_tolerance": config.get("minor_interval_tolerance"),
            "gap_tolerance_multiplier": config.get("gap_tolerance_multiplier"),
            "short_gap_multiplier": config.get("short_gap_multiplier"),
            "moderate_gap_multiplier": config.get("moderate_gap_multiplier"),
        },
        "concentration_thresholds": {
            "low": config.get("concentration_low_threshold_bq_m3"),
            "high": config.get("concentration_high_threshold_bq_m3"),
        },
        "slope_thresholds": {
            "stable": config.get("stable_slope_bq_m3_per_hour"),
            "trend": config.get("trend_slope_bq_m3_per_hour"),
            "sudden": config.get("sudden_change_bq_m3_per_hour"),
        },
        "rolling_windows": {
            "short": config.get("short_window_observations"),
            "medium": config.get("medium_window_observations"),
        },
        "persistence_rules": config.get("minimum_state_persistence_observations"),
        "hysteresis_parameters": config.get("concentration_hysteresis_bq_m3"),
        "confidence_score_rules": "deterministic score based on slope strength, duration, variability, and proximity to gaps; not probability",
        "confidence_formula": params.get("confidence_formula"),
        "prediction_split_configuration": summary.get("prediction_v2", {}).get("validation_policy", "chronological train/test split"),
        "random_seed": "N/A deterministic rules",
    }
    worksheet.append(["Field", "Value"])
    for key, value in metadata.items():
        worksheet.append([key, _stringify(value)])


def _write_research_context(worksheet, campaign):
    worksheet.append(["Section", "Field", "Value", "Evidence status"])
    for row in research_context_rows(campaign):
        worksheet.append(
            [
                row.get("section"),
                row.get("field"),
                _blank(row.get("value")),
                _blank(row.get("evidence_status")),
            ]
        )


def _write_dict_rows(worksheet, rows, headers):
    worksheet.append(headers)
    if not rows:
        worksheet.append(["N/A"] + [""] * (len(headers) - 1))
        return
    for row in rows:
        worksheet.append([_stringify(row.get(header)) for header in headers])


def _format_sheet(worksheet):
    worksheet.freeze_panes = "A2"
    if worksheet.max_row and worksheet.max_column:
        worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="EAF1F8")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    _apply_number_formats(worksheet)
    _wrap_long_text(worksheet)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), _max_width(worksheet.title, column_cells[0].column))


def _value(value, fallback="N/A"):
    if value is None or value == "":
        return fallback
    return value


def _blank(value):
    if value is None or value == "":
        return None
    return value


def _join(value):
    if not value:
        return "N/A"
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value) or "N/A"
    return str(value)


def _stringify(value):
    if value is None or value == "":
        return "N/A"
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    return value


def _number(value):
    if value is None or value == "":
        return "N/A"
    return round(float(value), 3)


def _datetime_cell(value):
    if not value:
        return "N/A"
    if isinstance(value, datetime):
        return _strip_tz(value)
    if isinstance(value, str):
        parsed = parse_datetime(value)
        if parsed:
            return _strip_tz(parsed)
    return value


def _date_range(start, end):
    if not start or not end:
        return "N/A"
    return f"{_display_datetime(start)} to {_display_datetime(end)}"


def _duration(start, end):
    if not start or not end:
        return "N/A"
    parsed_start = _datetime_cell(start)
    parsed_end = _datetime_cell(end)
    if isinstance(parsed_start, datetime) and isinstance(parsed_end, datetime):
        delta = parsed_end - parsed_start
        hours = delta.total_seconds() / 3600
        return round(hours, 2)
    return f"{start} to {end}"


def _improvement(baseline_value, model_value):
    if baseline_value in (None, "", 0) or model_value in (None, ""):
        return "N/A"
    baseline = float(baseline_value)
    if baseline == 0:
        return "N/A"
    return round(((baseline - float(model_value)) / baseline) * 100, 2)


def _strip_tz(value):
    if value.tzinfo:
        return value.replace(tzinfo=None)
    return value


def _display_datetime(value):
    parsed = _datetime_cell(value)
    if isinstance(parsed, datetime):
        return parsed.strftime("%Y-%m-%d %H:%M")
    return parsed


def _apply_number_formats(worksheet):
    datetime_headers = {"Start time", "End time", "Gap start time", "Gap end time", "Timestamp", "Parsed start", "Parsed end", "Interval start", "Local timestamp", "UTC timestamp"}
    one_decimal_headers = {"Radon", "Temperature", "Humidity", "Pressure", "Mean radon", "Max radon", "Actual radon", "Predicted radon", "Absolute error", "Radon mean", "Radon median", "Radon min", "Radon max", "Radon std", "Temperature mean", "Humidity mean", "Pressure mean"}
    three_decimal_headers = {"Baseline MAE", "Model MAE", "Baseline RMSE", "Model RMSE", "MAE", "RMSE", "R2", "Completeness ratio", "Score", "Skill score vs persistence"}
    percent_headers = {"MAE improvement %", "RMSE improvement %"}
    duration_headers = {"Duration", "Duration minutes"}

    headers = {cell.column: cell.value for cell in worksheet[1]}
    for column_index, header in headers.items():
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                if header in datetime_headers and isinstance(item.value, datetime):
                    item.number_format = "yyyy-mm-dd hh:mm"
                elif header in one_decimal_headers and isinstance(item.value, (int, float)):
                    item.number_format = "0.0"
                elif header in three_decimal_headers and isinstance(item.value, (int, float)):
                    item.number_format = "0.000"
                elif header in percent_headers and isinstance(item.value, (int, float)):
                    item.number_format = "0.00"
                elif header in duration_headers and isinstance(item.value, (int, float)):
                    item.number_format = "0.00"


def _wrap_long_text(worksheet):
    wrap_headers = {
        "Interpretation",
        "Reason/source",
        "Warnings/errors",
        "Detected overlap information",
        "Mapped columns",
        "Uploaded file name",
        "Regime/Label",
        "Prediction Insights",
        "Detected columns",
        "Environmental columns",
        "Missing values",
        "Interval distribution",
        "Quality flags",
        "Description",
        "Values",
        "Note",
        "Notes",
        "Regime counts",
        "Regime durations",
        "Most sensitive regimes",
        "Explanation",
        "Score reduction flags",
        "Value",
    }
    headers = {cell.column: cell.value for cell in worksheet[1]}
    for column_index, header in headers.items():
        if header in wrap_headers:
            for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                for item in cell:
                    item.alignment = Alignment(vertical="top", wrap_text=True)


def _max_width(sheet_name, column_index):
    if sheet_name in {"Ingestion Diagnostics", "Source File Inventory", "Overlap Conflicts", "Reproducibility Config"}:
        return 60
    if sheet_name == "Measurements" and column_index in (1, 6):
        return 28
    if sheet_name == "Segments" and column_index == 9:
        return 70
    if sheet_name in {"Prediction Insights", "Quality Flag Dictionary", "SIREM Readiness"}:
        return 100
    return 45
