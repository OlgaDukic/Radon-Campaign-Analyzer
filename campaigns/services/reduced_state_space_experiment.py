import csv
import json
from collections import defaultdict
from datetime import timedelta
from io import BytesIO, StringIO
from math import exp, isfinite, log, sqrt
from statistics import median

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook

from .baseline_prediction_experiment import run_baseline_prediction_experiment
from .documented_events import PHASES, default_event_cycles_for_campaign


VERSION = "reduced_state_space_experiment_v1"
HORIZONS = {"1h": 1.0, "3h": 3.0, "6h": 6.0}
FORECAST_MODES = ["F1_no_future_event_knowledge", "F2_known_intervention_scenario"]
R_SCENARIOS = {
    "R_low": {"observation_variance": 25.0, "description": "Low observation-noise scenario, not a RadonEye calibration claim."},
    "R_nominal": {"observation_variance": 100.0, "description": "Primary transparent nominal scenario."},
    "R_high": {"observation_variance": 400.0, "description": "High observation-noise sensitivity scenario."},
}
Q_GRID = [0.1, 1.0, 5.0, 10.0, 25.0, 50.0, 100.0, 250.0]
PRIMARY_R_SCENARIO = "R_nominal"
BASELINE_MODELS = ["persistence", "rolling_mean_3", "ridge_autoregressive"]
SMALL_PHASE_N = 5


def run_reduced_state_space_experiment(campaign):
    definitions = default_event_cycles_for_campaign(campaign)
    measurements = _measurement_rows(campaign)
    interval = _median_interval(measurements) or 1.0
    experiments = []
    if len(definitions) >= 2:
        experiments = [
            _run_experiment("Experiment A", definitions[0], definitions[1], measurements, interval),
            _run_experiment("Experiment B", definitions[1], definitions[0], measurements, interval),
        ]
    payload = {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "manifest": _manifest(campaign, definitions, measurements, interval),
        "experiments": experiments,
    }
    payload["forecast_rows"] = flatten_forecast_rows(payload)
    payload["parameter_audit"] = flatten_parameter_audit(payload)
    payload["q_selection_audit"] = flatten_q_selection_audit(payload)
    payload["overall_performance"] = _overall_performance(payload["forecast_rows"])
    payload["phase_specific_performance"] = _phase_specific_performance(payload["forecast_rows"])
    payload["uncertainty_performance"] = _uncertainty_performance(payload["forecast_rows"])
    payload["intervention_response"] = flatten_intervention_response(payload)
    payload["model_validity_flags"] = _model_validity_flags(payload["forecast_rows"], payload["parameter_audit"])
    payload["duplicate_forecast_key_count"] = _duplicate_forecast_key_count(payload["forecast_rows"])
    payload["baseline_reference"] = _baseline_reference(campaign)
    payload["fair_comparison"] = _fair_comparison(payload["forecast_rows"], payload["baseline_reference"])
    return payload


def flatten_forecast_rows(payload):
    return [row for experiment in payload["experiments"] for row in experiment["forecast_rows"]]


def flatten_parameter_audit(payload):
    return [row for experiment in payload["experiments"] for row in experiment["parameter_audit"]]


def flatten_q_selection_audit(payload):
    return [row for experiment in payload["experiments"] for row in experiment["q_selection_audit"]]


def flatten_intervention_response(payload):
    return [row for experiment in payload["experiments"] for row in experiment["intervention_response"]]


def build_reduced_state_space_csv(payload):
    return _csv(payload["overall_performance"])


def build_reduced_state_space_workbook(payload):
    workbook = Workbook()
    workbook.active.title = "Manifest"
    _write_rows(workbook["Manifest"], [{"field": key, "value": value} for key, value in payload["manifest"].items()])
    _write_rows(workbook.create_sheet("Overall Performance"), payload["overall_performance"])
    _write_rows(workbook.create_sheet("Fair Comparison"), payload["fair_comparison"])
    _write_rows(workbook.create_sheet("Phase Performance"), payload["phase_specific_performance"])
    _write_rows(workbook.create_sheet("Uncertainty"), payload["uncertainty_performance"])
    _write_rows(workbook.create_sheet("Intervention Response"), payload["intervention_response"])
    _write_rows(workbook.create_sheet("Parameter Audit"), payload["parameter_audit"])
    _write_rows(workbook.create_sheet("Q Selection Audit"), payload["q_selection_audit"])
    _write_rows(workbook.create_sheet("Forecast Rows"), payload["forecast_rows"])
    _write_rows(workbook.create_sheet("Validity Flags"), payload["model_validity_flags"])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 62)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_reduced_state_space_paper_outputs(campaign, output_dir=None):
    payload = run_reduced_state_space_experiment(campaign)
    base = output_dir or settings.BASE_DIR / "paper_outputs" / f"campaign_{campaign.pk}" / "reduced_state_space_experiment"
    base.mkdir(parents=True, exist_ok=True)
    (base / "reduced_state_space_experiment.json").write_text(json.dumps(payload, indent=2, ensure_ascii=True, default=str), encoding="utf-8")
    (base / "reduced_state_space_experiment.xlsx").write_bytes(build_reduced_state_space_workbook(payload).getvalue())
    exports = {
        "overall_performance.csv": payload["overall_performance"],
        "fair_comparison.csv": payload["fair_comparison"],
        "phase_specific_performance.csv": payload["phase_specific_performance"],
        "uncertainty_performance.csv": payload["uncertainty_performance"],
        "intervention_response.csv": payload["intervention_response"],
        "parameter_audit.csv": payload["parameter_audit"],
        "q_selection_audit.csv": payload["q_selection_audit"],
        "forecast_rows.csv": payload["forecast_rows"],
        "model_validity_flags.csv": payload["model_validity_flags"],
    }
    for filename, rows in exports.items():
        (base / filename).write_text(_csv(rows), encoding="utf-8")
    (base / "reduced_state_space_summary.md").write_text(_summary_markdown(payload), encoding="utf-8")
    for name, svg in _svg_outputs(payload).items():
        (base / name).write_text(svg, encoding="utf-8")
    return payload, base


def _run_experiment(label, train_cycle, test_cycle, measurements, interval):
    train_rows = _rows_in_cycle(measurements, train_cycle)
    test_rows = _rows_in_cycle(measurements, test_cycle)
    parameter_audit, parameters = _estimate_parameters(label, train_cycle, train_rows, interval)
    q_audit, selected_q = _select_q(label, train_cycle, train_rows, parameters, interval)
    forecast_rows = []
    state_trace = []
    for r_name, r_config in R_SCENARIOS.items():
        for mode in FORECAST_MODES:
            rows, trace = _filter_and_forecast(
                label,
                train_cycle,
                test_cycle,
                test_rows,
                parameters,
                selected_q,
                r_name,
                r_config["observation_variance"],
                mode,
                interval,
            )
            forecast_rows.extend(rows)
            state_trace.extend(trace)
    return {
        "label": label,
        "training_cycle": _cycle_payload(train_cycle),
        "test_cycle": _cycle_payload(test_cycle),
        "regime_parameters": parameters,
        "selected_q": selected_q,
        "parameter_audit": parameter_audit,
        "q_selection_audit": q_audit,
        "forecast_rows": forecast_rows,
        "state_trace": state_trace,
        "intervention_response": _intervention_response_rows(label, test_cycle, forecast_rows),
    }


def _estimate_parameters(label, cycle, rows, interval):
    groups = {"closed_reference": [], "rapid_removal": []}
    for prev, curr in zip(rows, rows[1:]):
        phase = _phase_for_time(cycle, prev["measured_at"])
        next_phase = _phase_for_time(cycle, curr["measured_at"])
        if not phase or phase != next_phase:
            continue
        regime = _regime_for_phase(phase)
        groups[regime].append((float(prev["radon_bq_m3"]), float(curr["radon_bq_m3"]), _hours(prev["measured_at"], curr["measured_at"])))
    audit = []
    parameters = {}
    for regime, pairs in groups.items():
        fit = _fit_ab_for_pairs(pairs, interval)
        flags = []
        if fit["sample_count"] < SMALL_PHASE_N:
            flags.append("SMALL_PHASE_SAMPLE")
        if fit["a"] <= 0.01 or fit["a"] >= 0.99:
            flags.append("PARAMETER_NEAR_BOUNDARY")
        if not isfinite(fit["a"]) or not isfinite(fit["b"]):
            flags.append("NUMERICAL_WARNING")
        fit["flags"] = "|".join(flags or ["FIT_OK"])
        parameters[regime] = fit
        audit.append(
            {
                "experiment": label,
                "training_cycle": cycle.cycle_label,
                "regime": regime,
                **fit,
                "estimation_rule": "constrained least squares on training cycle adjacent observations only",
            }
        )
    return audit, parameters


def _fit_ab_for_pairs(pairs, interval):
    if not pairs:
        return _parameter_payload(1.0, 0.0, 0, interval)
    xs = [pair[0] for pair in pairs]
    ys = [pair[1] for pair in pairs]
    x_mean = sum(xs) / len(xs)
    y_mean = sum(ys) / len(ys)
    denom = sum((x - x_mean) ** 2 for x in xs)
    a = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys)) / denom if denom else 1.0
    a = min(max(a, 0.0), 1.0)
    b = max(y_mean - a * x_mean, 0.0)
    return _parameter_payload(a, b, len(pairs), interval)


def _parameter_payload(a, b, sample_count, interval):
    if a <= 0:
        kappa = None
    elif a >= 1:
        kappa = 0.0
    else:
        kappa = -log(a) / interval
    ceq = b / (1 - a) if a < 0.999999 else None
    return {
        "a": _round(a, 6),
        "b": _round(b, 6),
        "sample_count": sample_count,
        "kappa_h_minus_1": _round(kappa, 6) if kappa is not None else None,
        "C_eq_bq_m3": _round(ceq, 6) if ceq is not None else None,
        "identifiability_flag": "C_EQ_NOT_STABLY_DEFINED" if ceq is None else "IDENTIFIABLE_REDUCED_FORM",
    }


def _select_q(label, cycle, rows, parameters, interval):
    if len(rows) < 8:
        q = Q_GRID[len(Q_GRID) // 2]
        return [
            {
                "experiment": label,
                "training_cycle": cycle.cycle_label,
                "q_candidate": q,
                "validation_metric_rmse": None,
                "selected": True,
                "validation_rule": "insufficient rows for blocked validation; deterministic middle-grid fallback",
            }
        ], q
    split = max(3, int(len(rows) * 0.7))
    validation = rows[split:]
    audit = []
    best = (None, None)
    for q in Q_GRID:
        errors = _validation_errors(cycle, rows[:split], validation, parameters, q, R_SCENARIOS[PRIMARY_R_SCENARIO]["observation_variance"], interval)
        rmse = sqrt(sum(error * error for error in errors) / len(errors)) if errors else float("inf")
        if best[0] is None or rmse < best[0]:
            best = (rmse, q)
        audit.append(
            {
                "experiment": label,
                "training_cycle": cycle.cycle_label,
                "q_candidate": q,
                "validation_metric_rmse": _round(rmse),
                "selected": False,
                "validation_start": _iso(validation[0]["measured_at"]) if validation else None,
                "validation_end": _iso(validation[-1]["measured_at"]) if validation else None,
                "validation_rule": "blocked chronological validation inside training cycle only",
            }
        )
    selected = best[1]
    for row in audit:
        row["selected"] = row["q_candidate"] == selected
    return audit, selected


def _validation_errors(cycle, warmup, validation, parameters, q, r, interval):
    if not warmup or not validation:
        return []
    state = float(warmup[-1]["radon_bq_m3"])
    variance = max(r, 1.0)
    last_time = warmup[-1]["measured_at"]
    errors = []
    for row in validation:
        phase = _phase_for_time(cycle, last_time) or _phase_for_time(cycle, row["measured_at"])
        regime = _regime_for_phase(phase)
        dt = max(_hours(last_time, row["measured_at"]), interval)
        prior, prior_var = _propagate(state, variance, parameters[regime], q, dt, interval)
        errors.append(prior - float(row["radon_bq_m3"]))
        state, variance = kalman_update(prior, prior_var, float(row["radon_bq_m3"]), r)
        last_time = row["measured_at"]
    return errors


def _filter_and_forecast(label, train_cycle, test_cycle, rows, parameters, q, r_name, r_value, mode, interval):
    if not rows:
        return [], []
    state = float(rows[0]["radon_bq_m3"])
    variance = r_value
    last_time = rows[0]["measured_at"]
    by_time = {row["measured_at"]: row for row in rows}
    forecasts = []
    trace = []
    for index, row in enumerate(rows):
        current_time = row["measured_at"]
        if index == 0:
            prior, prior_var = state, variance
        else:
            phase_for_transition = _phase_for_time(test_cycle, last_time)
            regime = _regime_for_phase(phase_for_transition)
            dt = _hours(last_time, current_time)
            prior, prior_var = _propagate(state, variance, parameters[regime], q, dt, interval)
        observed = float(row["radon_bq_m3"])
        posterior, posterior_var = kalman_update(prior, prior_var, observed, r_value)
        innovation = observed - prior
        current_phase = _phase_for_time(test_cycle, current_time)
        trace.append(
            {
                "experiment": label,
                "mode": mode,
                "R_scenario": r_name,
                "timestamp": _iso(current_time),
                "phase": current_phase,
                "observed": _round(observed),
                "prior_state": _round(prior),
                "posterior_state": _round(posterior),
                "prior_variance": _round(prior_var),
                "posterior_variance": _round(posterior_var),
                "innovation": _round(innovation),
            }
        )
        for horizon_label, horizon_hours in HORIZONS.items():
            target_time = current_time + timedelta(hours=horizon_hours)
            target = by_time.get(target_time)
            valid = target is not None and _phase_for_time(test_cycle, target_time) is not None
            target_phase = _phase_for_time(test_cycle, target_time)
            mean, latent_var = _forecast_from_origin(posterior, posterior_var, current_time, target_time, test_cycle, parameters, q, mode, interval)
            observation_var = latent_var + r_value
            latent_lower = mean - 1.96 * sqrt(max(latent_var, 0.0))
            latent_upper = mean + 1.96 * sqrt(max(latent_var, 0.0))
            observation_lower = mean - 1.96 * sqrt(max(observation_var, 0.0))
            observation_upper = mean + 1.96 * sqrt(max(observation_var, 0.0))
            actual = float(target["radon_bq_m3"]) if target else None
            flags, status = _validity_flags(row, current_phase, target_phase, innovation, posterior_var, interval, target)
            forecasts.append(
                {
                    "experiment": label,
                    "training_cycle": train_cycle.cycle_label,
                    "test_cycle": test_cycle.cycle_label,
                    "forecast_mode": mode,
                    "R_scenario": r_name,
                    "Q_selected": q,
                    "horizon": horizon_label,
                    "forecast_origin": _iso(current_time),
                    "target_timestamp": _iso(target_time),
                    "observed_value": _round(actual) if actual is not None else None,
                    "filtered_posterior_at_origin": _round(posterior),
                    "forecast_mean": _round(mean),
                    "forecast_variance": _round(latent_var),
                    "latent_state_variance": _round(latent_var),
                    "future_observation_variance": _round(observation_var),
                    "latent_state_lower_interval": _round(latent_lower),
                    "latent_state_upper_interval": _round(latent_upper),
                    "future_observation_lower_interval": _round(observation_lower),
                    "future_observation_upper_interval": _round(observation_upper),
                    "lower_interval": _round(observation_lower),
                    "upper_interval": _round(observation_upper),
                    "interval_semantics": "future_observation_predictive_interval_includes_R",
                    "origin_phase": current_phase,
                    "target_phase": target_phase,
                    "event_knowledge_status": _event_knowledge_status(mode, current_time, target_time, test_cycle),
                    "valid_status": "valid" if valid else "excluded",
                    "exclusion_reason": None if valid else "missing_target_or_outside_documented_cycle",
                    "error": _round(mean - actual) if actual is not None else None,
                    "absolute_error": _round(abs(mean - actual)) if actual is not None else None,
                    "latent_state_covered_by_95_interval": (latent_lower <= actual <= latent_upper) if actual is not None else None,
                    "covered_by_95_interval": (observation_lower <= actual <= observation_upper) if actual is not None else None,
                    "future_observation_covered_by_95_interval": (observation_lower <= actual <= observation_upper) if actual is not None else None,
                    "validity_status": status,
                    "validity_flags": "|".join(flags),
                }
            )
        state, variance = posterior, posterior_var
        last_time = current_time
    return forecasts, trace


def kalman_update(prior_mean, prior_variance, observation, observation_variance):
    gain = prior_variance / (prior_variance + observation_variance) if (prior_variance + observation_variance) else 0.0
    posterior = prior_mean + gain * (observation - prior_mean)
    posterior_variance = (1 - gain) * prior_variance
    return posterior, max(posterior_variance, 1e-9)


def _forecast_from_origin(state, variance, origin, target, cycle, parameters, q, mode, interval):
    current = origin
    mean = state
    var = variance
    while current < target:
        step = min(interval, _hours(current, target))
        phase = _forecast_phase(cycle, origin, current, target, mode)
        regime = _regime_for_phase(phase)
        mean, var = _propagate(mean, var, parameters[regime], q, step, interval)
        current = current + timedelta(hours=step)
    return mean, var


def _forecast_phase(cycle, origin, current, target, mode):
    if mode == "F2_known_intervention_scenario" and origin <= cycle.rapid_removal_start <= target and current >= cycle.rapid_removal_start:
        return "rapid_removal"
    phase = _phase_for_time(cycle, current)
    if mode == "F1_no_future_event_knowledge" and origin < cycle.rapid_removal_start <= target:
        return "accumulation"
    return phase


def _propagate(mean, variance, parameter, q, dt, interval):
    a = float(parameter["a"])
    b = float(parameter["b"])
    power = dt / interval if interval else dt
    a_dt = a ** power if a >= 0 else 0.0
    if abs(1 - a) < 1e-9:
        b_dt = b * power
    else:
        b_dt = b * (1 - a_dt) / (1 - a)
    return a_dt * mean + b_dt, (a_dt * a_dt) * variance + q * max(dt / interval, 0.0)


def _overall_performance(rows):
    return _metric_rows([row for row in rows if row["valid_status"] == "valid"], ["experiment", "forecast_mode", "R_scenario", "horizon"])


def _phase_specific_performance(rows):
    return _metric_rows([row for row in rows if row["valid_status"] == "valid"], ["experiment", "forecast_mode", "R_scenario", "horizon", "target_phase"])


def _uncertainty_performance(rows):
    valid = [row for row in rows if row["valid_status"] == "valid"]
    grouped = defaultdict(list)
    for row in valid:
        grouped[(row["experiment"], row["forecast_mode"], row["R_scenario"], row["horizon"], row["target_phase"])].append(row)
    output = []
    for key, values in sorted(grouped.items()):
        output.append({**dict(zip(["experiment", "forecast_mode", "R_scenario", "horizon", "target_phase"], key)), **_interval_metrics(values)})
    return output


def _metric_rows(rows, group_keys):
    grouped = defaultdict(list)
    for row in rows:
        grouped[tuple(row[key] for key in group_keys)].append(row)
    output = []
    for key, values in sorted(grouped.items()):
        errors = [row["error"] for row in values if row["error"] is not None]
        abs_errors = [abs(error) for error in errors]
        output.append(
            {
                **dict(zip(group_keys, key)),
                "N": len(errors),
                "MAE": _round(sum(abs_errors) / len(abs_errors)) if abs_errors else None,
                "RMSE": _round(sqrt(sum(error * error for error in errors) / len(errors))) if errors else None,
                "bias": _round(sum(errors) / len(errors)) if errors else None,
                "median_AE": _round(median(abs_errors)) if abs_errors else None,
                "maximum_AE": _round(max(abs_errors)) if abs_errors else None,
                **_interval_metrics(values),
            }
        )
    return output


def _interval_metrics(rows):
    valid = [row for row in rows if row.get("observed_value") is not None]
    if not valid:
        return {"empirical_coverage": None, "mean_interval_width": None, "interval_score": None}
    alpha = 0.05
    widths = [row["future_observation_upper_interval"] - row["future_observation_lower_interval"] for row in valid]
    covered = [bool(row["future_observation_covered_by_95_interval"]) for row in valid]
    scores = []
    for row, width in zip(valid, widths):
        y = row["observed_value"]
        lower = row["future_observation_lower_interval"]
        upper = row["future_observation_upper_interval"]
        penalty = (2 / alpha) * (lower - y) if y < lower else (2 / alpha) * (y - upper) if y > upper else 0
        scores.append(width + penalty)
    return {
        "empirical_coverage": _round(sum(covered) / len(covered)),
        "mean_interval_width": _round(sum(widths) / len(widths)),
        "interval_score": _round(sum(scores) / len(scores)),
    }


def _baseline_reference(campaign):
    try:
        payload = run_baseline_prediction_experiment(campaign)
    except Exception as exc:  # pragma: no cover - defensive export path
        return {"available": False, "error": str(exc), "forecast_rows": []}
    return {"available": True, "algorithm_version": payload.get("algorithm_version"), "forecast_rows": [row for exp in payload.get("experiments", []) for row in exp.get("forecast_rows", [])]}


def _fair_comparison(state_rows, baseline_reference):
    if not baseline_reference.get("available"):
        return [{"note": "Locked baseline reference unavailable.", "error": baseline_reference.get("error")}]
    baseline_rows = baseline_reference["forecast_rows"]
    rows = []
    valid_state = [row for row in state_rows if row["valid_status"] == "valid" and row["R_scenario"] == PRIMARY_R_SCENARIO]
    for experiment in sorted({row["experiment"] for row in valid_state}):
        for mode in FORECAST_MODES:
            for horizon in HORIZONS:
                ss = [row for row in valid_state if row["experiment"] == experiment and row["forecast_mode"] == mode and row["horizon"] == horizon]
                state_by_key = {(row["forecast_origin"], row["target_timestamp"]): row for row in ss}
                baseline_by_model = {
                    model: {
                        (row["forecast_origin"], row["target_timestamp"]): row
                        for row in baseline_rows
                        if row["experiment"] == experiment and row["horizon"] == horizon and row["model"] == model
                    }
                    for model in BASELINE_MODELS
                }
                common = set(state_by_key)
                for values in baseline_by_model.values():
                    common &= set(values)
                state_metrics = _simple_metrics([state_by_key[key]["observed_value"] for key in common], [state_by_key[key]["forecast_mean"] for key in common])
                rows.append({"experiment": experiment, "forecast_mode": mode, "horizon": horizon, "model": "reduced_state_space", "common_target_count": len(common), **state_metrics})
                for model, model_rows in baseline_by_model.items():
                    metrics = _simple_metrics([model_rows[key]["actual"] for key in common], [model_rows[key]["predicted"] for key in common])
                    rows.append({"experiment": experiment, "forecast_mode": mode, "horizon": horizon, "model": model, "common_target_count": len(common), **metrics})
    return rows


def _simple_metrics(actual, predicted):
    if not actual:
        return {"MAE": None, "RMSE": None, "bias": None}
    errors = [p - a for a, p in zip(actual, predicted)]
    abs_errors = [abs(error) for error in errors]
    return {"MAE": _round(sum(abs_errors) / len(abs_errors)), "RMSE": _round(sqrt(sum(error * error for error in errors) / len(errors))), "bias": _round(sum(errors) / len(errors))}


def _intervention_response_rows(label, test_cycle, forecast_rows):
    rows = []
    for mode in FORECAST_MODES:
        for horizon in HORIZONS:
            candidates = sorted(
                [
                    row for row in forecast_rows
                    if row["forecast_mode"] == mode and row["R_scenario"] == PRIMARY_R_SCENARIO and row["horizon"] == horizon and row["valid_status"] == "valid"
                ],
                key=lambda row: row["target_timestamp"],
            )
            rapid = [row for row in candidates if row["target_phase"] == "rapid_removal"]
            pre_event = max([row for row in candidates if _parse_iso(row["forecast_origin"]) < test_cycle.rapid_removal_start], key=lambda row: row["forecast_origin"], default=None)
            first_rapid = rapid[0] if rapid else None
            max_rapid = max(rapid, key=lambda row: row["absolute_error"], default=None)
            rows.append(
                {
                    "experiment": label,
                    "forecast_mode": mode,
                    "horizon": horizon,
                    "pre_event_forecast_origin": pre_event["forecast_origin"] if pre_event else None,
                    "pre_event_forecast_error": pre_event["error"] if pre_event else None,
                    "first_rapid_removal_target_error": first_rapid["error"] if first_rapid else None,
                    "maximum_rapid_removal_error": max_rapid["absolute_error"] if max_rapid else None,
                    "timestamp_of_maximum_error": max_rapid["target_timestamp"] if max_rapid else None,
                    "adaptation_after_first_new_observation": _adaptation_error(candidates, test_cycle.rapid_removal_start, 1),
                    "adaptation_after_second_new_observation": _adaptation_error(candidates, test_cycle.rapid_removal_start, 2),
                    "adaptation_after_third_new_observation": _adaptation_error(candidates, test_cycle.rapid_removal_start, 3),
                    "interpretation": "F2 is a known-intervention scenario, not autonomous event prediction." if mode.startswith("F2") else "F1 has no future event knowledge before observations arrive.",
                }
            )
    return rows


def _adaptation_error(rows, rapid_start, observations_seen):
    target = rapid_start + timedelta(hours=observations_seen - 1)
    matches = [row for row in rows if _parse_iso(row["forecast_origin"]) >= target]
    return matches[0]["error"] if matches else None


def _model_validity_flags(forecast_rows, parameter_rows):
    rows = []
    for row in forecast_rows:
        if row["validity_flags"] and row["validity_flags"] != "FIT_OK":
            rows.append(
                {
                    "experiment": row["experiment"],
                    "forecast_mode": row["forecast_mode"],
                    "R_scenario": row["R_scenario"],
                    "horizon": row["horizon"],
                    "forecast_origin": row["forecast_origin"],
                    "target_timestamp": row["target_timestamp"],
                    "status": row["validity_status"],
                    "flags": row["validity_flags"],
                }
            )
    for row in parameter_rows:
        if row.get("flags") and row["flags"] != "FIT_OK":
            rows.append({"experiment": row["experiment"], "status": "caution", "flags": row["flags"], "source": f"parameter:{row['regime']}"})
    return rows


def _validity_flags(row, current_phase, target_phase, innovation, posterior_var, interval, target):
    flags = ["PROVISIONAL_EVENT_BOUNDARY"]
    if target is None:
        flags.append("MISSING_OBSERVATION")
    if current_phase is None or target_phase is None:
        flags.append("OUTSIDE_TRAINING_REGIME")
    if abs(innovation) > 50:
        flags.append("LARGE_INNOVATION")
    if posterior_var > 250:
        flags.append("HIGH_PREDICTIVE_UNCERTAINTY")
    if row.get("previous_interval_hours") and row["previous_interval_hours"] > interval * 1.5:
        flags.append("RECENT_GAP")
    status = "valid" if flags == ["PROVISIONAL_EVENT_BOUNDARY"] else "caution"
    if "MISSING_OBSERVATION" in flags or "OUTSIDE_TRAINING_REGIME" in flags:
        status = "unreliable"
    return flags, status


def _duplicate_forecast_key_count(rows):
    keys = [(row["experiment"], row["forecast_mode"], row["R_scenario"], row["horizon"], row["forecast_origin"], row["target_timestamp"]) for row in rows]
    return len(keys) - len(set(keys))


def _measurement_rows(campaign):
    rows = list(campaign.measurements.exclude(radon_bq_m3=None).order_by("measured_at", "id").values("measured_at", "radon_bq_m3", "regime"))
    previous = None
    for row in rows:
        row["previous_interval_hours"] = _hours(previous["measured_at"], row["measured_at"]) if previous else None
        previous = row
    return rows


def _rows_in_cycle(rows, cycle):
    start, end = _cycle_window(cycle)
    return [row for row in rows if start <= row["measured_at"] <= end]


def _cycle_window(cycle):
    starts = [getattr(cycle, f"{phase}_start") for phase in PHASES]
    ends = [getattr(cycle, f"{phase}_end") for phase in PHASES]
    return min(starts), max(ends)


def _phase_for_time(cycle, timestamp):
    if not timestamp:
        return None
    for phase in PHASES:
        if getattr(cycle, f"{phase}_start") <= timestamp <= getattr(cycle, f"{phase}_end"):
            return phase
    return None


def _regime_for_phase(phase):
    return "rapid_removal" if phase == "rapid_removal" else "closed_reference"


def _event_knowledge_status(mode, origin, target, cycle):
    if mode.startswith("F2") and origin <= cycle.rapid_removal_start <= target:
        return "known_intervention_transition_used_after_declared_start"
    if mode.startswith("F1") and origin < cycle.rapid_removal_start <= target:
        return "no_future_event_knowledge_closed_transition_used"
    return "current_phase_transition"


def _manifest(campaign, definitions, measurements, interval):
    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "algorithm_version": VERSION,
        "model_equations": "C_(k+1)=a_r*C_k+b_r+w_k; y_k=C_k+v_k",
        "state_vector": "x_k=[C_k]",
        "sirem_boundary": "Reduced one-zone apparent dynamics only; no ACH, beta, lambda_v, C_bm, C_out or exhalation-rate estimation.",
        "continuous_mapping": "a_r=exp(-kappa_r*delta_t); b_r=(1-a_r)*C_eq,r; kappa is apparent and is not ACH.",
        "training_rule": "Experiment A trains on Cycle 1 and tests on Cycle 2; Experiment B reverses direction. Test cycles are not used for parameters, Q, R or tuning.",
        "q_grid": Q_GRID,
        "q_selection_rule": "training-cycle blocked chronological validation only",
        "r_scenarios": R_SCENARIOS,
        "primary_r_scenario": PRIMARY_R_SCENARIO,
        "forecast_modes": FORECAST_MODES,
        "horizons": list(HORIZONS),
        "interval_hours": interval,
        "prediction_interval": "nominal 95% future-observation predictive interval; latent-state interval is also exported separately",
        "interval_definitions": {
            "latent_state_uncertainty_interval": "propagated latent-state variance only",
            "future_observation_predictive_interval": "propagated latent-state variance plus observation noise R",
        },
        "quality_handling": "Valid stored radon observations are used; missing targets are excluded; gaps propagate uncertainty without interpolation.",
        "cycle_windows": [_cycle_payload(cycle) for cycle in definitions],
        "timestamp_warning": "Stored timestamps are provisional and are not confirmed local window-opening times.",
        "generated_at": timezone.now().replace(microsecond=0).isoformat(),
    }


def _cycle_payload(cycle):
    payload = {"cycle_label": cycle.cycle_label, "evidence_status": cycle.evidence_status, "note": cycle.note}
    for phase in PHASES:
        payload[f"{phase}_start"] = _iso(getattr(cycle, f"{phase}_start"))
        payload[f"{phase}_end"] = _iso(getattr(cycle, f"{phase}_end"))
    return payload


def _summary_markdown(payload):
    lines = [
        "# Reduced State-Space Experiment Summary",
        "",
        f"Campaign: {payload['campaign_id']} - {payload['campaign_name']}",
        f"Algorithm version: {payload['algorithm_version']}",
        "",
        "This is exploratory research software, not a certified radon risk assessment or ventilation-control tool.",
        "",
        "## Model",
        "`C_(k+1)=a_r*C_k+b_r+w_k`, `y_k=C_k+v_k`; state vector `x_k=[C_k]`.",
        "The model estimates reduced apparent dynamics only and does not estimate ACH or material exhalation.",
        "",
        "## Key rows",
    ]
    for row in payload["overall_performance"][:12]:
        lines.append(f"- {row['experiment']} {row['forecast_mode']} {row['R_scenario']} {row['horizon']}: N={row['N']}, MAE={row['MAE']}, coverage={row['empirical_coverage']}")
    lines.extend(["", f"Duplicate forecast keys: {payload['duplicate_forecast_key_count']}"])
    return "\n".join(lines) + "\n"


def _svg_outputs(payload):
    primary = [row for row in payload["forecast_rows"] if row["R_scenario"] == PRIMARY_R_SCENARIO and row["horizon"] == "1h" and row["valid_status"] == "valid"]
    return {
        "observed_filtered_predicted_series.svg": _simple_svg(primary, "Observed, filtered and predicted series"),
        "f1_forecast_intervals.svg": _simple_svg([row for row in primary if row["forecast_mode"].startswith("F1")], "F1 forecast with intervals"),
        "f2_known_intervention_forecast_intervals.svg": _simple_svg([row for row in primary if row["forecast_mode"].startswith("F2")], "F2 known-intervention forecast with intervals"),
        "rapid_removal_f1_vs_f2.svg": _simple_svg([row for row in primary if row["target_phase"] == "rapid_removal"], "Rapid-removal F1 vs F2"),
        "innovations_over_time.svg": _simple_svg(primary, "Forecast errors through time"),
        "coverage_interval_width_by_phase.svg": _bar_svg(payload["uncertainty_performance"], "Coverage and interval width by phase"),
    }


def _simple_svg(rows, title):
    width, height, pad = 760, 260, 32
    rows = rows[:180]
    values = [row["observed_value"] for row in rows if row.get("observed_value") is not None] + [row["forecast_mean"] for row in rows if row.get("forecast_mean") is not None]
    if not values:
        values = [0, 1]
    lo, hi = min(values), max(values)
    span = max(hi - lo, 1)
    def points(key):
        pts = []
        selected = [row for row in rows if row.get(key) is not None]
        for idx, row in enumerate(selected):
            x = pad + idx * (width - 2 * pad) / max(len(selected) - 1, 1)
            y = height - pad - ((row[key] - lo) / span) * (height - 2 * pad)
            pts.append(f"{x:.1f},{y:.1f}")
        return " ".join(pts)
    return f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}"><rect width="100%" height="100%" fill="#fff"/><text x="24" y="24" font-size="16" font-family="Arial">{title}</text><polyline fill="none" stroke="#1f77b4" stroke-width="2" points="{points("observed_value")}"/><polyline fill="none" stroke="#2ca02c" stroke-width="2" points="{points("forecast_mean")}"/></svg>'


def _bar_svg(rows, title):
    rows = rows[:24]
    width, height = 760, 260
    bars = []
    for idx, row in enumerate(rows):
        value = float(row.get("empirical_coverage") or 0)
        x = 32 + idx * 28
        h = value * 180
        bars.append(f'<rect x="{x}" y="{220-h:.1f}" width="18" height="{h:.1f}" fill="#2f6fdf"/>')
    return f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}"><rect width="100%" height="100%" fill="#fff"/><text x="24" y="24" font-size="16" font-family="Arial">{title}</text>{"".join(bars)}</svg>'


def _csv(rows):
    output = StringIO()
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No rows."})
    return output.getvalue()


def _write_rows(sheet, rows):
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
    else:
        sheet.append(["No rows."])


def _excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value


def _median_interval(rows):
    intervals = [_hours(a["measured_at"], b["measured_at"]) for a, b in zip(rows, rows[1:]) if b["measured_at"] > a["measured_at"]]
    return _round(median(intervals), 6) if intervals else None


def _hours(start, end):
    return (end - start).total_seconds() / 3600 if start and end else 0.0


def _iso(value):
    return value.isoformat() if value else None


def _parse_iso(value):
    from django.utils.dateparse import parse_datetime

    return parse_datetime(value)


def _round(value, digits=3):
    return round(float(value), digits) if value is not None else None
