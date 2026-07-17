from collections import defaultdict
from datetime import timedelta
from io import BytesIO, StringIO
from math import sqrt
from statistics import median
import csv
import json

from django.utils import timezone
from openpyxl import Workbook

from .documented_events import default_event_cycles_for_campaign


VERSION = "baseline_prediction_experiment_v1"
HORIZONS = {"1h": timedelta(hours=1), "3h": timedelta(hours=3), "6h": timedelta(hours=6)}
MODELS = ["persistence", "rolling_mean_3", "ridge_autoregressive"]
ALPHA_CANDIDATES = [0.01, 0.1, 1.0, 10.0]
THRESHOLD_NOT_IDENTIFIABLE = "NOT_IDENTIFIABLE_FROM_CURRENT_DIRECT_FORECAST_OUTPUT"


def run_baseline_prediction_experiment(campaign, rolling_window=3):
    definitions = default_event_cycles_for_campaign(campaign)
    measurements = _measurement_rows(campaign)
    interval = _median_interval(measurements)
    experiments = []
    if len(definitions) >= 2:
        experiments = [
            _run_experiment("Experiment A", definitions[0], definitions[1], measurements, interval, rolling_window),
            _run_experiment("Experiment B", definitions[1], definitions[0], measurements, interval, rolling_window),
        ]
    payload = {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "manifest": _manifest(campaign, definitions, measurements, interval, rolling_window),
        "algorithm_version": VERSION,
        "models": MODELS,
        "horizons": list(HORIZONS),
        "lag_features": ["lag_1", "lag_2", "lag_3", "lag_6", "lag_12_when_available", "recent_first_difference", "rolling_mean_3", "rolling_mean_6"],
        "rolling_window": rolling_window,
        "alpha_candidates": ALPHA_CANDIDATES,
        "sampling_interval_hours": interval,
        "timezone_status": _timezone_status(measurements),
        "caution": (
            "Forecasts are unconditional on future occupant actions. Increased error at the onset of documented "
            "ventilation is expected and provides a reference for later intervention-aware physical modelling."
        ),
        "experiments": experiments,
    }
    payload["paper_ready_tables"] = _paper_ready_tables(payload)
    payload["leakage_audit"] = _leakage_audit(payload)
    payload["duplicate_forecast_key_count"] = _duplicate_forecast_key_count(flatten_forecasts(payload))
    return payload


def flatten_overall_metrics(payload):
    return [row for experiment in payload["experiments"] for row in experiment["overall_metrics"]]


def flatten_phase_metrics(payload):
    return [row for experiment in payload["experiments"] for row in experiment["phase_metrics"]]


def flatten_forecasts(payload):
    return [row for experiment in payload["experiments"] for row in experiment["forecast_rows"]]


def flatten_exclusions(payload):
    return [row for experiment in payload["experiments"] for row in experiment["exclusions"]]


def flatten_target_phase_metrics(payload):
    return [row for experiment in payload["experiments"] for row in experiment["target_phase_metrics"]]


def flatten_fair_comparison_metrics(payload):
    return [row for experiment in payload["experiments"] for row in experiment["fair_comparison_metrics"]]


def flatten_intervention_audit(payload):
    return [row for experiment in payload["experiments"] for row in experiment["intervention_response_audit"]]


def flatten_alpha_audit(payload):
    return [row for experiment in payload["experiments"] for row in experiment["alpha_selection_audit"]]


def build_baseline_prediction_csv(payload):
    output = StringIO()
    rows = flatten_overall_metrics(payload)
    headers = list(rows[0]) if rows else ["note"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"note": "No baseline prediction experiments available."})
    return output.getvalue()


def build_baseline_prediction_workbook(payload):
    workbook = Workbook()
    workbook.active.title = "Configuration"
    _write_rows(
        workbook["Configuration"],
        [
            {"field": "campaign_id", "value": payload["campaign_id"]},
            {"field": "campaign_name", "value": payload["campaign_name"]},
            {"field": "algorithm_version", "value": payload["algorithm_version"]},
            {"field": "manifest", "value": payload["manifest"]},
            {"field": "models", "value": ", ".join(payload["models"])},
            {"field": "horizons", "value": ", ".join(payload["horizons"])},
            {"field": "lag_features", "value": ", ".join(payload["lag_features"])},
            {"field": "alpha_candidates", "value": ", ".join(str(value) for value in payload["alpha_candidates"])},
            {"field": "sampling_interval_hours", "value": payload["sampling_interval_hours"]},
            {"field": "timezone_status", "value": payload["timezone_status"]},
            {"field": "caution", "value": payload["caution"]},
        ],
    )
    _write_rows(workbook.create_sheet("Overall Metrics"), flatten_overall_metrics(payload))
    _write_rows(workbook.create_sheet("Phase Metrics"), flatten_phase_metrics(payload))
    _write_rows(workbook.create_sheet("Target Phase Metrics"), flatten_target_phase_metrics(payload))
    _write_rows(workbook.create_sheet("Fair Comparison"), flatten_fair_comparison_metrics(payload))
    _write_rows(workbook.create_sheet("Intervention Audit"), flatten_intervention_audit(payload))
    _write_rows(workbook.create_sheet("Alpha Audit"), flatten_alpha_audit(payload))
    _write_rows(workbook.create_sheet("Table A Overall"), payload["paper_ready_tables"]["table_a_overall_predictive_performance"])
    _write_rows(workbook.create_sheet("Table B Fair"), payload["paper_ready_tables"]["table_b_fair_comparison_performance"])
    _write_rows(workbook.create_sheet("Table C Target Phase"), payload["paper_ready_tables"]["table_c_phase_specific_performance"])
    _write_rows(workbook.create_sheet("Table D Intervention"), payload["paper_ready_tables"]["table_d_intervention_response_audit"])
    _write_rows(workbook.create_sheet("Forecast Rows"), flatten_forecasts(payload))
    _write_rows(workbook.create_sheet("Exclusions"), flatten_exclusions(payload))
    _write_rows(
        workbook.create_sheet("Rapid Removal Notes"),
        [
            {"experiment": experiment["label"], **experiment["rapid_removal_diagnostics"]}
            for experiment in payload["experiments"]
        ],
    )
    _write_rows(
        workbook.create_sheet("Selected Alphas"),
        [
            {"experiment": experiment["label"], "horizon": horizon, "ridge_alpha": alpha}
            for experiment in payload["experiments"]
            for horizon, alpha in experiment["selected_alphas"].items()
        ],
    )
    _write_rows(
        workbook.create_sheet("Cycle Windows"),
        [
            {"experiment": experiment["label"], "role": role, **cycle}
            for experiment in payload["experiments"]
            for role, cycle in [("training", experiment["training_cycle"]), ("test", experiment["test_cycle"])]
        ],
    )
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = cell.font.copy(bold=True)
            for column in sheet.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _manifest(campaign, definitions, measurements, interval, rolling_window):
    starts = [row["measured_at"] for row in measurements if row.get("measured_at")]
    return {
        "campaign_id": campaign.id,
        "campaign_label": campaign.name,
        "measurement_start": _iso(min(starts)) if starts else None,
        "measurement_end": _iso(max(starts)) if starts else None,
        "cycle_labels": [cycle.cycle_label for cycle in definitions],
        "stored_cycle_windows": [_cycle_payload(cycle) for cycle in definitions],
        "timestamp_warning": "Stored timestamps are provisional and are not independently confirmed as local window-opening times.",
        "model_definitions": {
            "persistence": "C(t+h) = C(t)",
            "rolling_mean_3": f"mean of the latest {rolling_window} available observations up to t",
            "ridge_autoregressive": "ridge-linear direct model using radon-only lag and rolling features from observations at or before t",
        },
        "lag_list": ["lag_1", "lag_2", "lag_3", "lag_6", "lag_12_when_available", "first_difference", "rolling_mean_3", "rolling_mean_6"],
        "rolling_windows": [rolling_window, 6],
        "horizons": list(HORIZONS),
        "ridge_alpha_grid": ALPHA_CANDIDATES,
        "validation_rule": "blocked chronological validation inside the training cycle only",
        "scaling_status": "No scaler or standardisation is applied; alpha therefore depends on the scale of the radon-derived features.",
        "direct_forecast": True,
        "phase_assignment_rule": "Existing phase_metrics use forecast-origin phase; target_phase_metrics use target timestamp phase and are primary for paper-ready evaluation.",
        "quality_handling": "quality_affected targets are excluded; near-zero radon values are retained; MAPE is not used",
        "gap_handling": "examples crossing missing timestamps or lag-window sampling gaps are excluded; no interpolation is performed",
        "sampling_interval_hours": interval,
        "algorithm_version": VERSION,
        "generated_at": timezone.now().replace(microsecond=0).isoformat(),
    }


def _paper_ready_tables(payload):
    table_a = []
    for row in flatten_overall_metrics(payload):
        experiment = _experiment_by_label(payload, row["experiment"])
        table_a.append(
            {
                "Experiment": row["experiment"],
                "Train cycle": experiment["training_cycle"]["cycle_label"],
                "Test cycle": experiment["test_cycle"]["cycle_label"],
                "Model": row["model"],
                "Horizon": row["horizon"],
                "N": row["forecast_count"],
                "MAE": row["mae"],
                "RMSE": row["rmse"],
                "Bias": row["mean_bias"],
                "Median AE": row["median_absolute_error"],
                "Maximum AE": row["maximum_absolute_error"],
            }
        )
    table_b = [
        {
            "Experiment": row["experiment"],
            "Horizon": row["horizon"],
            "Common N": row["common_target_count"],
            "Model": row["model"],
            "MAE": row["mae"],
            "RMSE": row["rmse"],
            "Bias": row["bias"],
            "Relative MAE improvement vs persistence": row["relative_mae_improvement_vs_persistence"],
            "Relative RMSE improvement vs persistence": row["relative_rmse_improvement_vs_persistence"],
        }
        for row in flatten_fair_comparison_metrics(payload)
    ]
    table_c = [
        {
            "Experiment": row["experiment"],
            "Model": row["model"],
            "Horizon": row["horizon"],
            "Phase": row["target_phase"],
            "N": row["forecast_count"],
            "MAE": row["mae"],
            "RMSE": row["rmse"],
            "Bias": row["mean_bias"],
            "Maximum AE": row["maximum_absolute_error"],
        }
        for row in flatten_target_phase_metrics(payload)
    ]
    table_d = [
        {
            "Experiment": row["experiment"],
            "Model": row["model"],
            "Horizon": row["horizon"],
            "Pre-event forecast error": row["pre_event_forecast_error"],
            "First rapid-removal target error": row["first_rapid_removal_target_error"],
            "Maximum rapid-removal error": row["maximum_rapid_removal_error"],
            "Timestamp of maximum error": row["timestamp_of_maximum_error"],
            "Error after first new observation": row["error_after_first_new_observation"],
            "Error after second new observation": row["error_after_second_new_observation"],
            "Error after third new observation": row["error_after_third_new_observation"],
        }
        for row in flatten_intervention_audit(payload)
    ]
    return {
        "table_a_overall_predictive_performance": table_a,
        "table_b_fair_comparison_performance": table_b,
        "table_c_phase_specific_performance": table_c,
        "table_d_intervention_response_audit": table_d,
    }


def _experiment_by_label(payload, label):
    for experiment in payload["experiments"]:
        if experiment["label"] == label:
            return experiment
    return {}


def _leakage_audit(payload):
    rows = flatten_forecasts(payload)
    origin_before_target = all(_parse_iso(row["timestamp"]) < _parse_iso(row["target_timestamp"]) for row in rows)
    direct_horizons = all(
        (_parse_iso(row["target_timestamp"]) - _parse_iso(row["timestamp"])) == HORIZONS[row["horizon"]]
        for row in rows
    )
    return {
        "forecast_origin_before_target": origin_before_target,
        "direct_horizon_construction": direct_horizons,
        "feature_rule": "lag, difference and rolling features are constructed only from observations at or before forecast origin t",
        "ridge_training_rule": "Ridge is fit only on the development cycle for each experiment; test cycle rows are not used for fitting or alpha selection.",
        "validation_rule": "alpha selection uses blocked chronological validation inside the training cycle only",
        "preprocessing_rule": "no scaler or standardisation is used; no test-cycle preprocessing is fitted",
        "recursive_forecasting": False,
    }


def _duplicate_forecast_key_count(rows):
    keys = [
        (row["experiment"], row["model"], row["horizon"], row["timestamp"], row["target_timestamp"])
        for row in rows
    ]
    return len(keys) - len(set(keys))


def _run_experiment(label, train_cycle, test_cycle, measurements, interval, rolling_window):
    train_window = _cycle_window(train_cycle)
    test_window = _cycle_window(test_cycle)
    all_forecasts = []
    exclusions = []
    selected_alphas = {}
    alpha_selection_audit = []
    for horizon_label, horizon in HORIZONS.items():
        train_examples, train_exclusions = _examples(measurements, train_cycle, train_window, horizon, interval, rolling_window, training=True)
        test_examples, test_exclusions = _examples(measurements, test_cycle, test_window, horizon, interval, rolling_window, training=False)
        exclusions.extend(_exclusion_rows(label, horizon_label, "training", train_exclusions))
        exclusions.extend(_exclusion_rows(label, horizon_label, "test", test_exclusions))
        use_lag12 = bool(train_examples) and all("lag_12" in sample["features"] for sample in [*train_examples, *test_examples])
        alpha, alpha_audit = _select_alpha(train_examples, use_lag12)
        selected_alphas[horizon_label] = alpha
        alpha_selection_audit.extend(
            [
                {
                    "experiment": label,
                    "horizon": horizon_label,
                    "training_cycle": train_cycle.cycle_label,
                    "test_cycle": test_cycle.cycle_label,
                    **row,
                }
                for row in alpha_audit
            ]
        )
        ridge_coefficients = _fit_ridge_for_examples(train_examples, alpha, use_lag12)
        for sample in test_examples:
            predictions = _predictions(sample, ridge_coefficients, use_lag12)
            for model, prediction in predictions.items():
                signed_error = _round(prediction - sample["target"])
                all_forecasts.append(
                    {
                        "experiment": label,
                        "training_cycle": train_cycle.cycle_label,
                        "test_cycle": test_cycle.cycle_label,
                        "horizon": horizon_label,
                        "model": model,
                        "timestamp": _iso(sample["origin_time"]),
                        "forecast_origin": _iso(sample["origin_time"]),
                        "target_timestamp": _iso(sample["target_time"]),
                        "phase": sample["origin_phase"],
                        "origin_phase": sample["origin_phase"],
                        "target_phase": sample["target_phase"],
                        "actual": sample["target"],
                        "predicted": _round(prediction),
                        "error": signed_error,
                        "signed_error_predicted_minus_observed": signed_error,
                        "absolute_error": _round(abs(prediction - sample["target"])),
                    }
                )
    return {
        "label": label,
        "training_cycle": _cycle_payload(train_cycle),
        "test_cycle": _cycle_payload(test_cycle),
        "selected_alphas": selected_alphas,
        "alpha_selection_audit": alpha_selection_audit,
        "overall_metrics": _metric_rows(all_forecasts, ["horizon", "model"], {"experiment": label}),
        "phase_metrics": _metric_rows(all_forecasts, ["horizon", "model", "phase"], {"experiment": label}),
        "target_phase_metrics": _metric_rows(all_forecasts, ["horizon", "model", "target_phase"], {"experiment": label}),
        "fair_comparison_metrics": _fair_comparison_rows(label, train_cycle, test_cycle, all_forecasts),
        "forecast_rows": all_forecasts,
        "exclusions": exclusions,
        "rapid_removal_diagnostics": _rapid_removal_diagnostics(label, test_cycle, all_forecasts, measurements),
        "intervention_response_audit": _intervention_response_rows(label, train_cycle, test_cycle, all_forecasts),
        "plot": _plot_payload(test_cycle, measurements, all_forecasts),
        "rapid_removal_plot": _focused_plot_payload(test_cycle, measurements, all_forecasts),
    }


def _examples(measurements, cycle, cycle_window, horizon, interval, rolling_window, training):
    by_time = {row["measured_at"]: row for row in measurements}
    rows = [row for row in measurements if cycle_window[0] <= row["measured_at"] <= cycle_window[1]]
    examples = []
    exclusions = defaultdict(int)
    for origin in rows:
        origin_phase = _phase_for_time(cycle, origin["measured_at"])
        if not origin_phase:
            continue
        target_time = origin["measured_at"] + horizon
        if not (cycle_window[0] <= target_time <= cycle_window[1]):
            exclusions["target_outside_cycle_window"] += 1
            continue
        target = by_time.get(target_time)
        if not target:
            exclusions["missing_future_target"] += 1
            continue
        if _is_quality_affected(target):
            exclusions["quality_affected_target"] += 1
            continue
        history = _history_rows(measurements, origin["measured_at"], 11)
        sequence = history + [origin]
        if len(sequence) < 6:
            exclusions["insufficient_history"] += 1
            continue
        if _crosses_gap(sequence[-6:], interval) or _origin_target_crosses_gap(by_time, origin["measured_at"], target_time, interval):
            exclusions["crosses_gap"] += 1
            continue
        feature_map = {
            "lag_1": _radon(sequence[-1]),
            "lag_2": _radon(sequence[-2]),
            "lag_3": _radon(sequence[-3]),
            "lag_6": _radon(sequence[-6]),
            "diff_1": _radon(sequence[-1]) - _radon(sequence[-2]),
            "rolling_mean_3": _mean([_radon(row) for row in sequence[-3:]]),
            "rolling_mean_6": _mean([_radon(row) for row in sequence[-6:]]),
        }
        if len(sequence) >= 12:
            feature_map["lag_12"] = _radon(sequence[-12])
        examples.append(
            {
                "origin_time": origin["measured_at"],
                "target_time": target_time,
                "origin_phase": origin_phase,
                "target_phase": _phase_for_time(cycle, target_time),
                "current": _radon(origin),
                "rolling_mean": _mean([_radon(row) for row in sequence[-rolling_window:]]),
                "features": feature_map,
                "target": _radon(target),
            }
        )
    return examples, dict(exclusions)


def _predictions(sample, ridge_coefficients, use_lag12):
    features = _feature_vector(sample, use_lag12)
    return {
        "persistence": sample["current"],
        "rolling_mean_3": sample["rolling_mean"],
        "ridge_autoregressive": _predict(ridge_coefficients, features) if ridge_coefficients else sample["current"],
    }


def _select_alpha(samples, use_lag12):
    if len(samples) < 12:
        return 1.0, [
            {
                "alpha": 1.0,
                "validation_mae": None,
                "selected": True,
                "selection_metric": "mae",
                "training_rows": len(samples),
                "validation_rows": 0,
                "validation_rule": "insufficient rows for blocked chronological validation; default alpha used",
            }
        ]
    split = max(6, int(len(samples) * 0.7))
    train = samples[:split]
    validation = samples[split:]
    scores = []
    for alpha in ALPHA_CANDIDATES:
        coefficients = _fit_ridge_for_examples(train, alpha, use_lag12)
        predictions = [_predict(coefficients, _feature_vector(sample, use_lag12)) for sample in validation]
        scores.append(((_metrics([sample["target"] for sample in validation], predictions) or {}).get("mae"), alpha))
    valid = [row for row in scores if row[0] is not None]
    selected = min(valid, key=lambda row: (row[0], row[1]))[1] if valid else 1.0
    audit = []
    for score, alpha in scores:
        audit.append(
            {
                "alpha": alpha,
                "validation_mae": score,
                "selected": alpha == selected,
                "selection_metric": "mae",
                "training_rows": len(train),
                "validation_rows": len(validation),
                "training_start": _iso(train[0]["origin_time"]) if train else None,
                "training_end": _iso(train[-1]["origin_time"]) if train else None,
                "validation_start": _iso(validation[0]["origin_time"]) if validation else None,
                "validation_end": _iso(validation[-1]["origin_time"]) if validation else None,
                "validation_rule": "blocked chronological validation inside the training cycle",
            }
        )
    return selected, audit


def _fit_ridge_for_examples(samples, alpha, use_lag12):
    if len(samples) < 4:
        return None
    return _fit_ridge([_feature_vector(sample, use_lag12) for sample in samples], [sample["target"] for sample in samples], alpha)


def _feature_vector(sample, use_lag12):
    keys = ["lag_1", "lag_2", "lag_3", "lag_6", "diff_1", "rolling_mean_3", "rolling_mean_6"]
    if use_lag12:
        keys.insert(4, "lag_12")
    return [sample["features"][key] for key in keys]


def _metric_rows(forecasts, group_keys, extra):
    grouped = defaultdict(list)
    for row in forecasts:
        grouped[tuple(row[key] for key in group_keys)].append(row)
    rows = []
    for key, values in sorted(grouped.items()):
        metrics = _metrics([row["actual"] for row in values], [row["predicted"] for row in values])
        rows.append({**extra, **dict(zip(group_keys, key)), **metrics})
    return rows


def _metrics(actual, predicted):
    if not actual:
        return {"forecast_count": 0, "mae": None, "rmse": None, "mean_bias": None, "median_absolute_error": None, "maximum_absolute_error": None}
    errors = [p - a for a, p in zip(actual, predicted)]
    absolute = [abs(error) for error in errors]
    return {
        "forecast_count": len(actual),
        "mae": _round(sum(absolute) / len(absolute)),
        "rmse": _round(sqrt(sum(error * error for error in errors) / len(errors))),
        "mean_bias": _round(sum(errors) / len(errors)),
        "median_absolute_error": _round(median(absolute)),
        "maximum_absolute_error": _round(max(absolute)),
    }


def _fair_comparison_rows(label, train_cycle, test_cycle, forecasts):
    rows = []
    grouped = defaultdict(list)
    for row in forecasts:
        grouped[row["horizon"]].append(row)
    for horizon, horizon_rows in sorted(grouped.items()):
        keys_by_model = {
            model: {
                (row["timestamp"], row["target_timestamp"])
                for row in horizon_rows
                if row["model"] == model
            }
            for model in MODELS
        }
        common_keys = set.intersection(*(keys_by_model[model] for model in MODELS)) if all(keys_by_model.values()) else set()
        persistence_metrics = None
        model_metrics = {}
        for model in MODELS:
            subset = [
                row
                for row in horizon_rows
                if row["model"] == model and (row["timestamp"], row["target_timestamp"]) in common_keys
            ]
            metrics = _metrics([row["actual"] for row in subset], [row["predicted"] for row in subset])
            model_metrics[model] = metrics
            if model == "persistence":
                persistence_metrics = metrics
        for model, metrics in model_metrics.items():
            rows.append(
                {
                    "experiment": label,
                    "training_cycle": train_cycle.cycle_label,
                    "test_cycle": test_cycle.cycle_label,
                    "horizon": horizon,
                    "model": model,
                    "model_specific_target_count": len(keys_by_model[model]),
                    "common_target_count": len(common_keys),
                    "mae": metrics["mae"],
                    "rmse": metrics["rmse"],
                    "bias": metrics["mean_bias"],
                    "relative_mae_improvement_vs_persistence": _relative_improvement(persistence_metrics["mae"], metrics["mae"]) if persistence_metrics else None,
                    "relative_rmse_improvement_vs_persistence": _relative_improvement(persistence_metrics["rmse"], metrics["rmse"]) if persistence_metrics else None,
                    "difference_reason": "same valid target set" if len({len(keys) for keys in keys_by_model.values()}) == 1 else "model-specific valid target sets differ",
                }
            )
    return rows


def _relative_improvement(baseline, model_value):
    if baseline in (None, 0) or model_value is None:
        return None
    return _round(100 * (baseline - model_value) / baseline)


def _rapid_removal_diagnostics(label, test_cycle, forecasts, measurements):
    start, end = test_cycle.rapid_removal_start, test_cycle.rapid_removal_end
    before = [row for row in measurements if row["measured_at"] < start]
    during = [row for row in measurements if start <= row["measured_at"] <= end]
    phase_forecasts = [row for row in forecasts if row["target_phase"] == "rapid_removal"]
    largest = max(phase_forecasts, key=lambda row: row["absolute_error"], default=None)
    first_before = max([row for row in forecasts if row["timestamp"] < _iso(start)], key=lambda row: row["timestamp"], default=None)
    by_horizon = _rapid_removal_forecasts_by_horizon(forecasts, start)
    return {
        "experiment": label,
        "rapid_removal_start": _iso(start),
        "rapid_removal_end": _iso(end),
        "concentration_immediately_before": _radon(before[-1]) if before else None,
        "first_forecast_before_removal": first_before,
        "forecast_observed_at_horizons": by_horizon,
        "sudden_drop_onset_error": _sudden_drop_onset_error(forecasts, start),
        "largest_error_during_removal": largest,
        "adaptation_time_to_abs_error_lte_10_hours_by_model": _adaptation_times(forecasts, start, 10),
        "observed_time_to_lte_50_hours": _observed_threshold_time(during, start, 50),
        "observed_time_to_lte_30_hours": _observed_threshold_time(during, start, 30),
        "predicted_time_to_lte_50_hours_by_model": THRESHOLD_NOT_IDENTIFIABLE,
        "predicted_time_to_lte_30_hours_by_model": THRESHOLD_NOT_IDENTIFIABLE,
        "predicted_threshold_time_note": "Direct 1h/3h/6h forecasts are not a continuous or recursive trajectory, so threshold-crossing time is not identifiable from this output.",
        "interpretation": "Failure to anticipate an unknown opening is not treated as a software defect.",
    }


def _intervention_response_rows(label, train_cycle, test_cycle, forecasts):
    start, end = test_cycle.rapid_removal_start, test_cycle.rapid_removal_end
    rows = []
    for horizon in HORIZONS:
        for model in MODELS:
            model_rows = sorted(
                [row for row in forecasts if row["horizon"] == horizon and row["model"] == model],
                key=lambda row: (_parse_iso(row["target_timestamp"]), _parse_iso(row["timestamp"])),
            )
            pre_event = _last_before(model_rows, "timestamp", start)
            first_rapid = _first_target_in_window(model_rows, start, end)
            max_rapid = max(
                [row for row in model_rows if start <= _parse_iso(row["target_timestamp"]) <= end],
                key=lambda row: row["absolute_error"],
                default=None,
            )
            rows.append(
                {
                    "experiment": label,
                    "training_cycle": train_cycle.cycle_label,
                    "test_cycle": test_cycle.cycle_label,
                    "model": model,
                    "horizon": horizon,
                    "pre_event_forecast_origin": pre_event["timestamp"] if pre_event else None,
                    "pre_event_target_timestamp": pre_event["target_timestamp"] if pre_event else None,
                    "pre_event_forecast_error": pre_event["error"] if pre_event else None,
                    "first_rapid_removal_target_timestamp": first_rapid["target_timestamp"] if first_rapid else None,
                    "first_rapid_removal_target_error": first_rapid["error"] if first_rapid else None,
                    "maximum_rapid_removal_error": max_rapid["absolute_error"] if max_rapid else None,
                    "timestamp_of_maximum_error": max_rapid["target_timestamp"] if max_rapid else None,
                    "error_after_first_new_observation": _adaptation_error(model_rows, start, 1),
                    "error_after_second_new_observation": _adaptation_error(model_rows, start, 2),
                    "error_after_third_new_observation": _adaptation_error(model_rows, start, 3),
                    "interpretation": "pre-event forecasts are unconditional on future ventilation; adaptation rows occur after new rapid-removal measurements are available",
                }
            )
    return rows


def _last_before(rows, timestamp_key, boundary):
    candidates = [row for row in rows if _parse_iso(row[timestamp_key]) < boundary]
    return max(candidates, key=lambda row: _parse_iso(row[timestamp_key]), default=None)


def _first_target_in_window(rows, start, end):
    candidates = [row for row in rows if start <= _parse_iso(row["target_timestamp"]) <= end]
    return min(candidates, key=lambda row: _parse_iso(row["target_timestamp"]), default=None)


def _adaptation_error(rows, rapid_start, observations_seen):
    candidates = [
        row
        for row in rows
        if _parse_iso(row["timestamp"]) >= rapid_start + timedelta(hours=observations_seen - 1)
    ]
    row = min(candidates, key=lambda item: _parse_iso(item["timestamp"]), default=None)
    return row["error"] if row else None


def _rapid_removal_forecasts_by_horizon(forecasts, rapid_start):
    rows = {}
    for horizon in HORIZONS:
        target_time = rapid_start + HORIZONS[horizon]
        matches = [
            row
            for row in forecasts
            if row["horizon"] == horizon and _parse_iso(row["target_timestamp"]) == target_time
        ]
        rows[horizon] = {
            "target_timestamp": _iso(target_time),
            "actual": matches[0]["actual"] if matches else None,
            "predictions": {row["model"]: row["predicted"] for row in matches},
            "absolute_errors": {row["model"]: row["absolute_error"] for row in matches},
        }
    return rows


def _sudden_drop_onset_error(forecasts, rapid_start):
    rows = [row for row in forecasts if _parse_iso(row["target_timestamp"]) == rapid_start]
    if not rows:
        rows = [row for row in forecasts if _parse_iso(row["target_timestamp"]) > rapid_start]
    if not rows:
        return None
    return {row["model"]: row["absolute_error"] for row in sorted(rows, key=lambda item: (item["horizon"], item["model"]))}


def _adaptation_times(forecasts, rapid_start, error_threshold):
    adapted = {}
    for model in MODELS:
        rows = sorted(
            [
                row
                for row in forecasts
                if row["model"] == model and _parse_iso(row["target_timestamp"]) >= rapid_start
            ],
            key=lambda row: _parse_iso(row["target_timestamp"]),
        )
        adapted[model] = None
        for row in rows:
            if row["absolute_error"] <= error_threshold:
                adapted[model] = _round((_parse_iso(row["target_timestamp"]) - rapid_start).total_seconds() / 3600)
                break
    return adapted


def _plot_payload(cycle, measurements, forecasts):
    start, end = _cycle_window(cycle)
    return _plot_payload_for_window(cycle, measurements, forecasts, start, end)


def _focused_plot_payload(cycle, measurements, forecasts):
    start = cycle.rapid_removal_start - timedelta(hours=6)
    end = cycle.rapid_removal_end + timedelta(hours=8)
    return _plot_payload_for_window(cycle, measurements, forecasts, start, end)


def _plot_payload_for_window(cycle, measurements, forecasts, start, end):
    observed = [row for row in measurements if start <= row["measured_at"] <= end]
    if not observed:
        return {"observed_points": "", "prediction_series": {}, "phase_bands": []}
    first_time = observed[0]["measured_at"]
    total_seconds = max((observed[-1]["measured_at"] - first_time).total_seconds(), 1)
    window_forecasts = [row for row in forecasts if start <= _parse_iso(row["target_timestamp"]) <= end]
    values = [_radon(row) for row in observed] + [row["predicted"] for row in window_forecasts]
    min_value, max_value = min(values), max(values)
    span = max(max_value - min_value, 1)
    return {
        "observed_points": _points(observed, first_time, total_seconds, min_value, span),
        "prediction_series": {
            model: _points([{"measured_at": _parse_iso(row["target_timestamp"]), "radon_bq_m3": row["predicted"]} for row in window_forecasts if row["model"] == model], first_time, total_seconds, min_value, span)
            for model in MODELS
        },
        "phase_bands": _phase_bands(cycle, first_time, total_seconds),
        "rapid_removal_band": _single_phase_band(cycle.rapid_removal_start, cycle.rapid_removal_end, first_time, total_seconds),
    }


def _points(rows, first_time, total_seconds, min_value, span):
    points = []
    for row in sorted(rows, key=lambda item: item["measured_at"]):
        x = 28 + ((row["measured_at"] - first_time).total_seconds() / total_seconds) * 644
        y = 192 - ((_radon(row) - min_value) / span) * 164
        points.append(f"{x:.1f},{y:.1f}")
    return " ".join(points)


def _phase_bands(cycle, first_time, total_seconds):
    colors = {"baseline": "#e8eef7", "accumulation": "#fff0c9", "rapid_removal": "#d9f3ee", "post_event": "#eef8e8"}
    bands = []
    for phase in ["baseline", "accumulation", "rapid_removal", "post_event"]:
        start = getattr(cycle, f"{phase}_start")
        end = getattr(cycle, f"{phase}_end")
        x = 28 + ((start - first_time).total_seconds() / total_seconds) * 644
        x2 = 28 + ((end - first_time).total_seconds() / total_seconds) * 644
        bands.append({"phase": phase, "x": round(x, 1), "width": round(max(x2 - x, 1), 1), "color": colors[phase]})
    return bands


def _single_phase_band(start, end, first_time, total_seconds):
    x = 28 + ((start - first_time).total_seconds() / total_seconds) * 644
    x2 = 28 + ((end - first_time).total_seconds() / total_seconds) * 644
    return {"x": round(x, 1), "width": round(max(x2 - x, 1), 1)}


def _measurement_rows(campaign):
    return list(campaign.measurements.exclude(radon_bq_m3=None).order_by("measured_at", "id").values("measured_at", "radon_bq_m3", "regime"))


def _cycle_window(cycle):
    starts = [getattr(cycle, f"{phase}_start") for phase in ["baseline", "accumulation", "rapid_removal", "post_event"]]
    ends = [getattr(cycle, f"{phase}_end") for phase in ["baseline", "accumulation", "rapid_removal", "post_event"]]
    return min(starts), max(ends)


def _phase_for_time(cycle, timestamp):
    for phase in ["baseline", "accumulation", "rapid_removal", "post_event"]:
        if getattr(cycle, f"{phase}_start") <= timestamp <= getattr(cycle, f"{phase}_end"):
            return phase
    return None


def _history_rows(measurements, origin_time, max_lag):
    return [row for row in measurements if row["measured_at"] < origin_time][-max_lag:]


def _origin_target_crosses_gap(by_time, origin, target, interval):
    if not interval:
        return False
    current = origin
    while current < target:
        nxt = current + timedelta(hours=interval)
        if nxt > target:
            break
        if nxt not in by_time:
            return True
        current = nxt
    return False


def _crosses_gap(rows, interval):
    if not interval or len(rows) < 2:
        return False
    for previous, current in zip(rows, rows[1:]):
        if (current["measured_at"] - previous["measured_at"]).total_seconds() / 3600 > interval * 1.5:
            return True
    return False


def _exclusion_rows(experiment, horizon, role, exclusions):
    return [{"experiment": experiment, "horizon": horizon, "role": role, "reason": reason, "count": count} for reason, count in sorted(exclusions.items())]


def _median_interval(rows):
    intervals = [(b["measured_at"] - a["measured_at"]).total_seconds() / 3600 for a, b in zip(rows, rows[1:]) if b["measured_at"] > a["measured_at"]]
    return _round(median(intervals)) if intervals else None


def _fit_ridge(features, targets, alpha):
    design = [[1.0, *row] for row in features]
    size = len(design[0])
    xtx = [[0.0] * size for _ in range(size)]
    xty = [0.0] * size
    for row, target in zip(design, targets):
        for i in range(size):
            xty[i] += row[i] * target
            for j in range(size):
                xtx[i][j] += row[i] * row[j]
    for i in range(1, size):
        xtx[i][i] += alpha
    return _solve(xtx, xty)


def _solve(matrix, vector):
    n = len(vector)
    augmented = [matrix[i][:] + [vector[i]] for i in range(n)]
    for pivot in range(n):
        row = max(range(pivot, n), key=lambda idx: abs(augmented[idx][pivot]))
        augmented[pivot], augmented[row] = augmented[row], augmented[pivot]
        divisor = augmented[pivot][pivot] or 1e-12
        for col in range(pivot, n + 1):
            augmented[pivot][col] /= divisor
        for row in range(n):
            if row == pivot:
                continue
            factor = augmented[row][pivot]
            for col in range(pivot, n + 1):
                augmented[row][col] -= factor * augmented[pivot][col]
    return [augmented[i][-1] for i in range(n)]


def _predict(coefficients, features):
    return coefficients[0] + sum(coef * value for coef, value in zip(coefficients[1:], features))


def _write_rows(sheet, rows):
    headers = sorted({key for row in rows for key in row}) if rows else ["note"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
    else:
        sheet.append(["No rows."])


def _cycle_payload(cycle):
    payload = {"cycle_label": cycle.cycle_label, "evidence_status": cycle.evidence_status, "note": cycle.note}
    for phase in ["baseline", "accumulation", "rapid_removal", "post_event"]:
        payload[f"{phase}_start"] = _iso(getattr(cycle, f"{phase}_start"))
        payload[f"{phase}_end"] = _iso(getattr(cycle, f"{phase}_end"))
    return payload


def _observed_threshold_time(rows, start, threshold):
    for row in rows:
        if _radon(row) <= threshold:
            return _round((row["measured_at"] - start).total_seconds() / 3600)
    return None


def _timezone_status(rows):
    return "Stored measurement timestamps are timezone-aware." if all(timezone.is_aware(row["measured_at"]) for row in rows if row["measured_at"]) else "Some stored measurement timestamps are timezone-naive."


def _is_quality_affected(row):
    return (row.get("regime") or "").lower() == "quality_affected"


def _radon(row):
    return float(row["radon_bq_m3"])


def _mean(values):
    return sum(values) / len(values)


def _iso(value):
    return value.isoformat() if value else None


def _parse_iso(value):
    return datetime_from_iso(value)


def datetime_from_iso(value):
    from django.utils.dateparse import parse_datetime

    return parse_datetime(value)


def _round(value):
    return round(float(value), 3) if value is not None else None


def _excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, default=str)
    return value
