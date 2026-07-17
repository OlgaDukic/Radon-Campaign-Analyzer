from decimal import Decimal
from datetime import timedelta, timezone as datetime_timezone
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from zoneinfo import ZoneInfo

from django.conf import settings
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import AnalysisReport, Campaign, CampaignResearchContext, Measurement, UploadedFile
from .services.analysis import run_campaign_analysis
from .services.analysis_config import AnalysisConfig
from .services import apparent_dynamics_audit
from .services import baseline_prediction_experiment as baseline_experiment
from .services import reduced_state_space_experiment as state_space_experiment
from .services import state_space_validation_audit
from .services import final_model_results
from .services.analysis_profiles import build_config, parse_overrides
from .services.canonicalization import build_canonical_outputs
from .services.campaign_comparison import compare_campaigns
from .services.ingestion import parse_decimal, read_uploaded_file
from .services.prediction import evaluate_prediction_models
from .services.regime_sensitivity import build_regime_sensitivity
from .services.regimes import classify_regimes
from .services.reports import build_summary
from .services.resampling import build_hourly_resampling
from .services.sampling_gaps import detect_sampling_gaps
from .services.time_continuity import analyze_time_continuity
from .services.regime_v2 import classify_regimes_v2
from .services.episodes import build_episodes
from .services.documented_events import EventCycleDefinition, analyse_documented_cycles, default_event_cycles_for_campaign, parse_event_timestamp
from .services.sensitivity_v2 import build_sensitivity_v2
from .services.prediction_v2 import evaluate_prediction_v2
from .services.source_inventory import build_source_file_inventory
from .services.time_diagnostics import build_dst_diagnostics


class CampaignModelTests(TestCase):
    def test_campaign_string_representation(self):
        campaign = Campaign.objects.create(name="Winter Survey")

        self.assertEqual(str(campaign), "Winter Survey")

    def test_measurement_string_representation(self):
        campaign = Campaign.objects.create(name="School Survey")
        measurement = Measurement.objects.create(
            campaign=campaign,
            radon_bq_m3=Decimal("123.45"),
            room_name="Classroom A",
        )

        self.assertIn("123.45 Bq/m3", str(measurement))

    def test_analysis_service_creates_empty_report_without_files(self):
        campaign = Campaign.objects.create(name="Baseline Campaign")

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.status, AnalysisReport.Status.COMPLETE)
        self.assertEqual(report.summary_json["measurement_count"], 0)
        self.assertIn("Research prototype analysis complete", report.summary)


class CampaignViewTests(TestCase):
    def test_campaign_list_view(self):
        Campaign.objects.create(name="Residential Pilot")

        response = self.client.get(reverse("campaigns:campaign_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Residential Pilot")
        self.assertContains(response, "Campaign Home")
        self.assertContains(response, "campaign-card-grid")
        self.assertContains(response, "Compare Campaigns")
        self.assertContains(response, "Open dashboard")

    def test_create_campaign_view(self):
        response = self.client.post(
            reverse("campaigns:campaign_create"),
            {"name": "Office Pilot", "location": "Nis"},
        )

        campaign = Campaign.objects.get(name="Office Pilot")
        self.assertRedirects(response, reverse("campaigns:campaign_detail", args=[campaign.pk]))

    def test_campaign_detail_view(self):
        campaign = Campaign.objects.create(name="Lab Pilot")

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Lab Pilot")
        self.assertContains(response, "Run new analysis")

    def test_upload_csv_file_to_campaign(self):
        campaign = Campaign.objects.create(name="CSV Campaign")
        upload = SimpleUploadedFile(
            "measurements.csv",
            b"measured_at,radon_bq_m3\n2026-01-01,100\n",
            content_type="text/csv",
        )

        response = self.client.post(
            reverse("campaigns:upload_file", args=[campaign.pk]),
            {"file": upload},
        )

        self.assertRedirects(response, reverse("campaigns:campaign_detail", args=[campaign.pk]))
        uploaded_file = UploadedFile.objects.get(campaign=campaign)
        self.assertEqual(uploaded_file.original_name, "measurements.csv")

    def test_run_analysis_view(self):
        campaign = Campaign.objects.create(name="Analysis Campaign")

        response = self.client.post(reverse("campaigns:run_analysis", args=[campaign.pk]))

        self.assertRedirects(response, reverse("campaigns:campaign_detail", args=[campaign.pk]))
        self.assertEqual(campaign.analysis_reports.count(), 1)

    def test_campaign_detail_renders_research_dashboard(self):
        campaign = Campaign.objects.create(name="Dashboard Campaign")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Research prototype analysis complete.",
            summary_json={
                "measurement_count": 3,
                "segment_count": 1,
                "gap_count": 1,
                "regime_counts": {"stable_low": 1, "rising": 2},
                "prediction_metrics": {
                    "1h": {
                        "naive_baseline": {"samples": 3, "mae": 8.0, "rmse": 9.0},
                        "ridge": {"samples": 3, "mae": 4.0, "rmse": 5.0},
                    }
                },
                "prediction_metrics_by_regime": [
                    {
                        "horizon": "1h",
                        "model": "ridge",
                        "regime": "rising",
                        "samples": 3,
                        "mae": 4.0,
                        "rmse": 5.0,
                        "mae_improvement_percent": 50.0,
                        "rmse_improvement_percent": 44.44,
                    }
                ],
                "prediction_errors": [
                    {
                        "timestamp": "2026-01-01T01:00:00+00:00",
                        "horizon": "1h",
                        "model": "ridge",
                        "actual_radon": 150,
                        "predicted_radon": 140,
                        "absolute_error": 10,
                        "regime": "rising",
                        "segment_id": 1,
                    }
                ],
                "gaps": [{"from": "2026-01-01T01:00:00+00:00", "to": "2026-01-01T03:00:00+00:00", "minutes": 120.0}],
                "segments": [
                    {
                        "segment_id": 1,
                        "measurement_count": 3,
                        "segment_label": "elevated_dynamic",
                        "percent_above_100": 66.7,
                        "percent_above_200": 0.0,
                        "dynamic_percent": 66.7,
                        "interpretation_text": "Radon is frequently above 100 Bq/m3.",
                        "statistics": {"radon_bq_m3": {"mean": 120.0, "max": 150.0}},
                    }
                ],
                "ingestion_debug": [
                    {
                        "filename": "dashboard.xlsx",
                        "detected_sheets": ["Measurements"],
                        "raw_rows_read": 5,
                        "detected_header_row": 2,
                        "detected_columns": ["Date and time", "Radon"],
                        "mapped_columns": {
                            "timestamp": "Date and time",
                            "radon": "Radon",
                            "temperature": None,
                            "humidity": None,
                            "pressure": None,
                        },
                        "parsed_measurement_rows": 3,
                        "skipped_reason": "",
                    }
                ],
            },
            html_report="<article>Generated report</article>",
        )
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        for index, value in enumerate([90, 120, 150]):
            Measurement.objects.create(
                campaign=campaign,
                measured_at=start + timedelta(hours=index),
                radon_bq_m3=Decimal(str(value)),
                segment_id=1,
                regime="rising" if index else "stable_low",
            )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertEqual(report, response.context["latest_report"])
        self.assertContains(response, "Overview")
        self.assertContains(response, "Research Summary")
        self.assertContains(response, "Measurements")
        self.assertContains(response, "Maximum radon")
        self.assertContains(response, "Data Quality")
        self.assertContains(response, "Regimes & Episodes")
        self.assertContains(response, "Prediction")
        self.assertContains(response, "Files & Provenance")
        self.assertContains(response, "Radon Time Series")
        self.assertContains(response, "<svg", html=False)
        self.assertNotContains(response, "Generated report")
        self.assertNotContains(response, "stable_low")
        self.assertNotContains(response, 'type="file"', html=False)
        self.assertNotContains(response, "Run full Paper 1 analysis")

        regimes_response = self.client.get(reverse("campaigns:campaign_regimes", args=[campaign.pk]))
        self.assertContains(regimes_response, "Regime Summary")
        prediction_response = self.client.get(reverse("campaigns:campaign_prediction", args=[campaign.pk]))
        self.assertContains(prediction_response, "Prediction Insights")
        quality_response = self.client.get(reverse("campaigns:campaign_quality", args=[campaign.pk]))
        self.assertContains(quality_response, "Quality Flags")
        provenance_response = self.client.get(reverse("campaigns:campaign_provenance", args=[campaign.pk]))
        self.assertContains(provenance_response, "dashboard.xlsx")

    def test_campaign_detail_dashboard_handles_missing_summary_fields(self):
        campaign = Campaign.objects.create(name="Sparse Dashboard")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Sparse report.",
            summary_json={"regime_counts": {}, "segments": []},
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertContains(response, "Overview")
        self.assertContains(response, "N/A")
        self.assertContains(response, "Radon Time Series")
        self.assertContains(response, "Prediction")

    def test_campaign_analysis_section_routes_render_shared_navigation(self):
        campaign = Campaign.objects.create(name="Section Campaign", location="Lab")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Section report.",
            summary_json={
                "measurement_count": 1,
                "concentration_level_counts": {"LOW": 1},
                "dynamic_state_counts": {"STABLE": 1},
                "episode_type_counts": {},
                "prediction_metrics": {},
            },
        )
        route_names = [
            "campaign_detail",
            "campaign_quality",
            "campaign_regimes",
            "campaign_prediction",
            "campaign_sensitivity",
            "campaign_research_context",
            "campaign_provenance",
            "campaign_reports",
        ]

        for route_name in route_names:
            with self.subTest(route=route_name):
                response = self.client.get(reverse(f"campaigns:{route_name}", args=[campaign.pk]))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "Data Quality")
                self.assertContains(response, "Regimes & Episodes")
                self.assertContains(response, "Reports & Exports")
                self.assertContains(response, "Section Campaign")

    def test_campaign_report_pages_render_breadcrumb_and_all_campaigns_link(self):
        campaign = Campaign.objects.create(name="Breadcrumb Campaign", location="Lab")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Breadcrumb report.",
            summary_json={"measurement_count": 0},
        )
        route_names = [
            "campaign_detail",
            "campaign_quality",
            "campaign_regimes",
            "campaign_prediction",
            "campaign_sensitivity",
            "campaign_research_context",
            "campaign_provenance",
            "campaign_reports",
        ]

        for route_name in route_names:
            with self.subTest(route=route_name):
                response = self.client.get(reverse(f"campaigns:{route_name}", args=[campaign.pk]))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "Campaigns")
                self.assertContains(response, "Breadcrumb Campaign")
                self.assertContains(response, reverse("campaigns:campaign_list"))
                self.assertContains(response, "All campaigns")

    def test_campaign_list_compares_completed_campaigns_without_mutating_reports(self):
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        first = Campaign.objects.create(name="Compare A", location="Salerno")
        second = Campaign.objects.create(name="Compare B", location="Nis")
        for index, value in enumerate([100, 120, 140]):
            Measurement.objects.create(
                campaign=first,
                measured_at=start + timedelta(days=index),
                radon_bq_m3=Decimal(str(value)),
                temperature_c=Decimal("20.0"),
                humidity_percent=Decimal("50.0"),
                pressure_hpa=Decimal("1010.0"),
            )
        for index, value in enumerate([50, 60]):
            Measurement.objects.create(
                campaign=second,
                measured_at=start + timedelta(days=index),
                radon_bq_m3=Decimal(str(value)),
            )
        first_summary = {
            "profile_metadata": {"profile_name": "profile-a", "profile_version": "v1"},
            "analysis_config": {"concentration_low_threshold_bq_m3": 100, "concentration_high_threshold_bq_m3": 200},
            "source_file_inventory": [{"device_id": "2E81E"}],
            "concentration_level_counts": {"ELEVATED": 2, "LOW": 1},
            "sampling_gaps_compact_summary": {"long_gaps": 2},
            "episodes": [
                {"episode_type": "ACCUMULATION", "quality_status": "OK", "duration_hours": 3, "measurement_count": 4, "confidence_category": "HIGH"},
                {"episode_type": "QUALITY_AFFECTED", "quality_status": "QUALITY_AFFECTED", "duration_hours": 0, "measurement_count": 1, "confidence_category": "LOW"},
            ],
            "profile_applicability": {"status": "PROFILE_COMPATIBLE"},
        }
        second_summary = {
            "profile_metadata": {"profile_name": "profile-b", "profile_version": "v2"},
            "analysis_config": {"concentration_low_threshold_bq_m3": 150, "concentration_high_threshold_bq_m3": 250},
            "source_file_inventory": [{"device_id": "RadonEye"}],
            "concentration_level_counts": {"LOW": 2},
            "sampling_gaps_compact_summary": {"long_gaps": 1},
            "episodes": [],
            "profile_applicability": {"status": "PROFILE_COMPATIBLE_WITH_WARNINGS"},
        }
        first_report = AnalysisReport.objects.create(campaign=first, status=AnalysisReport.Status.COMPLETE, summary="A", summary_json=first_summary.copy())
        second_report = AnalysisReport.objects.create(campaign=second, status=AnalysisReport.Status.COMPLETE, summary="B", summary_json=second_summary.copy())

        response = self.client.get(reverse("campaigns:campaign_list"), {"compare": [first.pk, second.pk]})

        self.assertContains(response, "Compare Campaigns")
        self.assertContains(response, "Compare A")
        self.assertContains(response, "Compare B")
        self.assertContains(response, "2E81E")
        self.assertContains(response, "RadonEye")
        self.assertContains(response, "66.7%")
        self.assertContains(response, "100.0% (3 rows)")
        self.assertContains(response, "0.0% (0 rows)")
        self.assertContains(response, "not measured")
        self.assertContains(response, "Major gaps per 30 measurement days")
        self.assertContains(response, "Data-quality events per 1,000 measurements")
        self.assertContains(response, "different analysis-profile versions")
        self.assertEqual(AnalysisReport.objects.get(pk=first_report.pk).summary_json, first_summary)
        self.assertEqual(AnalysisReport.objects.get(pk=second_report.pk).summary_json, second_summary)

    def test_research_context_page_handles_absent_metadata(self):
        campaign = Campaign.objects.create(name="No Context Campaign")

        response = self.client.get(reverse("campaigns:campaign_research_context", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Research Context")
        self.assertContains(response, "Not available")
        self.assertEqual(CampaignResearchContext.objects.filter(campaign=campaign).count(), 1)

    def test_research_context_stores_optional_values_and_json_nulls(self):
        campaign = Campaign.objects.create(name="Context Campaign")
        original_summary = {"regime_counts": {"stable_low": 1}}
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Complete.",
            summary_json=original_summary.copy(),
        )

        response = self.client.post(
            reverse("campaigns:campaign_research_context", args=[campaign.pk]),
            {
                "floor_level": "2",
                "height_above_ground_m": "6.5",
                "room_volume_m3": "48.25",
                "room_volume_source": "ESTIMATED",
                "dominant_material": "tuff stone",
                "sensor_height_m": "1.2",
                "distance_from_nearest_opening_m": "",
                "sensor_moved_during_campaign": "NO",
                "direct_connection_to_soil": "UNKNOWN",
                "event_log_availability": "PARTIAL",
                "notes": "Research-context metadata is provisional.",
                "evidence_status": "ESTIMATED",
            },
        )

        self.assertRedirects(response, reverse("campaigns:campaign_research_context", args=[campaign.pk]))
        context = CampaignResearchContext.objects.get(campaign=campaign)
        self.assertEqual(context.floor_level, 2)
        self.assertEqual(str(context.room_volume_m3), "48.250")
        self.assertIsNone(context.distance_from_nearest_opening_m)

        json_response = self.client.get(reverse("campaigns:research_context_json", args=[campaign.pk]))
        payload = json_response.json()
        self.assertEqual(payload["metadata"]["dominant_material"], "tuff stone")
        self.assertIsNone(payload["metadata"]["distance_from_nearest_opening_m"])
        self.assertTrue(payload["readiness"]["grey_box_modelling_possible"])
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, original_summary)

    def test_documented_events_page_and_exports_render_default_radoneye_cycles(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 — pilot regime analysis")
        self._create_hourly_measurements(
            campaign,
            timezone.datetime(2024, 5, 24, 0, 47, tzinfo=timezone.get_current_timezone()),
            24 * 23,
            start_value=40,
        )

        response = self.client.get(reverse("campaigns:documented_events", args=[campaign.pk]))
        json_response = self.client.get(reverse("campaigns:documented_events_json", args=[campaign.pk]))
        csv_response = self.client.get(reverse("campaigns:documented_events_csv", args=[campaign.pk]))
        excel_response = self.client.get(reverse("campaigns:documented_events_excel", args=[campaign.pk]))

        self.assertContains(response, "Documented Event Analysis")
        self.assertContains(response, "Cycle 1")
        self.assertContains(response, "Cycle 2")
        self.assertContains(response, "documented ventilation period")
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual(len(json_response.json()["cycles"]), 2)
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn("text/csv", csv_response["Content-Type"])
        self.assertEqual(excel_response.status_code, 200)
        workbook = load_workbook(BytesIO(excel_response.content))
        self.assertIn("Documented Event Summary", workbook.sheetnames)
        self.assertIn("Phase Metrics", workbook.sheetnames)

    def test_documented_event_complete_cycle_metrics_are_reproducible(self):
        campaign = Campaign.objects.create(name="Documented Complete")
        start = timezone.datetime(2024, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        for index, value in enumerate([10, 20, 30, 40, 80, 120, 90, 40, 20, 20, 30, 40]):
            Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal(str(value)))
        definition = EventCycleDefinition.from_mapping(
            {
                "cycle_label": "Complete",
                "baseline_start": "2024-01-01 00:00",
                "baseline_end": "2024-01-01 02:00",
                "accumulation_start": "2024-01-01 03:00",
                "accumulation_end": "2024-01-01 06:00",
                "rapid_removal_start": "2024-01-01 06:00",
                "rapid_removal_end": "2024-01-01 08:00",
                "post_event_start": "2024-01-01 09:00",
                "post_event_end": "2024-01-01 11:00",
                "evidence_status": "DOCUMENTED",
            }
        )

        payload = analyse_documented_cycles(campaign, [definition])
        cycle = payload["cycles"][0]

        self.assertEqual(cycle["metrics"]["baseline"]["observation_count"], 3)
        self.assertEqual(cycle["metrics"]["baseline"]["mean"], 20.0)
        self.assertEqual(cycle["metrics"]["accumulation"]["maximum_concentration"], 120.0)
        self.assertEqual(cycle["metrics"]["accumulation"]["endpoint_change"], 50.0)
        self.assertEqual(cycle["metrics"]["accumulation"]["ordinary_linear_slope"], 16.667)
        self.assertEqual(cycle["metrics"]["rapid_removal"]["absolute_decrease"], 70.0)
        self.assertEqual(cycle["metrics"]["rapid_removal"]["time_to_first_lte_50_bq_m3_hours"], 1.0)
        self.assertEqual(cycle["metrics"]["rapid_removal"]["time_to_first_lte_30_bq_m3_hours"], 2.0)
        self.assertEqual(cycle["metrics"]["post_event"]["median"], 30.0)

    def test_documented_event_reports_missing_observations_and_cross_midnight_phase(self):
        campaign = Campaign.objects.create(name="Documented Missing")
        start = timezone.datetime(2024, 1, 1, 22, 0, tzinfo=timezone.get_current_timezone())
        for index, value in [(0, 100), (1, 110), (3, 130), (4, 80)]:
            Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal(str(value)))
        definition = EventCycleDefinition.from_mapping(
            {
                "cycle_label": "Cross midnight",
                "baseline_start": "2024-01-01 22:00",
                "baseline_end": "2024-01-02 02:00",
                "accumulation_start": "2024-01-01 22:00",
                "accumulation_end": "2024-01-02 02:00",
                "rapid_removal_start": "2024-01-02 01:00",
                "rapid_removal_end": "2024-01-02 02:00",
                "post_event_start": "2024-01-02 01:00",
                "post_event_end": "2024-01-02 02:00",
            }
        )

        cycle = analyse_documented_cycles(campaign, [definition])["cycles"][0]

        self.assertGreater(cycle["phase_quality"]["baseline"]["missing_observations"], 0)
        self.assertGreater(len(cycle["phase_quality"]["baseline"]["gaps"]), 0)
        self.assertTrue(cycle["quality_affected"])

    def test_documented_event_thresholds_can_be_unreached_and_parse_aware_or_naive_input(self):
        campaign = Campaign.objects.create(name="Documented Threshold")
        start = timezone.datetime(2024, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        for index, value in enumerate([100, 90, 80]):
            Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal(str(value)))
        aware = parse_event_timestamp("2024-01-01T00:00:00+00:00")
        naive = parse_event_timestamp("2024-01-01 00:00")
        definition = EventCycleDefinition(
            cycle_label="No thresholds",
            baseline_start=naive,
            baseline_end=parse_event_timestamp("2024-01-01 00:00"),
            accumulation_start=aware,
            accumulation_end=parse_event_timestamp("2024-01-01 02:00"),
            rapid_removal_start=parse_event_timestamp("2024-01-01 00:00"),
            rapid_removal_end=parse_event_timestamp("2024-01-01 02:00"),
            post_event_start=parse_event_timestamp("2024-01-01 02:00"),
            post_event_end=parse_event_timestamp("2024-01-01 02:00"),
        )

        cycle = analyse_documented_cycles(campaign, [definition])["cycles"][0]

        self.assertTrue(timezone.is_aware(aware))
        self.assertTrue(timezone.is_aware(naive))
        self.assertIsNone(cycle["metrics"]["rapid_removal"]["time_to_first_lte_50_bq_m3_hours"])
        self.assertIsNone(cycle["metrics"]["rapid_removal"]["time_to_first_lte_30_bq_m3_hours"])

    def test_documented_event_analysis_does_not_mutate_stored_report_data(self):
        campaign = Campaign.objects.create(name="Documented No Mutation")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Complete.",
            summary_json={"regime_counts": {"stable_low": 1}},
        )
        Measurement.objects.create(
            campaign=campaign,
            measured_at=timezone.datetime(2024, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone()),
            radon_bq_m3=Decimal("10"),
        )

        self.client.get(reverse("campaigns:documented_events_json", args=[campaign.pk]))

        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"regime_counts": {"stable_low": 1}})

    def test_default_radoneye_cycle_definitions_are_reproducible(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 — pilot regime analysis")
        definitions = default_event_cycles_for_campaign(campaign)

        self.assertEqual([definition.cycle_label for definition in definitions], ["Cycle 1", "Cycle 2"])
        self.assertEqual(definitions[0].accumulation_start.isoformat(), "2024-05-25T09:47:00+00:00")
        self.assertEqual(definitions[1].rapid_removal_end.isoformat(), "2024-06-14T01:44:00+00:00")

    def test_baseline_prediction_experiment_page_and_exports_are_read_only(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 pilot regime analysis")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Existing report.",
            summary_json={"existing_prediction_payload": {"keep": True}},
        )
        self._create_radoneye_default_cycle_measurements(campaign)
        measurement_count = campaign.measurements.count()

        response = self.client.get(reverse("campaigns:baseline_prediction_experiment", args=[campaign.pk]))
        json_response = self.client.get(reverse("campaigns:baseline_prediction_experiment_json", args=[campaign.pk]))
        csv_response = self.client.get(reverse("campaigns:baseline_prediction_experiment_csv", args=[campaign.pk]))
        excel_response = self.client.get(reverse("campaigns:baseline_prediction_experiment_excel", args=[campaign.pk]))

        self.assertContains(response, "Baseline Prediction Experiment")
        self.assertContains(response, "Experiment A")
        self.assertContains(response, "Experiment B")
        self.assertContains(response, "Focused Rapid-Removal Plot")
        self.assertContains(response, "Forecasts are unconditional on future occupant actions")
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual({experiment["label"] for experiment in json_response.json()["experiments"]}, {"Experiment A", "Experiment B"})
        self.assertEqual(csv_response.status_code, 200)
        self.assertIn("text/csv", csv_response["Content-Type"])
        workbook = load_workbook(BytesIO(excel_response.content))
        self.assertIn("Configuration", workbook.sheetnames)
        self.assertIn("Overall Metrics", workbook.sheetnames)
        self.assertIn("Phase Metrics", workbook.sheetnames)
        self.assertIn("Forecast Rows", workbook.sheetnames)
        self.assertIn("Exclusions", workbook.sheetnames)
        self.assertIn("Rapid Removal Notes", workbook.sheetnames)
        self.assertIn("Target Phase Metrics", workbook.sheetnames)
        self.assertIn("Fair Comparison", workbook.sheetnames)
        self.assertIn("Intervention Audit", workbook.sheetnames)
        self.assertIn("Alpha Audit", workbook.sheetnames)
        self.assertIn("Table A Overall", workbook.sheetnames)
        self.assertIn("Table B Fair", workbook.sheetnames)
        self.assertEqual(campaign.measurements.count(), measurement_count)
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"existing_prediction_payload": {"keep": True}})

    def test_baseline_prediction_experiment_uses_direct_horizons_and_reverse_cycle(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 pilot regime analysis")
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = baseline_experiment.run_baseline_prediction_experiment(campaign)

        self.assertEqual(payload["horizons"], ["1h", "3h", "6h"])
        experiments = {experiment["label"]: experiment for experiment in payload["experiments"]}
        self.assertEqual(experiments["Experiment A"]["training_cycle"]["cycle_label"], "Cycle 1")
        self.assertEqual(experiments["Experiment A"]["test_cycle"]["cycle_label"], "Cycle 2")
        self.assertEqual(experiments["Experiment B"]["training_cycle"]["cycle_label"], "Cycle 2")
        self.assertEqual(experiments["Experiment B"]["test_cycle"]["cycle_label"], "Cycle 1")
        for row in experiments["Experiment A"]["forecast_rows"][:30]:
            origin = baseline_experiment.datetime_from_iso(row["timestamp"])
            target = baseline_experiment.datetime_from_iso(row["target_timestamp"])
            self.assertEqual((target - origin).total_seconds() / 3600, int(row["horizon"].removesuffix("h")))
        self.assertTrue(experiments["Experiment A"]["phase_metrics"])
        self.assertTrue(experiments["Experiment A"]["rapid_removal_diagnostics"]["forecast_observed_at_horizons"])
        self.assertNotIn("mape", experiments["Experiment A"]["overall_metrics"][0])

    def test_baseline_prediction_lag_features_use_history_only(self):
        campaign = Campaign.objects.create(name="Lag Construction")
        start = timezone.datetime(2024, 1, 1, 0, 0, tzinfo=datetime_timezone.utc)
        for index in range(20):
            Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal(str(index)))
        definition = self._simple_cycle_definition(start, 19)
        rows = baseline_experiment._measurement_rows(campaign)

        examples, exclusions = baseline_experiment._examples(
            rows,
            definition,
            (start, start + timedelta(hours=19)),
            timedelta(hours=3),
            1.0,
            3,
            training=True,
        )

        example = next(row for row in examples if row["origin_time"] == start + timedelta(hours=12))
        self.assertEqual(example["target_time"], start + timedelta(hours=15))
        self.assertEqual(example["features"]["lag_1"], 12.0)
        self.assertEqual(example["features"]["lag_2"], 11.0)
        self.assertEqual(example["features"]["lag_3"], 10.0)
        self.assertEqual(example["features"]["lag_6"], 7.0)
        self.assertEqual(example["features"]["lag_12"], 1.0)
        self.assertEqual(example["features"]["diff_1"], 1.0)
        self.assertEqual(example["features"]["rolling_mean_3"], 11.0)
        self.assertEqual(example["features"]["rolling_mean_6"], 9.5)
        self.assertGreater(exclusions["insufficient_history"], 0)

    def test_baseline_prediction_excludes_gaps_and_quality_affected_targets(self):
        campaign = Campaign.objects.create(name="Gap Exclusions")
        start = timezone.datetime(2024, 1, 1, 0, 0, tzinfo=datetime_timezone.utc)
        for index in range(18):
            if index == 7:
                continue
            Measurement.objects.create(
                campaign=campaign,
                measured_at=start + timedelta(hours=index),
                radon_bq_m3=Decimal(str(100 + index)),
                regime="quality_affected" if index == 12 else "",
            )
        definition = self._simple_cycle_definition(start, 17)
        rows = baseline_experiment._measurement_rows(campaign)

        _examples, exclusions = baseline_experiment._examples(
            rows,
            definition,
            (start, start + timedelta(hours=17)),
            timedelta(hours=3),
            1.0,
            3,
            training=False,
        )

        self.assertGreater(exclusions["missing_future_target"], 0)
        self.assertGreater(exclusions["crosses_gap"], 0)
        self.assertGreater(exclusions["quality_affected_target"], 0)

    def test_baseline_prediction_near_zero_values_are_supported_and_deterministic(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 near zero")
        self._create_radoneye_default_cycle_measurements(campaign, near_zero=True)

        first_payload = baseline_experiment.run_baseline_prediction_experiment(campaign)
        second_payload = baseline_experiment.run_baseline_prediction_experiment(campaign)

        self.assertEqual(first_payload["experiments"][0]["overall_metrics"], second_payload["experiments"][0]["overall_metrics"])
        self.assertEqual(first_payload["experiments"][1]["selected_alphas"], second_payload["experiments"][1]["selected_alphas"])
        self.assertTrue(first_payload["experiments"][0]["overall_metrics"])
        self.assertTrue(first_payload["experiments"][0]["rapid_removal_plot"]["observed_points"])

    def test_baseline_prediction_audit_tables_prevent_duplicate_and_phase_ambiguity(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 audit")
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = baseline_experiment.run_baseline_prediction_experiment(campaign)
        rows = baseline_experiment.flatten_forecasts(payload)
        keys = {
            (row["experiment"], row["model"], row["horizon"], row["timestamp"], row["target_timestamp"])
            for row in rows
        }

        self.assertEqual(len(keys), len(rows))
        self.assertEqual(payload["duplicate_forecast_key_count"], 0)
        self.assertTrue(payload["leakage_audit"]["forecast_origin_before_target"])
        self.assertTrue(payload["leakage_audit"]["direct_horizon_construction"])
        self.assertTrue(baseline_experiment.flatten_fair_comparison_metrics(payload))
        self.assertTrue(baseline_experiment.flatten_target_phase_metrics(payload))
        self.assertTrue(payload["paper_ready_tables"]["table_b_fair_comparison_performance"])
        self.assertEqual(
            payload["experiments"][0]["rapid_removal_diagnostics"]["predicted_time_to_lte_50_hours_by_model"],
            baseline_experiment.THRESHOLD_NOT_IDENTIFIABLE,
        )
        for experiment in payload["experiments"]:
            for row in experiment["fair_comparison_metrics"]:
                self.assertEqual(row["model_specific_target_count"], row["common_target_count"])

    def test_baseline_prediction_alpha_audit_is_deterministic_and_training_only(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 alpha audit")
        self._create_radoneye_default_cycle_measurements(campaign)

        first_payload = baseline_experiment.run_baseline_prediction_experiment(campaign)
        second_payload = baseline_experiment.run_baseline_prediction_experiment(campaign)

        self.assertEqual(
            [experiment["selected_alphas"] for experiment in first_payload["experiments"]],
            [experiment["selected_alphas"] for experiment in second_payload["experiments"]],
        )
        alpha_rows = baseline_experiment.flatten_alpha_audit(first_payload)
        self.assertTrue(alpha_rows)
        self.assertTrue(all(row["validation_rule"] == "blocked chronological validation inside the training cycle" for row in alpha_rows))
        self.assertTrue(any(row["selected"] for row in alpha_rows))
        for row in alpha_rows:
            self.assertLessEqual(row["training_start"], row["training_end"])
            self.assertLessEqual(row["validation_start"], row["validation_end"])
            self.assertLess(row["training_end"], row["validation_start"])

    def test_apparent_dynamics_exact_recovery_and_fixed_floor_endpoint(self):
        ceq = 240.0
        kappa = 0.23
        c0 = 80.0
        times = [0, 1, 2, 4, 7, 11]
        values = [ceq + (c0 - ceq) * pow(2.718281828459045, -kappa * t) for t in times]

        free_fit = apparent_dynamics_audit.fit_free_equilibrium(times, values)
        floor = 40.0
        removal_start = 260.0
        removal_values = [floor + (removal_start - floor) * pow(2.718281828459045, -kappa * t) for t in times]
        floor_fit = apparent_dynamics_audit.fit_fixed_floor(times, removal_values, floor)
        endpoint = apparent_dynamics_audit.endpoint_kappa(removal_values[0], removal_values[-1], floor, times[-1] - times[0])

        self.assertAlmostEqual(free_fit["kappa"], kappa, places=2)
        self.assertAlmostEqual(free_fit["ceq"], ceq, places=1)
        self.assertAlmostEqual(floor_fit["kappa"], kappa, places=2)
        self.assertAlmostEqual(endpoint, kappa, places=2)

    def test_apparent_dynamics_endpoint_not_identifiable_and_flat_flags(self):
        self.assertEqual(apparent_dynamics_audit.endpoint_kappa(100, 110, 50, 2), apparent_dynamics_audit.NOT_IDENTIFIABLE)
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 flat apparent")
        definitions = default_event_cycles_for_campaign(campaign)
        start = definitions[0].baseline_start
        for index in range(8):
            Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal("100"))

        payload = apparent_dynamics_audit.run_apparent_dynamics_audit(campaign)
        baseline_d1 = next(row for row in payload["phase_fit_results"] if row["cycle_label"] == "Cycle 1" and row["phase"] == "baseline" and "D1" in row["fit_name"])

        self.assertIn("INSUFFICIENT_DYNAMIC_RANGE", baseline_d1["flags"])
        self.assertEqual(payload["manifest"]["yellow_tuff_measurement_status"], apparent_dynamics_audit.YELLOW_TUFF_STATUS)

    def test_apparent_dynamics_volume_sensitivity_is_read_only_and_linear(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 apparent volume")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Keep me.",
            summary_json={"regime_counts": {"stable_low": 1}},
        )
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = apparent_dynamics_audit.run_apparent_dynamics_audit(campaign)
        rows = [row for row in payload["volume_sensitivity"] if row["cycle_label"] == "Cycle 1" and row["phase"] == "accumulation" and row["fit_name"] == "accumulation_D1_free_equilibrium"]
        low = next(row for row in rows if float(row["volume_m3"]) == 100.0)
        high = next(row for row in rows if float(row["volume_m3"]) == 150.0)

        self.assertEqual(low["kappa_h_minus_1"], high["kappa_h_minus_1"])
        self.assertEqual(low["C_eq_bq_m3"], high["C_eq_bq_m3"])
        self.assertAlmostEqual(
            float(high["effective_net_source_loading_Bq_per_h"]) / float(low["effective_net_source_loading_Bq_per_h"]),
            1.5,
            places=3,
        )
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"regime_counts": {"stable_low": 1}})

    def test_apparent_dynamics_routes_and_exports_render(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 apparent routes")
        self._create_radoneye_default_cycle_measurements(campaign)

        page = self.client.get(reverse("campaigns:apparent_dynamics_audit", args=[campaign.pk]))
        json_response = self.client.get(reverse("campaigns:apparent_dynamics_audit_json", args=[campaign.pk]))
        csv_response = self.client.get(reverse("campaigns:apparent_dynamics_audit_csv", args=[campaign.pk]))
        excel_response = self.client.get(reverse("campaigns:apparent_dynamics_audit_excel", args=[campaign.pk]))

        self.assertContains(page, "Apparent Dynamics Audit")
        self.assertContains(page, "Apparent relaxation rates are not ACH")
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual(json_response.json()["manifest"]["yellow_tuff_measurement_status"], apparent_dynamics_audit.YELLOW_TUFF_STATUS)
        self.assertEqual(csv_response.status_code, 200)
        workbook = load_workbook(BytesIO(excel_response.content))
        self.assertIn("Phase Fits", workbook.sheetnames)
        self.assertIn("Boundary Sensitivity", workbook.sheetnames)
        self.assertIn("Volume Sensitivity", workbook.sheetnames)

    def test_reduced_state_space_kalman_equations_and_exact_recovery(self):
        posterior, posterior_var = state_space_experiment.kalman_update(100.0, 25.0, 120.0, 25.0)

        self.assertAlmostEqual(posterior, 110.0)
        self.assertAlmostEqual(posterior_var, 12.5)

        pairs = [(100.0, 95.0, 1.0), (95.0, 91.0, 1.0), (91.0, 87.8, 1.0), (87.8, 85.24, 1.0)]
        fit = state_space_experiment._fit_ab_for_pairs(pairs, 1.0)

        self.assertAlmostEqual(fit["a"], 0.8, places=3)
        self.assertAlmostEqual(fit["b"], 15.0, places=3)

    def test_reduced_state_space_no_leakage_q_selection_and_modes(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 state space")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Keep report.",
            summary_json={"locked": True},
        )
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = state_space_experiment.run_reduced_state_space_experiment(campaign)
        first = payload["experiments"][0]
        train_end = state_space_experiment._parse_iso(first["training_cycle"]["post_event_end"])
        test_start = state_space_experiment._parse_iso(first["test_cycle"]["baseline_start"])

        self.assertLess(train_end, test_start)
        self.assertEqual(payload["duplicate_forecast_key_count"], 0)
        self.assertTrue(all(row["forecast_origin"] < row["target_timestamp"] for row in payload["forecast_rows"]))
        self.assertTrue(all(row["validation_rule"].startswith("blocked chronological validation") for row in payload["q_selection_audit"]))
        self.assertTrue(any(row["selected"] for row in payload["q_selection_audit"]))
        self.assertTrue(any(row["forecast_mode"].startswith("F1") for row in payload["forecast_rows"]))
        self.assertTrue(any(row["forecast_mode"].startswith("F2") for row in payload["forecast_rows"]))
        self.assertTrue(
            any(
                row["event_knowledge_status"] == "no_future_event_knowledge_closed_transition_used"
                for row in payload["forecast_rows"]
            )
        )
        self.assertTrue(
            any(
                row["event_knowledge_status"] == "known_intervention_transition_used_after_declared_start"
                for row in payload["forecast_rows"]
            )
        )
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"locked": True})

    def test_reduced_state_space_routes_exports_and_missing_targets(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 state routes")
        self._create_radoneye_default_cycle_measurements(campaign)
        Measurement.objects.filter(campaign=campaign).order_by("measured_at")[20].delete()

        page = self.client.get(reverse("campaigns:reduced_state_space_experiment", args=[campaign.pk]))
        json_response = self.client.get(reverse("campaigns:reduced_state_space_experiment_json", args=[campaign.pk]))
        csv_response = self.client.get(reverse("campaigns:reduced_state_space_experiment_csv", args=[campaign.pk]))
        excel_response = self.client.get(reverse("campaigns:reduced_state_space_experiment_excel", args=[campaign.pk]))

        self.assertContains(page, "Reduced State-Space Experiment")
        self.assertContains(page, "Apparent dynamics are not ACH")
        payload = json_response.json()
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual(payload["manifest"]["state_vector"], "x_k=[C_k]")
        self.assertTrue(any("MISSING_OBSERVATION" in row["validity_flags"] for row in payload["forecast_rows"]))
        self.assertEqual(csv_response.status_code, 200)
        workbook = load_workbook(BytesIO(excel_response.content))
        self.assertIn("Overall Performance", workbook.sheetnames)
        self.assertIn("Fair Comparison", workbook.sheetnames)
        self.assertIn("Forecast Rows", workbook.sheetnames)

    def test_state_space_validation_open_loop_local_level_and_no_mutation(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 validation audit")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Locked.",
            summary_json={"do_not_change": True},
        )
        self._create_radoneye_default_cycle_measurements(campaign)

        first = state_space_validation_audit.run_state_space_validation_audit(campaign)
        second = state_space_validation_audit.run_state_space_validation_audit(campaign)

        self.assertEqual(first["ablation_performance"], second["ablation_performance"])
        self.assertTrue(any(row["model"] == "open_loop_reduced_transition" for row in first["ablation_performance"]))
        self.assertTrue(any(row["model"] == "generic_local_level_kalman" for row in first["ablation_performance"]))
        self.assertTrue(any(row["model"] == "reduced_sirem_informed_kalman" for row in first["ablation_performance"]))
        self.assertTrue(any(row["answer"] in {"yes", "no"} for row in first["scientific_decision"]))
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"do_not_change": True})

    def test_state_space_validation_q_f1_and_interval_semantics(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 validation semantics")
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = state_space_validation_audit.run_state_space_validation_audit(campaign)

        self.assertTrue(all("blocked chronological validation" in row["validation_rule"] for row in payload["local_level_q_selection_audit"]))
        self.assertTrue(any(row["selected"] for row in payload["local_level_q_selection_audit"]))
        self.assertTrue(payload["event_knowledge_audit"])
        self.assertTrue(all(not row["future_phase_label_used_for_transition"] for row in payload["event_knowledge_audit"]))
        self.assertTrue(all(not row["rapid_removal_transition_used_before_event"] for row in payload["event_knowledge_audit"]))
        calibration = payload["uncertainty_calibration"]
        self.assertTrue(calibration)
        self.assertTrue(all("Future-observation interval includes observation noise R" in row["semantic_note"] for row in calibration))
        self.assertTrue(
            any(
                row["future_observation_mean_interval_width"] > row["latent_state_mean_interval_width"]
                for row in calibration
                if row["future_observation_mean_interval_width"] and row["latent_state_mean_interval_width"]
            )
        )

    def test_state_space_validation_boundary_determinism_and_exports(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 validation outputs")
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = state_space_validation_audit.run_state_space_validation_audit(campaign)
        workbook = state_space_validation_audit.build_state_space_validation_workbook(payload)

        self.assertTrue(payload["rapid_parameter_stability"])
        self.assertEqual(
            payload["rapid_parameter_stability"],
            state_space_validation_audit.run_state_space_validation_audit(campaign)["rapid_parameter_stability"],
        )
        loaded = load_workbook(BytesIO(workbook.getvalue()))
        self.assertIn("Ablation Performance", loaded.sheetnames)
        self.assertIn("Uncertainty Calibration", loaded.sheetnames)
        self.assertIn("Fair Comparison", loaded.sheetnames)

    def test_final_state_space_interval_semantics_and_forecast_means(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 final intervals")
        self._create_radoneye_default_cycle_measurements(campaign)

        state_payload = state_space_experiment.run_reduced_state_space_experiment(campaign)
        row = next(
            item
            for item in state_payload["forecast_rows"]
            if item["valid_status"] == "valid" and item["R_scenario"] == state_space_experiment.PRIMARY_R_SCENARIO
        )
        r_value = state_space_experiment.R_SCENARIOS[row["R_scenario"]]["observation_variance"]

        self.assertAlmostEqual(row["future_observation_variance"], row["latent_state_variance"] + r_value, places=3)
        self.assertEqual(row["interval_semantics"], "future_observation_predictive_interval_includes_R")
        self.assertIn("latent_state_lower_interval", row)
        self.assertIn("future_observation_lower_interval", row)

        validation = state_space_validation_audit.run_state_space_validation_audit(campaign)
        match = next(
            item
            for item in validation["fair_comparison"]
            if item["experiment"] == row["experiment"]
            and item["forecast_mode"] == row["forecast_mode"]
            and item["horizon"] == row["horizon"]
            and item["model"] == "reduced_sirem_informed_kalman"
        )
        self.assertIsNotNone(match["MAE"])

    def test_final_model_results_are_deterministic_and_read_only(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 final package")
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Locked.",
            summary_json={"baseline_locked": True},
        )
        self._create_radoneye_default_cycle_measurements(campaign)

        first = final_model_results.build_final_model_results(campaign)
        second = final_model_results.build_final_model_results(campaign)

        self.assertEqual(first["table_1_overall_performance"], second["table_1_overall_performance"])
        self.assertEqual(first["table_2_sequential_ablation"], second["table_2_sequential_ablation"])
        self.assertEqual(first["table_3_predictive_uncertainty"], second["table_3_predictive_uncertainty"])
        self.assertTrue(first["table_1_overall_performance"])
        self.assertTrue(first["model_validity_summary"])
        self.assertIn("Future-observation predictive interval", first["interpretation_markdown"])
        self.assertEqual(AnalysisReport.objects.get(pk=report.pk).summary_json, {"baseline_locked": True})

    def test_final_model_workbook_contains_locked_tables(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 final workbook")
        self._create_radoneye_default_cycle_measurements(campaign)

        payload = final_model_results.build_final_model_results(campaign)
        workbook = final_model_results.build_final_results_workbook(payload)
        loaded = load_workbook(BytesIO(workbook.getvalue()))

        self.assertIn("Table 1 Overall", loaded.sheetnames)
        self.assertIn("Table 3 Uncertainty", loaded.sheetnames)
        self.assertIn("Model Validity", loaded.sheetnames)
        self.assertEqual(loaded["Table 3 Uncertainty"]["D1"].value, "Nominal coverage")

    def _create_hourly_measurements(self, campaign, start, count, start_value=100):
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(start_value + (index % 180))),
                )
                for index in range(count)
            ]
        )

    def _create_radoneye_default_cycle_measurements(self, campaign, near_zero=False):
        rows = []
        for cycle in default_event_cycles_for_campaign(campaign):
            start, end = baseline_experiment._cycle_window(cycle)
            current = start
            while current <= end:
                phase = baseline_experiment._phase_for_time(cycle, current)
                phase_start = getattr(cycle, f"{phase}_start")
                phase_hour = int((current - phase_start).total_seconds() / 3600)
                if near_zero and phase == "baseline":
                    value = phase_hour % 3
                elif phase == "baseline":
                    value = 42 + (phase_hour % 6)
                elif phase == "accumulation":
                    value = 75 + min(185, phase_hour * 2)
                elif phase == "rapid_removal":
                    value = max(18, 250 - phase_hour * 82)
                else:
                    value = 28 + (phase_hour % 8)
                rows.append(Measurement(campaign=campaign, measured_at=current, radon_bq_m3=Decimal(str(value))))
                current += timedelta(hours=1)
        Measurement.objects.bulk_create(rows)

    def _simple_cycle_definition(self, start, end_hour):
        return EventCycleDefinition(
            cycle_label="Synthetic",
            baseline_start=start,
            baseline_end=start + timedelta(hours=4),
            accumulation_start=start + timedelta(hours=5),
            accumulation_end=start + timedelta(hours=13),
            rapid_removal_start=start + timedelta(hours=14),
            rapid_removal_end=start + timedelta(hours=15),
            post_event_start=start + timedelta(hours=16),
            post_event_end=start + timedelta(hours=end_hour),
            evidence_status="SYNTHETIC",
        )

    def test_campaign_detail_links_to_excel_report(self):
        campaign = Campaign.objects.create(name="Excel Link Campaign")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Report ready.",
            summary_json={},
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertContains(response, "Download full audit report")
        self.assertContains(response, reverse("campaigns:export_excel_report", args=[campaign.pk]))
        self.assertContains(response, "Download compact report")

    def test_campaign_detail_does_not_embed_large_per_row_summary_arrays(self):
        campaign = Campaign.objects.create(name="Large Dashboard")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(80 + index % 30)),
                    segment_id=1,
                    regime="stable_low",
                )
                for index in range(600)
            ]
        )
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Large report.",
            summary_json={
                "measurement_count": 600,
                "segments": [],
                "regime_counts": {"stable_low": 600},
                "measurement_regimes_v2": [
                    {"timestamp": (start + timedelta(hours=index)).isoformat(), "large_marker": "do-not-render"}
                    for index in range(600)
                ],
                "canonical_hourly_data": [{"interval_start": index, "large_marker": "do-not-render"} for index in range(600)],
            },
            html_report="<article>" + ("large-html-marker " * 1000) + "</article>",
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertLess(len(response.content), 250_000)
        self.assertNotContains(response, "do-not-render")
        self.assertContains(response, "chart-data")

    def test_campaign_overview_limits_cards_episodes_and_warnings(self):
        campaign = Campaign.objects.create(name="Focused Overview")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.bulk_create(
            [
                Measurement(campaign=campaign, measured_at=start + timedelta(hours=index), radon_bq_m3=Decimal("100"))
                for index in range(12)
            ]
        )
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Focused report.",
            summary_json={
                "measurement_count": 12,
                "gap_count": 6,
                "concentration_level_counts": {"LOW": 7, "ELEVATED": 5},
                "episode_type_counts": {"ACCUMULATION": 4, "DECLINE": 3},
                "profile_applicability": {"status": "PROFILE_PARTIAL"},
                "quality_flag_counts": {"UNKNOWN_SENSOR_RESOLUTION": 1, "MISSING_TEMPERATURE": 12},
                "regime_confidence_summary": {"confidence_category_counts": {"LOW": 4, "HIGH": 8}},
                "episodes": [
                    {
                        "episode_type": "ACCUMULATION",
                        "start": f"2026-01-{index + 1:02d}T00:00:00+00:00",
                        "end": f"2026-01-{index + 1:02d}T03:00:00+00:00",
                        "duration_hours": 3.0,
                        "measurement_count": 4,
                        "max_radon": 100 + index,
                        "confidence_category": "HIGH",
                        "quality_status": "OK",
                    }
                    for index in range(12)
                ],
                "regime_counts": {"stable_low": 12},
            },
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(response.content.decode().count('class="summary-card"'), 8)
        self.assertLessEqual(response.content.decode().count(">View episode</a>"), 5)
        self.assertLessEqual(response.content.decode().count('class="warning-card'), 3)
        self.assertNotContains(response, "stable_low")
        self.assertContains(response, "Accumulation")
        self.assertContains(response, "Open Regimes & Episodes")
        self.assertContains(response, "Show more")
        self.assertContains(response, "12 Jan 2026, 00:00-03:00")
        self.assertNotContains(response, "2026-01-12T00:00:00+00:00")

    def test_campaign_overview_has_clear_card_actions_and_consistent_terminology(self):
        campaign = Campaign.objects.create(name="Polished Overview")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Polished report.",
            summary_json={
                "measurement_count": 5,
                "gap_count": 1,
                "concentration_level_counts": {"LOW": 3, "HIGH": 2},
                "dynamic_state_counts": {"STABLE_HIGH": 5},
                "episode_type_counts": {"DECLINE": 1},
                "profile_applicability": {"status": "PROFILE_PARTIAL"},
            },
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))
        content = response.content.decode()

        self.assertContains(response, "Run new analysis")
        self.assertLess(content.index("Run new analysis"), content.index("Add files"))
        self.assertContains(response, "View measurements")
        self.assertContains(response, "Review gaps")
        self.assertContains(response, "Review profile")
        self.assertContains(response, "Open Data Quality")
        self.assertContains(response, "Open regimes")
        self.assertContains(response, "View prediction results")
        self.assertContains(response, "Review sensitivity")
        self.assertContains(response, "View source files")
        self.assertContains(response, "Open reports")
        self.assertNotContains(response, ">M<", html=False)
        self.assertNotContains(response, ">MEAN<", html=False)
        self.assertNotContains(response, ">MAX<", html=False)
        self.assertNotContains(response, ">EH<", html=False)
        self.assertContains(response, "Sensitivity & Uncertainty")
        self.assertContains(response, "Sensitivity & Confidence")
        self.assertNotContains(response, "campaign-context-strip")
        self.assertLess(len(response.content), 100_000)

    def test_campaign_overview_uses_scientifically_consistent_metrics(self):
        campaign = Campaign.objects.create(name="Scientific Overview")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.create(
            campaign=campaign,
            measured_at=start,
            radon_bq_m3=Decimal("120"),
            temperature_c=Decimal("20.5"),
            humidity_percent=Decimal("50"),
            pressure_hpa=Decimal("1010"),
        )
        Measurement.objects.create(campaign=campaign, measured_at=start + timedelta(hours=1), radon_bq_m3=Decimal("130"))
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Scientific report.",
            summary_json={
                "measurement_count": 2,
                "profile_applicability": {"status": "PROFILE_COMPATIBLE_WITH_WARNINGS"},
                "analysis_config": {"concentration_low_threshold_bq_m3": 100, "concentration_high_threshold_bq_m3": 200},
                "concentration_level_counts": {"ELEVATED": 2},
                "dynamic_state_counts": {"QUALITY_AFFECTED": 10, "STABLE": 2},
                "sampling_gaps_compact_summary": {"long_gaps": 1, "moderate_gaps": 9},
                "regime_confidence_summary": {
                    "confidence_distribution_by_dynamic_state": {
                        "QUALITY_AFFECTED": {"HIGH": 10},
                        "STABLE": {"HIGH": 2},
                    }
                },
                "episodes": [
                    {
                        "episode_type": "QUALITY_AFFECTED",
                        "quality_status": "QUALITY_AFFECTED",
                        "duration_hours": 0,
                        "measurement_count": 1,
                        "max_radon": 500,
                        "confidence_category": "LOW",
                    },
                    {
                        "episode_type": "ACCUMULATION",
                        "quality_status": "OK",
                        "duration_hours": 3,
                        "measurement_count": 4,
                        "max_radon": 180,
                        "confidence_category": "HIGH",
                        "start": "2026-01-01T00:00:00+00:00",
                        "end": "2026-01-01T03:00:00+00:00",
                    },
                ],
            },
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertContains(response, "Compatible")
        self.assertContains(response, "With warnings")
        self.assertContains(response, "Temperature: 1 rows (50.0%)")
        self.assertContains(response, "Relative humidity: 1 rows (50.0%)")
        self.assertContains(response, "Pressure: 1 rows (50.0%)")
        self.assertContains(response, "CO2: not measured")
        self.assertContains(response, "Major gaps")
        self.assertContains(response, "Long campaign breaks")
        self.assertContains(response, "Episodes")
        self.assertContains(response, ">1</strong>", html=False)
        self.assertContains(response, "Accumulation")
        self.assertNotContains(response, "Data quality affected</td>")
        self.assertContains(response, "Profile categories: &gt;= 100 and &gt;= 200 Bq/m3", html=True)
        self.assertContains(response, "Displaying ${payload.returned_count.toLocaleString()} downsampled points")

    def test_chart_data_endpoint_limits_points_and_preserves_extremes(self):
        campaign = Campaign.objects.create(name="Chart Data Campaign")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        values = [100 + (index % 20) for index in range(300)]
        values[123] = 999
        values[200] = 5
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(value)),
                    segment_id=1 if index < 150 else 2,
                    regime="rising" if index % 2 else "stable_low",
                )
                for index, value in enumerate(values)
            ]
        )

        response = self.client.get(reverse("campaigns:campaign_chart_data", args=[campaign.pk]), {"max_points": 50})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        returned_values = {point["radon_bq_m3"] for point in payload["points"]}
        self.assertLessEqual(payload["returned_count"], 50)
        self.assertEqual(payload["source_count"], 300)
        self.assertIn(999.0, returned_values)
        self.assertIn(5.0, returned_values)

    def test_measurements_endpoint_is_paginated_and_filterable(self):
        campaign = Campaign.objects.create(name="Paged Measurements")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(100 + index)),
                    segment_id=1 if index < 40 else 2,
                    regime="rising" if index >= 40 else "stable_low",
                )
                for index in range(75)
            ]
        )

        response = self.client.get(
            reverse("campaigns:campaign_measurements", args=[campaign.pk]),
            {"page": 2, "page_size": 25, "segment_id": 1, "format": "json"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["count"], 40)
        self.assertEqual(payload["page"], 2)
        self.assertEqual(len(payload["results"]), 15)

    def test_campaign_detail_limits_large_summary_tables_to_twenty_rows(self):
        campaign = Campaign.objects.create(name="Large Tables Campaign")
        segments = []
        segment_v2 = []
        prediction_readiness = []
        ingestion_debug = []
        for index in range(30):
            segment_id = index + 1
            segments.append(
                {
                    "segment_id": segment_id,
                    "measurement_count": 10,
                    "segment_label": f"label_{segment_id}",
                    "percent_above_100": 0,
                    "percent_above_200": 0,
                    "dynamic_percent": 0,
                    "interpretation_text": f"segment-preview-{segment_id}",
                    "statistics": {"radon_bq_m3": {"mean": 100 + index, "max": 120 + index}},
                }
            )
            segment_v2.append(
                {
                    "segment_id": segment_id,
                    "start": "2026-01-01",
                    "end": "2026-01-02",
                    "mean_radon": 100,
                    "max_radon": 120,
                    "concentration_level_proportions": f"level-preview-{segment_id}",
                    "dynamic_state_proportions": "stable",
                    "episode_counts_by_type": {},
                    "segment_quality_score": 1,
                }
            )
            prediction_readiness.append(
                {
                    "segment_id": segment_id,
                    "category": f"readiness-preview-{segment_id}",
                    "prediction_readiness_score": 1,
                }
            )
            ingestion_debug.append(
                {
                    "filename": f"file-preview-{segment_id}.csv",
                    "detected_sheets": ["CSV"],
                    "raw_rows_read": 1,
                    "detected_header_row": 1,
                    "detected_columns": ["Time", "Radon"],
                    "mapped_columns": {"timestamp": "Time", "radon": "Radon"},
                    "parsed_measurement_rows": 1,
                    "skipped_reason": "",
                }
            )
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Large table report.",
            summary_json={
                "measurement_count": 0,
                "segments": segments,
                "segment_count": len(segments),
                "segment_v2_summaries": segment_v2,
                "prediction_readiness": prediction_readiness,
                "ingestion_debug": ingestion_debug,
                "source_file_inventory": [
                    {"filename": f"source-preview-{index + 1}.csv", "raw_rows": 1}
                    for index in range(30)
                ],
                "regime_counts": {"stable_low": 30},
            },
        )

        response = self.client.get(reverse("campaigns:campaign_regimes", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "segment-preview-21")

        provenance_response = self.client.get(reverse("campaigns:campaign_provenance", args=[campaign.pk]))
        self.assertContains(provenance_response, "file-preview-20.csv")
        self.assertNotContains(provenance_response, "file-preview-21.csv")
        self.assertContains(provenance_response, "Showing 20")

        prediction_response = self.client.get(reverse("campaigns:campaign_prediction", args=[campaign.pk]))
        self.assertContains(prediction_response, "readiness-preview-20")
        self.assertNotContains(prediction_response, "readiness-preview-21")

    def test_regimes_episode_table_distinguishes_raw_and_classification_trends_for_radoneye_interval(self):
        campaign = Campaign.objects.create(name="RadonEye Salerno 2024 — pilot regime analysis")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="RadonEye interval audit.",
            summary_json={
                "regime_counts": {},
                "concentration_level_counts": {},
                "dynamic_state_counts": {},
                "episode_type_counts": {"ACCUMULATION": 1, "DECLINE": 2, "SUDDEN_DROP_EVENT": 1},
                "episodes": [
                    {
                        "episode_type": "ACCUMULATION",
                        "legacy_episode_label": "accumulation",
                        "start": "2024-05-26T09:47:00+00:00",
                        "end": "2024-05-26T11:47:00+00:00",
                        "duration_hours": 2,
                        "starting_radon": 185.0,
                        "ending_radon": 170.0,
                        "absolute_concentration_change": -15.0,
                        "robust_episode_slope_bq_m3_per_hour": -7.5,
                        "mean_slope_bq_m3_per_hour": 1.889,
                        "dynamic_state_distribution": {"RISING": {"count": 3, "percent": 100.0}},
                        "confidence_category": "MEDIUM",
                        "dominant_reason_codes": ["SHORT_SLOPE_POSITIVE"],
                        "regime_algorithm_version": "regime_analysis_v2.2",
                        "episode_algorithm_version": "episode_analysis_v2.2",
                    },
                    {
                        "episode_type": "DECLINE",
                        "legacy_episode_label": "controlled or natural decline",
                        "start": "2024-05-26T12:47:00+00:00",
                        "end": "2024-05-26T14:47:00+00:00",
                        "duration_hours": 2,
                        "starting_radon": 161.0,
                        "ending_radon": 188.0,
                        "absolute_concentration_change": 27.0,
                        "robust_episode_slope_bq_m3_per_hour": 13.5,
                        "mean_slope_bq_m3_per_hour": -4.667,
                        "dynamic_state_distribution": {"FALLING": {"count": 3, "percent": 100.0}},
                        "confidence_category": "MEDIUM",
                        "dominant_reason_codes": ["SHORT_SLOPE_NEGATIVE"],
                        "regime_algorithm_version": "regime_analysis_v2.2",
                        "episode_algorithm_version": "episode_analysis_v2.2",
                    },
                    {
                        "episode_type": "DECLINE",
                        "legacy_episode_label": "controlled or natural decline",
                        "start": "2024-05-28T16:47:00+00:00",
                        "end": "2024-05-28T18:47:00+00:00",
                        "duration_hours": 2,
                        "starting_radon": 225.0,
                        "ending_radon": 227.0,
                        "absolute_concentration_change": 2.0,
                        "robust_episode_slope_bq_m3_per_hour": 1.0,
                        "mean_slope_bq_m3_per_hour": -10.555,
                        "dynamic_state_distribution": {"FALLING": {"count": 3, "percent": 100.0}},
                        "confidence_category": "MEDIUM",
                        "dominant_reason_codes": ["SHORT_SLOPE_NEGATIVE"],
                        "regime_algorithm_version": "regime_analysis_v2.2",
                        "episode_algorithm_version": "episode_analysis_v2.2",
                    },
                    {
                        "episode_type": "SUDDEN_DROP_EVENT",
                        "legacy_episode_label": "sudden drop",
                        "start": "2024-05-29T23:47:00+00:00",
                        "end": "2024-05-29T23:47:00+00:00",
                        "duration_hours": 0,
                        "measurement_count": 1,
                        "starting_radon": 46.0,
                        "ending_radon": 46.0,
                        "absolute_concentration_change": 0,
                        "robust_episode_slope_bq_m3_per_hour": None,
                        "mean_slope_bq_m3_per_hour": -55.333,
                        "dynamic_state_distribution": {"SUDDEN_DROP": {"count": 1, "percent": 100.0}},
                        "confidence_category": "HIGH",
                        "dominant_reason_codes": ["SUDDEN_NEGATIVE_CHANGE"],
                        "regime_algorithm_version": "regime_analysis_v2.2",
                        "episode_algorithm_version": "episode_analysis_v2.2",
                    },
                ],
            },
        )

        response = self.client.get(
            reverse("campaigns:campaign_regimes", args=[campaign.pk]),
            {"date_from": "2024-05-24", "date_to": "2024-05-31"},
        )

        self.assertContains(response, "V2 physical label")
        self.assertContains(response, "Legacy label")
        self.assertContains(response, "Raw endpoint delta")
        self.assertContains(response, "Raw endpoint slope")
        self.assertContains(response, "V2 classification trend slope")
        self.assertContains(response, "-7.5")
        self.assertContains(response, "1.889")
        self.assertContains(response, "13.5")
        self.assertContains(response, "-4.667")
        self.assertContains(response, "Raw start/end trend and v2 classification trend have opposite signs.")
        self.assertContains(response, "SUDDEN_DROP_EVENT")
        self.assertContains(response, "2024-05-29T23:47:00+00:00")

    def test_episode_table_endpoint_delta_is_computed_from_displayed_start_and_end_radon(self):
        campaign = Campaign.objects.create(name="RadonEye Endpoint Delta Audit")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Endpoint delta audit.",
            summary_json={
                "regime_counts": {},
                "concentration_level_counts": {},
                "dynamic_state_counts": {},
                "episode_type_counts": {"UNSTABLE_TRANSITION": 1, "STABLE_HIGH": 1},
                "episodes": [
                    {
                        "episode_type": "UNSTABLE_TRANSITION",
                        "legacy_episode_label": "unstable transition",
                        "start": "2024-05-28T06:47:00+00:00",
                        "end": "2024-05-28T15:47:00+00:00",
                        "duration_hours": 7.0,
                        "measurement_count": 10,
                        "starting_radon": 251.0,
                        "ending_radon": 273.0,
                        "absolute_concentration_change": -13.0,
                        "robust_episode_slope_bq_m3_per_hour": -4.333,
                        "mean_slope_bq_m3_per_hour": 1.0,
                        "dynamic_state_distribution": {"UNSTABLE_TRANSITION": {"count": 4, "percent": 100.0}},
                    },
                    {
                        "episode_type": "STABLE_HIGH",
                        "legacy_episode_label": "stable high",
                        "start": "2024-05-28T23:47:00+00:00",
                        "end": "2024-05-29T10:47:00+00:00",
                        "duration_hours": 7.0,
                        "measurement_count": 12,
                        "starting_radon": 228.0,
                        "ending_radon": 210.0,
                        "absolute_concentration_change": -3.0,
                        "robust_episode_slope_bq_m3_per_hour": -0.6,
                        "mean_slope_bq_m3_per_hour": 0.778,
                        "dynamic_state_distribution": {"STABLE": {"count": 6, "percent": 100.0}},
                    },
                ],
            },
        )

        response = self.client.get(
            reverse("campaigns:campaign_regimes", args=[campaign.pk]),
            {"date_from": "2024-05-28", "date_to": "2024-05-30"},
        )

        self.assertContains(response, "Elapsed span (h)")
        self.assertContains(response, "Effective observed duration (h)")
        self.assertContains(response, "Stored delta field")
        self.assertContains(response, "22.0")
        self.assertContains(response, "2.444")
        self.assertContains(response, "-18.0")
        self.assertContains(response, "-1.636")
        self.assertContains(response, "Stored delta field differs from displayed endpoint delta")
        self.assertContains(response, ">9.0<", html=False)
        self.assertContains(response, ">11.0<", html=False)
        self.assertContains(response, ">7.0<", html=False)

    def test_measurements_page_is_server_side_paginated(self):
        campaign = Campaign.objects.create(name="Measurement Page")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(100 + index)),
                    segment_id=1,
                    regime="stable_low",
                )
                for index in range(75)
            ]
        )

        response = self.client.get(reverse("campaigns:campaign_measurements", args=[campaign.pk]), {"page_size": 50})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rows: 75")
        self.assertContains(response, "Showing 50 of 75")
        self.assertContains(response, "149.00")
        self.assertNotContains(response, "150.00")

    def test_measurements_page_size_hard_limit_is_enforced(self):
        campaign = Campaign.objects.create(name="Measurement Hard Limit")
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.bulk_create(
            [
                Measurement(
                    campaign=campaign,
                    measured_at=start + timedelta(hours=index),
                    radon_bq_m3=Decimal(str(100 + index)),
                    segment_id=1,
                )
                for index in range(250)
            ]
        )

        response = self.client.get(
            reverse("campaigns:campaign_measurements", args=[campaign.pk]),
            {"page_size": 999, "format": "json"},
        )

        payload = response.json()
        self.assertEqual(payload["page_size"], 200)
        self.assertEqual(len(payload["results"]), 200)

    def test_campaign_detail_large_fixture_response_stays_below_one_mb(self):
        campaign = Campaign.objects.create(name="Campaign 7 Style Fixture")
        segments = [
            {
                "segment_id": index + 1,
                "measurement_count": 10,
                "segment_label": "low_stable",
                "percent_above_100": 0,
                "percent_above_200": 0,
                "dynamic_percent": 0,
                "interpretation_text": f"large-segment-{index + 1}",
                "statistics": {"radon_bq_m3": {"mean": 100, "max": 110}},
            }
            for index in range(600)
        ]
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Large fixture.",
            summary_json={
                "measurement_count": 600,
                "segment_count": 600,
                "segments": segments,
                "segment_v2_summaries": [{"segment_id": index + 1, "level": f"large-v2-{index + 1}"} for index in range(600)],
                "prediction_readiness": [{"segment_id": index + 1, "category": f"large-ready-{index + 1}"} for index in range(600)],
                "ingestion_debug": [{"filename": f"large-file-{index + 1}.csv"} for index in range(600)],
                "regime_counts": {"stable_low": 600},
            },
            html_report="<article>" + ("large report " * 10000) + "</article>",
        )

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertLess(len(response.content), 1_000_000)
        self.assertNotContains(response, "large-segment-21")
        self.assertNotContains(response, "large report large report")

    def test_campaign_detail_links_to_dedicated_paper1_analysis_form(self):
        campaign = Campaign.objects.create(name="Paper Form Campaign")

        response = self.client.get(reverse("campaigns:campaign_detail", args=[campaign.pk]))

        self.assertContains(response, "Run new analysis")
        self.assertNotContains(response, "Run full Paper 1 analysis")
        self.assertNotContains(response, "Europe/Rome")

        form_response = self.client.get(reverse("campaigns:run_campaign_analysis", args=[campaign.pk]))
        self.assertContains(form_response, "Run New Analysis")
        self.assertContains(form_response, "Europe/Rome")
        self.assertContains(form_response, "Advanced options")

    @patch("campaigns.views.run_paper1_analysis")
    def test_run_campaign_analysis_post_calls_shared_runner(self, runner):
        campaign = Campaign.objects.create(name="Paper Runner Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="runner.csv",
            file=SimpleUploadedFile("runner.csv", b"Time,Radon\n2026-01-01 00:00,100\n"),
        )
        runner.return_value = {
            "status": "success",
            "canonical_valid_rows": 1,
            "canonical_hourly_rows": 1,
        }

        response = self.client.post(
            reverse("campaigns:run_campaign_analysis", args=[campaign.pk]),
            {
                "timezone": "Europe/Rome",
                "resample": "1H",
                "gap_tolerance": "1.5",
                "rebuild_canonical": "on",
                "run_sensitivity": "on",
                "export_excel": "on",
            },
            follow=True,
        )

        self.assertRedirects(response, reverse("campaigns:campaign_detail", args=[campaign.pk]))
        runner.assert_called_once()
        self.assertContains(response, "Paper 1 analysis complete")

    @patch("campaigns.views.run_paper1_analysis")
    def test_run_campaign_analysis_invalid_gap_tolerance_is_graceful(self, runner):
        campaign = Campaign.objects.create(name="Invalid Gap Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="runner.csv",
            file=SimpleUploadedFile("runner.csv", b"Time,Radon\n2026-01-01 00:00,100\n"),
        )

        response = self.client.post(
            reverse("campaigns:run_campaign_analysis", args=[campaign.pk]),
            {
                "timezone": "Europe/Rome",
                "resample": "1H",
                "gap_tolerance": "-1",
            },
            follow=True,
        )

        runner.assert_not_called()
        self.assertContains(response, "Paper 1 analysis could not start")

    @patch("campaigns.views.run_paper1_analysis")
    def test_run_campaign_analysis_requires_uploaded_files(self, runner):
        campaign = Campaign.objects.create(name="No Upload Campaign")

        response = self.client.post(
            reverse("campaigns:run_campaign_analysis", args=[campaign.pk]),
            {
                "timezone": "Europe/Rome",
                "resample": "1H",
                "gap_tolerance": "1.5",
            },
            follow=True,
        )

        runner.assert_not_called()
        self.assertContains(response, "Upload at least one monitoring file")

    def test_missing_campaign_for_paper1_run_returns_404(self):
        response = self.client.post(
            reverse("campaigns:run_campaign_analysis", args=[999999]),
            {"timezone": "Europe/Rome", "resample": "1H", "gap_tolerance": "1.5"},
        )

        self.assertEqual(response.status_code, 404)

    def test_latest_paper1_summary_is_displayed(self):
        campaign = Campaign.objects.create(name="Latest Paper Summary")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Complete.",
            summary_json={
                "paper1_run_summary": {
                    "status": "success",
                    "run_timestamp": "2026-07-09T08:00:00+00:00",
                    "timezone": "Europe/Rome",
                    "resample": "1H",
                    "gap_tolerance": 1.5,
                    "rebuild_canonical": True,
                    "run_sensitivity": True,
                    "export_excel": True,
                    "raw_imported_rows": 10,
                    "exact_duplicate_rows_removed": 2,
                    "duplicate_conflict_rows": 1,
                    "canonical_valid_rows": 7,
                    "canonical_hourly_rows": 3,
                    "timezone_audit_rows": 7,
                    "dst_ambiguous_count": 1,
                    "dst_nonexistent_count": 0,
                    "total_sampling_irregularities": 4,
                    "short_gaps": 2,
                    "long_gaps": 1,
                    "regime_labels_found": ["stable_low"],
                    "prediction_horizons_evaluated": ["1h"],
                    "models_evaluated": ["naive_baseline"],
                    "small_sample_warning_count": 0,
                },
                "segments": [],
                "regime_counts": {"stable_low": 1},
            },
        )

        response = self.client.get(reverse("campaigns:campaign_quality", args=[campaign.pk]))

        self.assertContains(response, "timezone_audit_rows")
        self.assertContains(response, "dst_ambiguous_count")
        self.assertContains(response, "total_sampling_irregularities")

    def test_artifact_links_only_show_for_existing_expected_files(self):
        campaign = Campaign.objects.create(name="Artifact Campaign")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Complete.",
            summary_json={"segments": [], "regime_counts": {}},
        )
        with TemporaryDirectory() as tempdir:
            base_dir = Path(tempdir)
            output_dir = base_dir / "paper_outputs" / f"campaign_{campaign.pk}"
            output_dir.mkdir(parents=True)
            (output_dir / "paper1_validation_report.md").write_text("validation", encoding="utf-8")
            with override_settings(BASE_DIR=base_dir):
                response = self.client.get(reverse("campaigns:campaign_reports", args=[campaign.pk]))

        self.assertContains(response, "Open validation report")
        self.assertContains(response, "not generated yet")


class AnalysisPipelineTests(TestCase):
    def test_parse_decimal_accepts_decimal_commas(self):
        self.assertEqual(parse_decimal("123,45"), Decimal("123.45"))

    def test_csv_pipeline_merges_duplicates_segments_and_stores_measurements(self):
        campaign = Campaign.objects.create(name="Aranet CSV")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="aranet.csv",
            file=SimpleUploadedFile(
                "aranet.csv",
                (
                    "Time;Radon (Bq/m3);Temperature (C);Humidity (%);Pressure (hPa)\n"
                    "2026-01-01 00:00;100,5;20,1;45,2;1001,5\n"
                    "2026-01-01 00:00;;20,3;;\n"
                    "2026-01-01 00:30;110;20,4;46;1002\n"
                    "2026-01-01 02:00;130;21;48;1004\n"
                ).encode("utf-8"),
                content_type="text/csv",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.status, AnalysisReport.Status.COMPLETE)
        self.assertEqual(report.summary_json["measurement_count"], 3)
        self.assertEqual(report.summary_json["segment_count"], 2)
        self.assertEqual(report.summary_json["gap_count"], 1)

        measurements = list(campaign.measurements.order_by("measured_at"))
        self.assertEqual(len(measurements), 3)
        self.assertEqual(measurements[0].radon_bq_m3, Decimal("100.50"))
        self.assertEqual(measurements[0].temperature_c, Decimal("20.10"))
        self.assertEqual([measurement.segment_id for measurement in measurements], [1, 1, 2])

        first_segment = report.summary_json["segments"][0]
        self.assertEqual(first_segment["statistics"]["radon_bq_m3"]["count"], 2)
        self.assertEqual(first_segment["statistics"]["radon_bq_m3"]["mean"], 105.25)

    def test_xlsx_pipeline_detects_aranet_columns(self):
        campaign = Campaign.objects.create(name="Aranet XLSX")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Date time", "Radon", "Temp", "Humidity", "Pressure"])
        worksheet.append(["2026-02-01 10:00", "90", "19,5", "44", "998"])
        worksheet.append(["2026-02-01 10:30", "95", "19,7", "45", "999"])
        buffer = BytesIO()
        workbook.save(buffer)

        UploadedFile.objects.create(
            campaign=campaign,
            original_name="aranet.xlsx",
            file=SimpleUploadedFile(
                "aranet.xlsx",
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.status, AnalysisReport.Status.COMPLETE)
        self.assertEqual(report.summary_json["measurement_count"], 2)
        self.assertEqual(campaign.measurements.count(), 2)
        self.assertEqual(campaign.measurements.first().temperature_c, Decimal("19.50"))

    def test_xlsx_pipeline_skips_metadata_rows_before_aranet_header(self):
        campaign = Campaign.objects.create(name="Aranet Metadata XLSX")
        workbook = Workbook()
        metadata = workbook.active
        metadata.title = "About"
        metadata.append(["Aranet export"])
        metadata.append(["Device", "Airthings-like metadata"])
        data = workbook.create_sheet("Measurements")
        data.append(["Aranet4 data export"])
        data.append(["Sensor name", "Living room"])
        data.append(["Serial number", "12345"])
        data.append([])
        data.append(["Date and time", "Radon concentration (Bq/m³)", "Temperature (°C)", "Humidity (%)", "Pressure (hPa)"])
        data.append(["2026-05-01 08:00", "101,5", "20,2", "44,1", "1000,5"])
        data.append(["2026-05-01 09:00", "110,0", "20,3", "44,5", "1001,0"])
        buffer = BytesIO()
        workbook.save(buffer)

        uploaded = UploadedFile.objects.create(
            campaign=campaign,
            original_name="aranet-metadata.xlsx",
            file=SimpleUploadedFile(
                "aranet-metadata.xlsx",
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )

        rows, column_map, debug = read_uploaded_file(uploaded)
        report = run_campaign_analysis(campaign)

        self.assertEqual(len(rows), 2)
        self.assertEqual(column_map.timestamp, "Date and time")
        self.assertEqual(debug["detected_sheets"], ["About", "Measurements"])
        self.assertEqual(debug["selected_sheet"], "Measurements")
        self.assertEqual(debug["detected_header_row"], 5)
        self.assertEqual(debug["mapped_columns"]["radon"], "Radon concentration (Bq/m³)")
        self.assertEqual(debug["parsed_measurement_rows"], 2)
        self.assertEqual(report.summary_json["measurement_count"], 2)
        self.assertEqual(report.summary_json["ingestion_debug"][0]["selected_sheet"], "Measurements")
        self.assertEqual(campaign.measurements.count(), 2)

    def test_no_measurements_report_shows_file_failure_reason(self):
        campaign = Campaign.objects.create(name="Bad Aranet XLSX")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"
        worksheet.append(["Aranet metadata only"])
        worksheet.append(["No table here"])
        buffer = BytesIO()
        workbook.save(buffer)
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="bad-aranet.xlsx",
            file=SimpleUploadedFile(
                "bad-aranet.xlsx",
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.summary_json["measurement_count"], 0)
        self.assertIn("no measurements were imported", report.summary.lower())
        self.assertEqual(report.summary_json["ingestion_debug"][0]["filename"], "bad-aranet.xlsx")
        self.assertIn("Could not find a header row", report.summary_json["ingestion_debug"][0]["skipped_reason"])
        self.assertIn("Ingestion Diagnostics", report.html_report)


class RegimeClassificationTests(TestCase):
    def test_classify_regimes_uses_one_hour_radon_change(self):
        start = timezone.datetime(2026, 3, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        values = [
            Decimal("80"),
            Decimal("90"),
            Decimal("130"),
            Decimal("260"),
            Decimal("310"),
            Decimal("170"),
            Decimal("130"),
            Decimal("150"),
        ]
        rows = [
            {
                "measured_at": start + timedelta(hours=index),
                "radon_bq_m3": value,
                "temperature_c": None,
                "humidity_percent": None,
                "pressure_hpa": None,
                "segment_id": 1,
            }
            for index, value in enumerate(values)
        ]

        classified = classify_regimes(rows)

        self.assertEqual(
            [row["regime"] for row in classified],
            [
                "stable_low",
                "stable_low",
                "rising",
                "sudden_rise",
                "high_episode",
                "sudden_drop",
                "falling",
                "stable_elevated",
            ],
        )

    def test_pipeline_stores_regimes_and_report_statistics(self):
        campaign = Campaign.objects.create(name="Regime Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="regimes.csv",
            file=SimpleUploadedFile(
                "regimes.csv",
                (
                    "Time,Radon,Temperature,Humidity,Pressure\n"
                    "2026-03-01 00:00,80,20,45,1000\n"
                    "2026-03-01 01:00,90,20,45,1000\n"
                    "2026-03-01 02:00,130,20,45,1000\n"
                    "2026-03-01 03:00,260,20,45,1000\n"
                    "2026-03-01 04:00,310,20,45,1000\n"
                    "2026-03-01 05:00,170,20,45,1000\n"
                    "2026-03-01 06:00,130,20,45,1000\n"
                    "2026-03-01 07:00,150,20,45,1000\n"
                ).encode("utf-8"),
                content_type="text/csv",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.summary_json["concentration_level_counts"]["LOW"], 2)
        self.assertGreaterEqual(report.summary_json["confirmed_dynamic_state_counts"]["SUDDEN_RISE"], 1)
        self.assertGreaterEqual(report.summary_json["candidate_dynamic_state_counts"]["RISING"], 1)
        self.assertIn("Per-Measurement Regime Counts", report.html_report)
        self.assertIn("sudden_rise", report.html_report)
        self.assertEqual(
            list(campaign.measurements.order_by("measured_at").values_list("regime", flat=True)),
            [
                "quality_affected",
                "rising",
                "rising",
                "sudden_rise",
                "rising",
                "sudden_drop",
                "unstable_transition",
                "unstable_transition",
            ],
        )

    def test_campaign_detail_exposes_regime_statistics(self):
        campaign = Campaign.objects.create(name="Visible Regimes")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Analysis complete.",
            summary_json={"regime_counts": {"stable_low": 1}, "segments": []},
            html_report="<article><h1>Regime Statistics</h1></article>",
        )

        response = self.client.get(reverse("campaigns:campaign_regimes", args=[campaign.pk]))

        self.assertContains(response, "Regime Summary")
        self.assertContains(response, "stable_low")
        self.assertNotContains(response, "Generated HTML Report")


class SegmentInterpretationTests(TestCase):
    def test_segment_interpretation_labels_exposure_and_dynamics(self):
        start = timezone.datetime(2026, 6, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = []
        rows.extend(_segment_rows(start, 1, [70, 75, 80, 78], ["stable_low"] * 4))
        rows.extend(_segment_rows(start, 2, [70, 82, 95, 105, 90], ["stable_low", "rising", "falling", "stable_elevated", "stable_low"]))
        rows.extend(_segment_rows(start, 3, [105, 125, 135, 145], ["stable_elevated", "rising", "stable_elevated", "falling"]))
        rows.extend(_segment_rows(start, 4, [120, 220, 330, 180], ["stable_elevated", "rising", "high_episode", "falling"]))
        rows.extend(_segment_rows(start, 5, [90, 95], ["stable_low", "stable_low"]))

        summary = build_summary(rows, gaps=[], uploaded_file_count=1)
        segments = {segment["segment_id"]: segment for segment in summary["segments"]}

        self.assertEqual(segments[1]["segment_label"], "low_stable")
        self.assertEqual(segments[1]["percent_above_100"], 0.0)
        self.assertEqual(segments[1]["dynamic_percent"], 0.0)
        self.assertEqual(segments[2]["segment_label"], "low_dynamic")
        self.assertEqual(segments[2]["percent_above_100"], 20.0)
        self.assertEqual(segments[2]["dynamic_percent"], 40.0)
        self.assertEqual(segments[3]["segment_label"], "elevated_dynamic")
        self.assertEqual(segments[3]["percent_above_100"], 100.0)
        self.assertEqual(segments[4]["segment_label"], "high_episode")
        self.assertEqual(segments[4]["percent_above_200"], 50.0)
        self.assertEqual(segments[5]["segment_label"], "insufficient_data")
        self.assertIn("Fewer than three", segments[5]["interpretation_text"])

    def test_pipeline_report_includes_segment_interpretation(self):
        campaign = Campaign.objects.create(name="Segment Interpretation Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="segments.csv",
            file=SimpleUploadedFile(
                "segments.csv",
                (
                    "Time,Radon,Temperature,Humidity,Pressure\n"
                    "2026-06-01 00:00,90,20,45,1000\n"
                    "2026-06-01 01:00,115,20,45,1000\n"
                    "2026-06-01 02:00,140,20,45,1000\n"
                    "2026-06-01 03:00,165,20,45,1000\n"
                ).encode("utf-8"),
                content_type="text/csv",
            ),
        )

        report = run_campaign_analysis(campaign)
        segment = report.summary_json["segments"][0]

        self.assertEqual(segment["segment_label"], "elevated_dynamic")
        self.assertIn("percent_above_100", segment)
        self.assertIn("dynamic_percent", segment)
        self.assertIn("Segment Interpretation", report.html_report)
        self.assertIn("naive baseline", report.html_report)


class PredictionModelTests(TestCase):
    def test_prediction_models_compute_metrics_for_1h_and_6h_horizons(self):
        start = timezone.datetime(2026, 4, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {
                "measured_at": start + timedelta(hours=index),
                "radon_bq_m3": Decimal(str(100 + index * 10)),
                "segment_id": 1,
            }
            for index in range(10)
        ]

        evaluation = evaluate_prediction_models(rows)
        metrics = evaluation["overall"]

        self.assertEqual(metrics["1h"]["naive_baseline"]["samples"], 7)
        self.assertEqual(metrics["1h"]["naive_baseline"]["mae"], 10.0)
        self.assertEqual(metrics["1h"]["ridge"]["samples"], 7)
        self.assertLess(metrics["1h"]["ridge"]["mae"], metrics["1h"]["naive_baseline"]["mae"])
        self.assertEqual(metrics["6h"]["naive_baseline"]["samples"], 2)
        self.assertEqual(metrics["6h"]["naive_baseline"]["mae"], 60.0)
        self.assertTrue(evaluation["by_regime"])
        self.assertTrue(evaluation["errors"])

    def test_prediction_samples_do_not_cross_segment_boundaries(self):
        start = timezone.datetime(2026, 4, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {
                "measured_at": start + timedelta(hours=index),
                "radon_bq_m3": Decimal(str(100 + index * 5)),
                "segment_id": 1,
            }
            for index in range(3)
        ]
        rows.extend(
            {
                "measured_at": start + timedelta(hours=10 + index),
                "radon_bq_m3": Decimal(str(200 + index * 5)),
                "segment_id": 2,
            }
            for index in range(3)
        )

        metrics = evaluate_prediction_models(rows)["overall"]

        self.assertEqual(metrics["1h"]["naive_baseline"]["samples"], 0)
        self.assertEqual(metrics["6h"]["naive_baseline"]["samples"], 0)

    def test_pipeline_stores_prediction_metrics_and_report_output(self):
        campaign = Campaign.objects.create(name="Prediction Campaign")
        csv_lines = ["Time,Radon,Temperature,Humidity,Pressure"]
        for index in range(10):
            csv_lines.append(
                f"2026-04-01 {index:02d}:00,{100 + index * 10},20,45,1000"
            )
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="prediction.csv",
            file=SimpleUploadedFile(
                "prediction.csv",
                "\n".join(csv_lines).encode("utf-8"),
                content_type="text/csv",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertEqual(report.summary_json["prediction_metrics"]["1h"]["naive_baseline"]["samples"], 7)
        self.assertEqual(report.summary_json["prediction_metrics"]["6h"]["naive_baseline"]["samples"], 2)
        self.assertIn("prediction_metrics_by_regime", report.summary_json)
        self.assertIn("prediction_errors", report.summary_json)
        self.assertTrue(report.summary_json["prediction_metrics_by_regime"])
        self.assertTrue(report.summary_json["prediction_errors"])
        self.assertIn("Model Performance", report.html_report)
        self.assertIn("ridge", report.html_report)

    def test_prediction_evaluation_handles_missing_regime_labels(self):
        start = timezone.datetime(2026, 4, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {
                "measured_at": start + timedelta(hours=index),
                "radon_bq_m3": Decimal(str(100 + index * 10)),
                "segment_id": 1,
            }
            for index in range(5)
        ]

        evaluation = evaluate_prediction_models(rows)

        regimes = {row["regime"] for row in evaluation["by_regime"]}
        self.assertEqual(regimes, {"unclassified"})
        self.assertEqual(evaluation["errors"][0]["regime"], "unclassified")

    def test_campaign_detail_exposes_model_performance(self):
        campaign = Campaign.objects.create(name="Visible Models")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Analysis complete.",
            summary_json={
                "regime_counts": {},
                "prediction_metrics": {
                    "1h": {
                        "naive_baseline": {"samples": 3, "mae": 5.0, "rmse": 6.0},
                        "ridge": {"samples": 3, "mae": 2.0, "rmse": 3.0},
                    }
                },
                "segments": [],
            },
            html_report="<article><h1>Model Performance</h1></article>",
        )

        response = self.client.get(reverse("campaigns:campaign_prediction", args=[campaign.pk]))

        self.assertContains(response, "Overall Metrics")
        self.assertContains(response, "naive_baseline")
        self.assertContains(response, "ridge")


class ExcelExportTests(TestCase):
    def test_excel_report_export_returns_workbook(self):
        campaign = Campaign.objects.create(name="Export Campaign", location="Lab A")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="export.csv",
            file=SimpleUploadedFile("export.csv", b"Time,Radon\n2026-01-01 00:00,100\n"),
        )
        CampaignResearchContext.objects.create(
            campaign=campaign,
            floor_level=1,
            room_volume_m3=Decimal("42.500"),
            dominant_material="concrete",
            evidence_status="PROVIDED",
        )
        report = AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Research prototype analysis complete.",
            summary_json={
                "measurement_count": 2,
                "segment_count": 1,
                "gap_count": 1,
                "regime_counts": {"stable_low": 1, "rising": 1},
                "prediction_metrics": {
                    "1h": {
                        "naive_baseline": {"samples": 1, "mae": 10.0, "rmse": 10.0},
                        "ridge": {"samples": 1, "mae": 5.0, "rmse": 6.0},
                    }
                },
                "prediction_metrics_by_regime": [
                    {
                        "horizon": "1h",
                        "model": "ridge",
                        "regime": "rising",
                        "samples": 1,
                        "mae": 5.0,
                        "rmse": 6.0,
                        "mae_improvement_percent": 50.0,
                        "rmse_improvement_percent": 40.0,
                    }
                ],
                "prediction_errors": [
                    {
                        "timestamp": "2026-01-01T01:00:00+00:00",
                        "horizon": "1h",
                        "model": "ridge",
                        "actual_radon": 110,
                        "predicted_radon": 105,
                        "absolute_error": 5,
                        "regime": "rising",
                        "segment_id": 1,
                    }
                ],
                "gaps": [{"from": "2026-01-01T00:00:00+00:00", "to": "2026-01-01T02:00:00+00:00", "minutes": 120}],
                "segments": [
                    {
                        "segment_id": 1,
                        "start": "2026-01-01T00:00:00+00:00",
                        "end": "2026-01-01T01:00:00+00:00",
                        "segment_label": "low_dynamic",
                        "dominant_regime": "rising",
                        "interpretation_text": "Low but dynamic.",
                        "statistics": {"radon_bq_m3": {"mean": 105.0, "max": 110.0}},
                    }
                ],
                "ingestion_debug": [
                    {
                        "filename": "export.csv",
                        "parsed_measurement_rows": 2,
                        "skipped_rows": 0,
                        "skipped_reason": "",
                        "detected_sheets": ["CSV"],
                        "detected_header_row": 1,
                        "mapped_columns": {"timestamp": "Time", "radon": "Radon"},
                    }
                ],
            },
        )
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        Measurement.objects.create(
            campaign=campaign,
            measured_at=start,
            radon_bq_m3=Decimal("100"),
            temperature_c=Decimal("20.1"),
            humidity_percent=Decimal("45.0"),
            pressure_hpa=Decimal("1001.0"),
            segment_id=1,
            regime="stable_low",
        )
        Measurement.objects.create(
            campaign=campaign,
            measured_at=start + timedelta(hours=1),
            radon_bq_m3=Decimal("110"),
            segment_id=1,
            regime="rising",
        )

        response = self.client.get(reverse("campaigns:export_excel_report", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(f"radon_campaign_{campaign.pk}_report.xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary",
                "Segments",
                "Regime Counts",
                "Prediction Metrics",
                "Prediction Insights",
                "Prediction by Regime",
                "Prediction Errors",
                "Gaps",
                "Ingestion Diagnostics",
                "Measurements",
                "Source File Inventory",
                "Canonical Dataset Summary",
                "Canonical Hourly Data",
                "Quality Flags",
                "Quality Flag Dictionary",
                "Sampling Diagnostics",
                "Overlap Conflicts",
                "DST Diagnostics",
                "Resampling Summary",
                "Regime Sensitivity",
                "Prediction Skill by Regime",
                "Prediction Readiness",
                "SIREM Readiness",
                "Reproducibility Config",
                "Row Reconciliation Summary",
                "Campaign Summary",
                "Data Quality Summary",
                "Intervals and Gaps",
                "Measurement Regimes",
                "Episodes",
                "Regime Parameters",
                "Regime Confidence",
                "Important Episodes",
                "Feature Diagnostics",
                "Sudden Event Audit",
                "Episode Reasons",
                "Elevated Period Phases",
                "Profile Applicability",
                "Adaptive Recommendations",
                "Standardized Summary",
                "Transition Merge Audit",
                "Level Sensitivity",
                "Dynamic Sensitivity",
                "Prediction Summary",
                "Prediction Intervals",
                "Largest Errors",
                "Methodology Metadata",
                "Research Context",
            ],
        )
        self.assertEqual(workbook["Summary"]["B2"].value, "Export Campaign")
        self.assertEqual(workbook["Summary"]["B3"].value, "Lab A")
        self.assertEqual(workbook["Summary"]["B6"].value, 2)
        self.assertEqual(workbook["Summary"]["B7"].value, 1)
        self.assertEqual(workbook["Segments"]["A2"].value, 1)
        self.assertEqual(workbook["Segments"]["G2"].value, "low_dynamic")
        self.assertEqual(workbook["Regime Counts"]["A2"].value, "stable_low")
        self.assertEqual(workbook["Prediction Metrics"]["A2"].value, "1h")
        self.assertEqual(workbook["Prediction Insights"]["A1"].value, "Prediction Insights")
        self.assertEqual(workbook["Prediction by Regime"]["C2"].value, "rising")
        self.assertEqual(workbook["Prediction Errors"]["G2"].value, "rising")
        self.assertEqual(workbook["Row Reconciliation Summary"]["A1"].value, "Field")
        self.assertEqual(workbook["Methodology Metadata"]["A1"].value, "Field")
        self.assertEqual(workbook["Research Context"]["A1"].value, "Section")
        self.assertEqual(workbook["Research Context"]["C2"].value, 1)
        self.assertEqual(workbook["Research Context"]["D2"].value, "PROVIDED")
        self.assertEqual(workbook["Gaps"]["C2"].value, 120)
        self.assertEqual(workbook["Ingestion Diagnostics"]["A2"].value, "export.csv")
        self.assertEqual(workbook["Measurements"]["B2"].value, 100.0)
        for sheet_name in workbook.sheetnames:
            self.assertEqual(workbook[sheet_name].freeze_panes, "A2")
            self.assertTrue(workbook[sheet_name].auto_filter.ref)
            self.assertTrue(workbook[sheet_name]["A1"].font.bold)
            self.assertEqual(workbook[sheet_name]["A1"].alignment.vertical, "center")
        self.assertEqual(workbook["Measurements"]["A2"].number_format, "yyyy-mm-dd hh:mm")
        self.assertEqual(workbook["Measurements"]["B2"].number_format, "0.0")
        self.assertEqual(workbook["Measurements"]["C2"].number_format, "0.0")
        self.assertEqual(workbook["Prediction Metrics"]["D2"].number_format, "0.000")
        self.assertTrue(workbook["Ingestion Diagnostics"]["H2"].alignment.wrap_text)
        self.assertEqual(report, campaign.analysis_reports.first())

    def test_excel_report_export_handles_missing_optional_fields(self):
        campaign = Campaign.objects.create(name="Sparse Export")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Sparse report.",
            summary_json={"segments": [{}], "prediction_metrics": {"6h": {"ridge": {"samples": 0}}}},
        )

        response = self.client.get(reverse("campaigns:export_excel_report", args=[campaign.pk]))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook["Summary"]["B2"].value, "Sparse Export")
        self.assertEqual(workbook["Summary"]["B3"].value, "N/A")
        self.assertEqual(workbook["Segments"]["A2"].value, "N/A")
        self.assertEqual(workbook["Prediction Metrics"]["A2"].value, "6h")
        self.assertIn("Research Context", workbook.sheetnames)
        self.assertIsNone(workbook["Research Context"]["C2"].value)

    def test_compact_excel_report_export_omits_full_measurement_sheets(self):
        campaign = Campaign.objects.create(name="Compact Export")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="Report ready.",
            summary_json={"regime_counts": {"stable_low": 1}, "segments": []},
        )

        response = self.client.get(reverse("campaigns:export_excel_report", args=[campaign.pk]), {"mode": "compact"})

        self.assertEqual(response.status_code, 200)
        self.assertIn(f"radon_campaign_{campaign.pk}_compact_report.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Summary", workbook.sheetnames)
        self.assertIn("Regime Counts", workbook.sheetnames)
        self.assertIn("Research Context", workbook.sheetnames)
        self.assertNotIn("Measurements", workbook.sheetnames)
        self.assertNotIn("Canonical Hourly Data", workbook.sheetnames)


class PaperOneResearchWorkflowTests(TestCase):
    def test_source_inventory_reports_start_end_and_row_counts(self):
        campaign = Campaign.objects.create(name="Inventory Campaign")
        uploaded = UploadedFile.objects.create(
            campaign=campaign,
            original_name="AranetRn+ 2E81E_test.csv",
            file=SimpleUploadedFile("inventory.csv", b"Time,Radon\n"),
        )
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {
                "source_file": uploaded,
                "measured_at": start + timedelta(minutes=10 * index),
                "radon_bq_m3": Decimal("100"),
                "temperature_c": Decimal("20"),
                "humidity_percent": None,
                "pressure_hpa": None,
            }
            for index in range(3)
        ]
        inventory = build_source_file_inventory(
            rows,
            [{"source_file_id": uploaded.id, "filename": uploaded.original_name, "raw_rows_read": 4, "detected_columns": ["Time", "Radon"], "parsed_measurement_rows": 3}],
        )

        self.assertEqual(inventory[0]["device_id"], "2E81E")
        self.assertEqual(inventory[0]["raw_rows"], 4)
        self.assertEqual(inventory[0]["imported_measurement_rows"], 3)
        self.assertEqual(inventory[0]["nominal_sampling_interval_minutes"], 10.0)

    def test_canonicalization_deduplicates_exact_rows_and_preserves_provenance(self):
        campaign = Campaign.objects.create(name="Canonical Campaign")
        first = UploadedFile.objects.create(campaign=campaign, original_name="a.csv", file=SimpleUploadedFile("a.csv", b""))
        second = UploadedFile.objects.create(campaign=campaign, original_name="b.csv", file=SimpleUploadedFile("b.csv", b""))
        timestamp = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"source_file": first, "measured_at": timestamp, "radon_bq_m3": Decimal("100"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
            {"source_file": second, "measured_at": timestamp, "radon_bq_m3": Decimal("100"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
        ]

        outputs = build_canonical_outputs(rows, AnalysisConfig())

        self.assertEqual(outputs["canonical_dataset_summary"]["exact_duplicates_removed"], 1)
        self.assertEqual(outputs["canonical_records_preview"][0]["source_count"], 2)

    def test_canonicalization_flags_duplicate_conflicts(self):
        campaign = Campaign.objects.create(name="Conflict Campaign")
        first = UploadedFile.objects.create(campaign=campaign, original_name="a.csv", file=SimpleUploadedFile("a.csv", b""))
        second = UploadedFile.objects.create(campaign=campaign, original_name="b.csv", file=SimpleUploadedFile("b.csv", b""))
        timestamp = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"source_file": first, "measured_at": timestamp, "radon_bq_m3": Decimal("100"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
            {"source_file": second, "measured_at": timestamp, "radon_bq_m3": Decimal("130"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
        ]

        outputs = build_canonical_outputs(rows, AnalysisConfig())

        self.assertEqual(outputs["canonical_dataset_summary"]["conflicts"], 1)
        self.assertIn("DUPLICATE_CONFLICT", outputs["canonical_records_preview"][0]["quality_flags"])

    def test_sampling_aware_gaps_handle_10_minute_and_60_minute_data(self):
        config = AnalysisConfig(gap_tolerance_multiplier=1.5)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        regular_10 = [{"measured_at": start + timedelta(minutes=10 * index)} for index in range(6)]
        regular_60 = [{"measured_at": start + timedelta(hours=index)} for index in range(6)]
        missing_10 = [{"measured_at": start}, {"measured_at": start + timedelta(minutes=10)}, {"measured_at": start + timedelta(minutes=40)}]

        self.assertEqual(detect_sampling_gaps(regular_10, config), [])
        self.assertEqual(detect_sampling_gaps(regular_60, config), [])
        self.assertEqual(detect_sampling_gaps(missing_10, config)[0]["gap_class"], "GAP_SHORT")

    def test_dst_ambiguity_is_flagged_for_autumn_fallback(self):
        config = AnalysisConfig(timezone_name="Europe/Rome")
        ambiguous = timezone.datetime(2026, 10, 25, 2, 30, tzinfo=ZoneInfo("Europe/Rome"))
        rows = [{"measured_at": ambiguous, "radon_bq_m3": Decimal("100")}]

        diagnostics = build_dst_diagnostics(rows, config)

        self.assertIn("DST_AMBIGUOUS", diagnostics[0]["flags"])

    def test_hourly_resampling_counts_completeness_and_flags_low_completeness(self):
        config = AnalysisConfig(completeness_threshold=0.75)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start, "radon_bq_m3": Decimal("100"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
            {"measured_at": start + timedelta(minutes=10), "radon_bq_m3": Decimal("110"), "temperature_c": None, "humidity_percent": None, "pressure_hpa": None},
        ]

        outputs = build_hourly_resampling(rows, config)

        self.assertEqual(outputs["canonical_hourly_data"][0]["radon_count"], 2)
        self.assertIn("LOW_COMPLETENESS", outputs["canonical_hourly_data"][0]["quality_flags"])

    def test_regime_threshold_sensitivity_returns_agreement_and_counts(self):
        config = AnalysisConfig()
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(value))}
            for index, value in enumerate([80, 110, 320])
        ]

        sensitivity = build_regime_sensitivity(rows, config)

        self.assertEqual(len(sensitivity), 5)
        self.assertIn("regime_counts", sensitivity[0])
        self.assertIn("percentage_agreement_with_baseline", sensitivity[0])

    def test_pipeline_summary_contains_paper_one_outputs(self):
        campaign = Campaign.objects.create(name="Paper One Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="paper.csv",
            file=SimpleUploadedFile(
                "paper.csv",
                b"Time,Radon,Temperature,Humidity,Pressure\n2026-01-01 00:00,100,20,45,1000\n2026-01-01 01:00,110,20,45,1000\n2026-01-01 02:00,120,20,45,1000\n",
            ),
        )

        report = run_campaign_analysis(campaign)

        self.assertIn("source_file_inventory", report.summary_json)
        self.assertIn("canonical_dataset_summary", report.summary_json)
        self.assertIn("quality_flag_counts", report.summary_json)
        self.assertIn("sampling_diagnostics", report.summary_json)
        self.assertIn("sirem_readiness", report.summary_json)

    def test_analyze_campaign_command_creates_paper_outputs(self):
        campaign = Campaign.objects.create(name="Command Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="command.csv",
            file=SimpleUploadedFile(
                "command.csv",
                (
                    "Time,Radon,Temperature,Humidity,Pressure\n"
                    "2026-01-01 00:00,100,20,45,1000\n"
                    "2026-01-01 01:00,110,20,45,1000\n"
                    "2026-01-01 02:00,120,20,45,1000\n"
                ).encode("utf-8"),
            ),
        )

        with TemporaryDirectory() as tempdir:
            call_command(
                "analyze_campaign",
                campaign.id,
                "--timezone",
                "Europe/Rome",
                "--resample",
                "1H",
                "--gap-tolerance",
                "1.5",
                "--rebuild-canonical",
                "--run-sensitivity",
                "--export-excel",
                "--output-dir",
                tempdir,
                verbosity=0,
            )
            output_dir = Path(tempdir)

            self.assertTrue((output_dir / f"radon_campaign_{campaign.id}_report.xlsx").exists())
            self.assertTrue((output_dir / "row_reconciliation_summary.csv").exists())
            self.assertTrue((output_dir / "dst_diagnostics_compact_summary.csv").exists())
            self.assertTrue((output_dir / "sampling_gaps_compact_summary.csv").exists())
            self.assertTrue((output_dir / "paper1_validation_report.md").exists())


class RegimeAnalysisV2Tests(TestCase):
    def test_time_continuity_classifies_regular_and_gap_intervals(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start, "radon_bq_m3": Decimal("100")},
            {"measured_at": start + timedelta(minutes=60), "radon_bq_m3": Decimal("105")},
            {"measured_at": start + timedelta(minutes=180), "radon_bq_m3": Decimal("120")},
            {"measured_at": start + timedelta(minutes=420), "radon_bq_m3": Decimal("130")},
            {"measured_at": start + timedelta(minutes=1020), "radon_bq_m3": Decimal("140")},
            {"measured_at": start + timedelta(minutes=1860), "radon_bq_m3": Decimal("150")},
        ]

        result = analyze_time_continuity(rows, config)
        classes = [item["interval_class"] for item in result["intervals"]]

        self.assertEqual(classes[0], "REGULAR_INTERVAL")
        self.assertEqual(classes[1], "SHORT_GAP")
        self.assertEqual(classes[2], "MODERATE_GAP")
        self.assertEqual(classes[3], "LONG_GAP")
        self.assertEqual(classes[4], "LONG_GAP")
        self.assertEqual(result["summary"]["segment_count"], 5)

    def test_time_continuity_handles_duplicate_and_out_of_order_timestamps(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=1), "radon_bq_m3": Decimal("110")},
            {"measured_at": start, "radon_bq_m3": Decimal("100")},
            {"measured_at": start, "radon_bq_m3": Decimal("100")},
        ]

        result = analyze_time_continuity(rows, config)

        self.assertEqual(result["summary"]["out_of_order_timestamp_count"], 1)
        self.assertEqual(result["summary"]["duplicated_timestamp_count"], 1)

    def test_regime_v2_separates_level_and_dynamic_state_without_crossing_gap(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start, "radon_bq_m3": Decimal("80")},
            {"measured_at": start + timedelta(hours=1), "radon_bq_m3": Decimal("120")},
            {"measured_at": start + timedelta(hours=5), "radon_bq_m3": Decimal("260")},
            {"measured_at": start + timedelta(hours=6), "radon_bq_m3": Decimal("180")},
        ]
        continuity = analyze_time_continuity(rows, config)

        classified = classify_regimes_v2(continuity["rows"], config)

        self.assertEqual(classified[0]["concentration_level"], "LOW")
        self.assertEqual(classified[1]["concentration_level"], "ELEVATED")
        self.assertEqual(classified[2]["dynamic_state"], "QUALITY_AFFECTED")
        self.assertNotEqual(classified[2].get("slope_bq_m3_per_hour"), 140)

    def test_regime_v2_detects_sudden_rise_drop_and_confidence_reasons(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(value))}
            for index, value in enumerate([80, 90, 220, 100])
        ]
        continuity = analyze_time_continuity(rows, config)

        classified = classify_regimes_v2(continuity["rows"], config)
        states = [row["dynamic_state"] for row in classified]

        self.assertIn("SUDDEN_RISE", states)
        self.assertIn("SUDDEN_DROP", states)
        self.assertTrue(classified[-1]["regime_confidence_reasons"])

    def test_episode_grouping_and_statistics(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(value))}
            for index, value in enumerate([80, 95, 120, 150, 180])
        ]
        continuity = analyze_time_continuity(rows, config)
        classified = classify_regimes_v2(continuity["rows"], config)

        episodes = build_episodes(classified, continuity["gaps"], config)

        self.assertTrue(episodes)
        self.assertIn("episode_type", episodes[0])
        self.assertIn("mean_radon", episodes[0])

    def test_regime_v2_candidate_confirmed_persistence_and_hysteresis(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(value))}
            for index, value in enumerate([95, 105, 95, 92, 88, 91, 89])
        ]
        continuity = analyze_time_continuity(rows, config)

        classified = classify_regimes_v2(continuity["rows"], config)

        self.assertIn("candidate_dynamic_state", classified[2])
        self.assertIn("confirmed_dynamic_state", classified[2])
        self.assertEqual(classified[1]["concentration_level"], "ELEVATED")
        self.assertEqual(classified[2]["concentration_level"], "ELEVATED")
        self.assertEqual(classified[4]["concentration_level"], "LOW")
        self.assertTrue(any(row["candidate_dynamic_state"] != row["confirmed_dynamic_state"] for row in classified))

    def test_regime_v2_classifies_low_rising_elevated_stable_high_falling(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60, minimum_state_persistence_observations=1)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        values = [
            (0, 40), (1, 50), (2, 80),
            (8, 120), (9, 122), (10, 124), (11, 126), (12, 128), (13, 130), (14, 132),
            (20, 280), (21, 260), (22, 240), (23, 220), (24, 210),
        ]
        rows = [{"measured_at": start + timedelta(hours=hour), "radon_bq_m3": Decimal(str(value))} for hour, value in values]
        continuity = analyze_time_continuity(rows, config)

        classified = classify_regimes_v2(continuity["rows"], config)

        self.assertTrue(any(row["concentration_level"] == "LOW" and row["confirmed_dynamic_state"] == "RISING" for row in classified))
        self.assertTrue(any(row["concentration_level"] == "ELEVATED" and row["confirmed_dynamic_state"] == "STABLE" for row in classified))
        self.assertTrue(any(row["concentration_level"] == "HIGH" and row["confirmed_dynamic_state"] in {"FALLING", "SUDDEN_DROP"} for row in classified))

    def test_episodes_end_at_gaps_and_include_v2_statistics(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start, "radon_bq_m3": Decimal("80")},
            {"measured_at": start + timedelta(hours=1), "radon_bq_m3": Decimal("120")},
            {"measured_at": start + timedelta(hours=5), "radon_bq_m3": Decimal("125")},
            {"measured_at": start + timedelta(hours=6), "radon_bq_m3": Decimal("126")},
        ]
        continuity = analyze_time_continuity(rows, config)
        classified = classify_regimes_v2(continuity["rows"], config)

        episodes = build_episodes(classified, continuity["gaps"], config, campaign_id=7, analysis_report_id=11)

        self.assertGreaterEqual(len(episodes), 2)
        self.assertTrue(all(episode["campaign_id"] == 7 for episode in episodes))
        self.assertTrue(all(episode["analysis_report_id"] == 11 for episode in episodes))
        self.assertTrue(all("confidence_reason_codes" in episode for episode in episodes))
        self.assertTrue(any(episode["quality_status"] == "QUALITY_AFFECTED" for episode in episodes))

    def test_dynamic_sensitivity_reports_agreement_and_episode_preservation(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(70 + index * 8))}
            for index in range(10)
        ]
        continuity = analyze_time_continuity(rows, config)
        classified = classify_regimes_v2(continuity["rows"], config)

        sensitivity = build_sensitivity_v2(classified, config)

        self.assertTrue(sensitivity["dynamic_sensitivity"])
        self.assertIn("cohen_kappa", sensitivity["dynamic_sensitivity"][0])
        self.assertIn("baseline_episode_preservation_percent", sensitivity["dynamic_sensitivity"][0])
        self.assertIn("episode_count_by_type", sensitivity["dynamic_sensitivity"][0])

    def test_v2_sensitivity_and_prediction_intervals(self):
        config = AnalysisConfig(expected_sampling_interval_minutes=60)
        start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
        rows = [
            {"measured_at": start + timedelta(hours=index), "radon_bq_m3": Decimal(str(80 + index * 10))}
            for index in range(12)
        ]
        continuity = analyze_time_continuity(rows, config)
        classified = classify_regimes_v2(continuity["rows"], config)

        sensitivity = build_sensitivity_v2(classified, config)
        prediction = evaluate_prediction_v2(classified, config)

        self.assertIn("level_sensitivity", sensitivity)
        self.assertIn("dynamic_sensitivity", sensitivity)
        self.assertIn("prediction_intervals", prediction)
        self.assertTrue(prediction["prediction_intervals"])


class RegimePortabilityTests(TestCase):
    def test_same_physical_rise_is_consistent_across_sampling_intervals(self):
        for interval_minutes in (10, 30, 60, 120):
            config = build_config(
                "default_radon_hourly",
                overrides={"expected_sampling_interval_minutes": interval_minutes},
            ).with_time_windows(interval_minutes)
            start = timezone.datetime(2026, 1, 1, 0, 0, tzinfo=timezone.get_current_timezone())
            rows = [
                {
                    "measured_at": start + timedelta(minutes=interval_minutes * index),
                    "radon_bq_m3": Decimal(str(50 + (12 * interval_minutes / 60) * index)),
                }
                for index in range(max(6, int(360 / interval_minutes) + 1))
            ]
            continuity = analyze_time_continuity(rows, config)
            classified = classify_regimes_v2(continuity["rows"], config)
            states = [row["confirmed_dynamic_state"] for row in classified]

            self.assertIn("RISING", states, f"interval {interval_minutes} min should detect the same per-hour rise")
            self.assertTrue(all(row.get("adjacent_slope_bq_m3_per_hour") in (None, 12.0) for row in classified))

    def test_profile_override_and_applicability_are_exported(self):
        config = build_config(
            "high_noise_sensor",
            overrides=parse_overrides(["trend_slope_bq_m3_per_hour=12", "expected_sampling_interval_minutes=30"]),
        )
        campaign = Campaign.objects.create(name="Profile Campaign")
        UploadedFile.objects.create(
            campaign=campaign,
            original_name="profile.csv",
            file=SimpleUploadedFile(
                "profile.csv",
                (
                    "Time,Radon\n"
                    "2026-01-01 00:00,50\n"
                    "2026-01-01 00:30,56\n"
                    "2026-01-01 01:00,62\n"
                    "2026-01-01 01:30,68\n"
                ).encode("utf-8"),
            ),
        )

        report = run_campaign_analysis(campaign, config=config)

        self.assertEqual(report.summary_json["profile_metadata"]["profile_name"], "high_noise_sensor")
        self.assertEqual(report.summary_json["profile_metadata"]["overrides"]["trend_slope_bq_m3_per_hour"], 12)
        self.assertIn("profile_applicability", report.summary_json)
        self.assertIn("adaptive_recommendations", report.summary_json)
        self.assertIn("standardized_campaign_summary", report.summary_json)

    def test_compare_campaigns_reads_completed_reports_without_rerun(self):
        campaign = Campaign.objects.create(name="Compare Campaign")
        AnalysisReport.objects.create(
            campaign=campaign,
            status=AnalysisReport.Status.COMPLETE,
            summary="done",
            summary_json={
                "standardized_campaign_summary": {
                    "profile_name": "default_radon_hourly",
                    "profile_version": "2026-07-v1",
                    "valid_row_count": 10,
                },
                "profile_applicability": {"status": "PROFILE_COMPATIBLE", "reason_codes": []},
            },
        )

        rows = compare_campaigns([campaign.id], profile="default_radon_hourly")

        self.assertEqual(rows[0]["campaign_id"], campaign.id)
        self.assertEqual(rows[0]["valid_row_count"], 10)
        self.assertEqual(rows[0]["profile_compatibility"], "PROFILE_COMPATIBLE")


def _segment_rows(start, segment_id, radon_values, regimes):
    return [
        {
            "measured_at": start + timedelta(hours=segment_id * 24 + index),
            "radon_bq_m3": Decimal(str(value)),
            "temperature_c": None,
            "humidity_percent": None,
            "pressure_hpa": None,
            "segment_id": segment_id,
            "regime": regime,
        }
        for index, (value, regime) in enumerate(zip(radon_values, regimes))
    ]
